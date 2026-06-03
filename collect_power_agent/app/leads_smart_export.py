"""leads_smart_export.py — Smart tiered export of leads (web agencies) by reseller potential.

Reads leads + contacts from Firestore, keeps only leads with valid email contacts,
scores and groups into 5 tiers based on reseller fit for BlueSearch (selling SEO
services through agency resellers).

Tier 1 — Prime Reseller  : ai_reseller_potential=high  + score >= 75 (A priority)
Tier 2 — Strong Prospect : ai_reseller_potential=high  + score 55-74  OR  WordPress+SMB
Tier 3 — Good Prospect   : ai_reseller_potential=medium + score >= 55
Tier 4 — Possible        : ai_reseller_potential=medium + score < 55
Tier 5 — Low Priority    : ai_reseller_potential=low   or no AI classification

Bonus signals that lift a tier:
  - WordPress/WooCommerce specialisation → +1 tier
  - SMB client base → +1 tier
  - care_plan / hosting in specialisation → +1 tier

Usage:
    python app\\leads_smart_export.py --countries UK
    python app\\leads_smart_export.py --countries UK IN NO --out exports\\leads_resellers.xlsx
    python app\\leads_smart_export.py --countries NO SE DK --min-score 50
"""
from __future__ import annotations

import threading as _threading

# Guards firebase_admin.initialize_app against concurrent init
_local_fb_lock = _threading.Lock()
import sys
import argparse
import importlib.util
import os
import re
from pathlib import Path as _Path
sys.path.insert(0, str(_Path(__file__).parent))  # make functions/ importable
from functions.utils import clean_str, resolve_country, ISO_TO_CC, email_matches_name, normalize_url
from functions.excel_builder import write_contacts_sheet, make_header_cell, save_workbook, TIER_COLORS, TIER_TEXT
from functions.config import cfg
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Email validation
# ---------------------------------------------------------------------------

_EMAIL_RE  = re.compile(r'^[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-]+(?:\.[a-zA-Z0-9\-]+)+$')
_EMAIL_BAD = re.compile(
    r'('
    r'noreply|no-reply|donotreply|do-not-reply|unsubscribe|optout|opt-out|'
    r'example|localhost|invalid|dummy|placeholder|fake|sample|'
    r'test@|testmail|mailtest|'
    r'guided-selling|guided_selling|automated|automailer|autorespond|'
    r'besttemplate|template@|campaign@|mailer-daemon|'
    r'bounce@|bounced@|ndr@|postmaster|mailer@|'
    r'my-orders|my-packages|order-notify|orders-packages|'
    r'tracking@|shipment@|invoice@|notification@|'
    r'vtex\\.|system@|robot@|bot@|no_reply|_noreply'
    r')',
    re.IGNORECASE
)

_VALID_TLD_RE = re.compile(r'\.[a-zA-Z]{2,}$')  # requires 2+ letter TLD
_NUMERIC_DOMAIN_RE = re.compile(r'@[0-9]+\.')    # block @0., @1., @123. domains



def _valid_email(email: str) -> bool:
    # Reject strings with control chars or JSON-artifact characters
    # e.g. 'partner","slug":"anne-sofie...' leaked from JSON parsing
    if any(ord(c) < 32 or ord(c) == 127 for c in email):
        return False
    if any(c in email for c in ('"', '{', '}', '\\', '<', '>')):
        return False
    if not email or '@' not in email:
        return False
    local, _, domain = email.partition('@')
    if not local or not domain:
        return False
    if not _EMAIL_RE.match(email):
        return False
    # Reject single-char TLDs (.x, .y) and numeric-only TLDs (.1, .123)
    if not _VALID_TLD_RE.search(domain):
        return False
    # Reject numeric domain prefix like @0., @1., @123.
    if _NUMERIC_DOMAIN_RE.search(email):
        return False
    if _EMAIL_BAD.search(email):
        return False
    # Reject hex/UUID local parts (automated system addresses)
    if len(local) >= 16 and re.fullmatch(r'[0-9a-f\-]+', local):
        return False
    # Reject all-digit local parts
    if re.fullmatch(r'[0-9\-]+', local):
        return False
    return True

# ---------------------------------------------------------------------------
# Tier classification
# ---------------------------------------------------------------------------

WP_TAGS    = {'wordpress', 'woocommerce', 'elementor', 'divi', 'wp'}
SMB_TAGS   = {'smb', 'local', 'small business'}
CARE_TAGS  = {'care_plan', 'hosting', 'maintenance'}


def _score_lead(lead: dict, email_count: int) -> tuple[int, str]:
    potential  = (lead.get('ai_reseller_potential') or '').lower()
    score      = int(float(lead.get('reseller_score') or 0))
    priority   = (lead.get('priority') or '').upper()
    specs      = [s.lower() for s in (lead.get('ai_specialisation') or [])]
    client_base = (lead.get('ai_client_base') or '').lower()

    is_wp   = any(t in specs for t in WP_TAGS)
    is_smb  = client_base in ('smb', 'local') or any(t in specs for t in SMB_TAGS)
    is_care = any(t in specs for t in CARE_TAGS)

    # Base tier from AI potential + score
    if potential == 'high' and score >= 75:
        tier = 1
    elif potential == 'high':
        tier = 2
    elif potential == 'medium' and score >= 55:
        tier = 3
    elif potential == 'medium':
        tier = 4
    else:
        tier = 5

    # Bonus: bump up one tier for strong reseller signals
    bonuses = sum([is_wp, is_smb, is_care, email_count >= 2])
    if bonuses >= 2 and tier > 1:
        tier -= 1
    elif bonuses == 1 and tier > 2:
        tier -= 1

    labels = {
        1: '1 - Prime Reseller',
        2: '2 - Strong Prospect',
        3: '3 - Good Prospect',
        4: '4 - Possible',
        5: '5 - Low Priority',
    }
    return tier, labels[tier]

# ---------------------------------------------------------------------------
# Firestore
# ---------------------------------------------------------------------------

def _load_secrets():
    """Load Firebase credentials from env (FIREBASE_KEY_JSON or FIREBASE_CREDENTIALS)."""
    from dotenv import load_dotenv
    load_dotenv()
    from functions.firebase_cred import get_firebase_cred
    return get_firebase_cred()


def _init_firestore(fb_key):
    import firebase_admin
    from firebase_admin import firestore
    import firebase_admin.credentials as creds
    # fb_key may already be a Certificate object from get_firebase_cred()
    if isinstance(fb_key, creds.Certificate):
        cred = fb_key
    elif fb_key:
        cred = (fb_key if isinstance(fb_key, creds.Base) else creds.Certificate(fb_key))
    else:
        cred = creds.Certificate(
            cfg.FIREBASE_CREDENTIALS or "config/serviceAccountKey.json")
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    return firestore.client()

# ---------------------------------------------------------------------------
# email_contacts writer helpers
# ---------------------------------------------------------------------------

def _doc_id(email: str) -> str:
    """Stable Firestore document ID from an email address."""
    return re.sub(r'[^a-zA-Z0-9_]', '_', email.lower())


def _derive_name_from_email(email: str) -> tuple[str, str]:
    """For personal emails with no known name, derive first/full from local part.
    john.smith@… → ('John', 'John Smith')
    """
    local = email.split('@')[0]
    parts = re.split(r'[._\-]', local)
    parts = [p.capitalize() for p in parts if p.isalpha() and len(p) > 1]
    if not parts:
        return ('', '')
    return parts[0], ' '.join(parts)


def _write_email_contacts(db, rows: list[dict], campaign: str | None,
                           dry_run: bool = False) -> int:
    """Write leads contacts to email_contacts Firestore collection.

    Strategy (same as site pipeline):
    - Document ID = _doc_id(email)  →  natural deduplication by email.
    - Data fields always overwritten (merge=True in batches of 400).
    - Lifecycle fields (status, created_at) only set for NEW documents.
    - Parallel pre-scan + batch writes — no per-doc round trips.
    Returns count of docs written.
    """
    from concurrent.futures import ThreadPoolExecutor as _TPE

    BATCH_SIZE = 400
    WORKERS    = 10
    col        = db.collection('email_contacts')
    now_ts     = datetime.now(timezone.utc).isoformat()

    # ── Build valid rows ───────────────────────────────────────────────────
    valid_rows = []
    for row in rows:
        email = (row.get('email') or '').strip()
        if not email:
            continue
        name = clean_str((row.get('name') or '').strip())
        # Clear name if it doesn't correspond to the email address
        if name and not email_matches_name(email, name):
            name = ''
        # For personal emails with no known name, derive from email local part
        if not name and (row.get('email_type') or '') == 'personal':
            first, name = _derive_name_from_email(email)
        else:
            first = (name.split() or [''])[0]
        doc_id = _doc_id(email)
        valid_rows.append((doc_id, email, name, first, row))

    if not valid_rows:
        print('  [write] No valid contacts to write.', flush=True)
        return 0

    total = len(valid_rows)
    print(f'  [write] {total} contacts to write…', flush=True)

    if dry_run:
        from collections import Counter as _Ctr
        tier_c = _Ctr(row.get('tier', 0) for _, _, _, _, row in valid_rows)
        prio_c = _Ctr(str(row.get('outreach_priority', '?')) for _, _, _, _, row in valid_rows)
        print(f'  [DRY] Would write {total} contacts to email_contacts', flush=True)
        print(f'  [DRY] Tiers:    { dict(sorted(tier_c.items())) }', flush=True)
        print(f'  [DRY] Priority: { dict(sorted(prio_c.items())) }', flush=True)
        return total

    # ── Step A: parallel pre-scan for existing docs ────────────────────────
    print(f'  [write] Step A: scanning {total} existing docs…', flush=True)
    doc_ids = [r[0] for r in valid_rows]
    existing_ids: set[str] = set()

    def _check_batch(ids):
        refs  = [col.document(i) for i in ids]
        snaps = db.get_all(refs)
        return {s.id for s in snaps if s.exists}

    id_batches = [doc_ids[i:i+BATCH_SIZE] for i in range(0, len(doc_ids), BATCH_SIZE)]
    with _TPE(max_workers=WORKERS) as pool:
        for result in pool.map(_check_batch, id_batches):
            existing_ids |= result
    new_count = total - len(existing_ids)
    print(f'  [write] Step A done: {len(existing_ids)} existing, {new_count} new', flush=True)

    # ── Step B: build doc dicts ────────────────────────────────────────────
    data_docs:      list[tuple[str, dict]] = []
    lifecycle_docs: list[tuple[str, dict]] = []

    for doc_id, email, name, first, row in valid_rows:
        doc = {
            # Identity
            'doc_id':            doc_id,
            'approved':           '',
            'lead_id_leads':     row.get('lead_id_leads', ''),
            'mark_leads':        True,
            'email':             email,
            'name':              name,
            'title':             clean_str(row.get('title', '')),
            'phone':             row.get('phone', ''),
            'linkedin':          row.get('linkedin', ''),
            # Source agency
            'domain':            row.get('domain', ''),
            'website':           normalize_url(row.get('website', '') or ''),
            'company':           row.get('company', ''),
            'country':           resolve_country(row),
            # Classification
            'ai_sector':         row.get('ai_sector', ''),
            'ai_platform':       row.get('platform', ''),
            'ai_potential':      row.get('ai_potential', ''),
            'ai_client_base':    row.get('client_base', ''),
            'ai_summary':        row.get('summary', ''),
            # Contact scoring
            'tier':              row.get('tier', 0),
            'tier_label':        row.get('tier_label', ''),
            'email_type':        row.get('email_type', ''),
            'contact_type':      row.get('contact_type', ''),
            'outreach_priority': row.get('outreach_priority', 4),
            'reseller_score':    row.get('reseller_score', 0),
            # Origin
            'category_leads':    row.get('source', ''),
            # Pipeline metadata
            # Mail-merge
            'personalisation': {
                'name':      first,
                'full_name': name,
            },
        }
        data_docs.append((doc_id, doc))
        if doc_id not in existing_ids:
            lifecycle_docs.append((doc_id, {'status': 'pending', 'created_at': now_ts, 'campaign': campaign or ''}))

    # ── Step C: batch write data fields ───────────────────────────────────
    total_batches = (len(data_docs) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f'  [write] Step C: writing data in {total_batches} batch(es)…', flush=True)
    done = 0
    for i in range(0, len(data_docs), BATCH_SIZE):
        chunk = data_docs[i:i+BATCH_SIZE]
        batch = db.batch()
        for doc_id, doc in chunk:
            batch.set(col.document(doc_id), doc, merge=True)
        batch.commit()
        done += len(chunk)
        print(f'  [write]   {done}/{len(data_docs)} data docs written', flush=True)

    # ── Step D: lifecycle fields — new docs only ───────────────────────────
    if lifecycle_docs:
        print(f'  [write] Step D: lifecycle on {len(lifecycle_docs)} new docs…', flush=True)
        done = 0
        for i in range(0, len(lifecycle_docs), BATCH_SIZE):
            chunk = lifecycle_docs[i:i+BATCH_SIZE]
            batch = db.batch()
            for doc_id, lc in chunk:
                batch.set(col.document(doc_id), lc, merge=True)
            batch.commit()
            done += len(chunk)

    print(f'  [write] Done: {len(data_docs)} written ({len(lifecycle_docs)} new, {len(existing_ids)} updated)', flush=True)
    return len(data_docs)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_data(db, countries: list[str] | None, min_score: int,
               outreach_priority: int | None = None,
               collection: str = 'leads') -> list[dict]:
    from google.cloud.firestore_v1.base_query import FieldFilter

    from concurrent.futures import ThreadPoolExecutor
    SCAN_PARTITIONS = 32
    WORKERS         = 20
    LEAD_BATCH      = 300   # get_all batch size

    rows: list[dict] = []
    skipped = 0

    # Step 1: scan contacts collectionGroup in parallel partitions
    print('[leads-export] Step 1: scanning contacts in parallel partitions...', flush=True)
    cg = db.collection_group('contacts')

    def _scan_partition(query):
        local: dict[str, list[dict]] = {}
        for doc in query.stream():
            c = doc.to_dict() or {}
            if not _valid_email((c.get('email') or '').strip()):
                continue
            parts = doc.reference.path.split('/')   # {collection}/{lead_id}/contacts/{cid}
            if len(parts) >= 4:
                local.setdefault(parts[1], []).append(c)
        return local

    try:
        queries = [pt.query() for pt in cg.get_partitions(SCAN_PARTITIONS)]
    except Exception as exc:
        print(f'[leads-export]   partitioning unavailable ({exc}); single-pass scan', flush=True)
        queries = [cg.order_by('__name__')]

    contact_map: dict[str, list[dict]] = {}
    total_contacts = 0
    done_parts = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        for local in pool.map(_scan_partition, queries):
            for lid, lst in local.items():
                contact_map.setdefault(lid, []).extend(lst)
                total_contacts += len(lst)
            done_parts += 1
            print(f'[leads-export]   partition {done_parts}/{len(queries)} done  '
                  f'{total_contacts} valid contacts in {len(contact_map)} agencies', flush=True)
    print(f'[leads-export] Step 1 done: {total_contacts} valid contacts in {len(contact_map)} agencies', flush=True)

    # Step 2: batch-fetch parent lead docs in parallel
    print('[leads-export] Step 2: batch-fetching parent leads...', flush=True)
    leads_col = db.collection(collection)
    lead_data: dict[str, dict] = {}
    lead_ids  = list(contact_map.keys())

    def _fetch_batch(batch_ids):
        return db.get_all([leads_col.document(i) for i in batch_ids])

    batches = [lead_ids[i:i+LEAD_BATCH] for i in range(0, len(lead_ids), LEAD_BATCH)]
    n_batches = len(batches)
    print(f'[leads-export]   {len(lead_ids)} leads → {n_batches} batch(es) of {LEAD_BATCH}', flush=True)
    done_b = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        for res in pool.map(_fetch_batch, batches):
            for d in res:
                if d.exists:
                    lead_data[d.id] = d.to_dict() or {}
            done_b += 1
            print(f'[leads-export]   batch {done_b}/{n_batches}  ({len(lead_data)} leads fetched so far)', flush=True)
    print(f'[leads-export] Step 2 done: {len(lead_data)} lead docs fetched', flush=True)

    # Step 3: merge + filter
    total_leads = len(contact_map)
    print(f'[leads-export] Step 3: merging and filtering {total_leads} agencies...', flush=True)
    done_l = 0
    for lead_id, valid in contact_map.items():
        done_l += 1
        if done_l % 100 == 0 or done_l == total_leads:
            print(f'[leads-export]   {done_l}/{total_leads} agencies processed  '
                  f'{len(rows)} contacts kept  {skipped} skipped', flush=True)
        lead = lead_data.get(lead_id)
        if not lead:
            skipped += 1
            continue

        if lead.get('country') == '*':
            skipped += 1
            continue

        if countries:
            c = resolve_country(lead)
            if c not in countries:
                skipped += 1
                continue

        score = int(float(lead.get('reseller_score') or 0))
        if score < min_score:
            skipped += 1
            continue

        # Outreach priority filter on contacts
        if outreach_priority is not None:
            valid = [c for c in valid
                     if c.get("outreach_priority") is None
                     or int(c.get("outreach_priority", 4)) <= outreach_priority]
        if not valid:
            skipped += 1
            continue

        tier, tier_label = _score_lead(lead, len(valid))

        specs = lead.get('ai_specialisation') or []
        base = {
            'tier':           tier,
            'tier_label':     tier_label,
            'domain':         lead.get('domain', ''),
            'website':        normalize_url(lead.get('website', '') or ''),
            'company':        lead.get('company', ''),
            'country':        resolve_country(lead),
            'reseller_score': score,
            'priority':       lead.get('priority', ''),
            'ai_potential':   lead.get('ai_reseller_potential', ''),
            'ai_sector':      lead.get('ai_sector', ''),
            'specialisation': ', '.join(specs[:6]),
            'client_base':    lead.get('ai_client_base', ''),
            'platform':       lead.get('ai_platform', ''),
            'summary':        (lead.get('ai_summary') or lead.get('description') or '')[:120],
            'reasons':        (lead.get('reasons') or '')[:100],
            'email_count':    len(valid),
            'mark_leads':   True,
            'category_leads': 'catalog' if lead.get('found_by_catalog') == 'yes' else 'search',
        }

        for contact in valid:
            rows.append({
                **base,
                'lead_id_leads':     lead_id,
                'email':             contact.get('email', ''),
                'name':              clean_str(contact.get('name', ''))
                                     if email_matches_name(contact.get('email',''), clean_str(contact.get('name',''))) else '',
                'title':             clean_str(contact.get('title', '')),
                'phone':             contact.get('phone', ''),
                'linkedin':          contact.get('linkedin', ''),
                'email_type':        contact.get('email_type', ''),
                'contact_type':      contact.get('contact_type', ''),
                'outreach_priority': contact.get('outreach_priority', ''),
            })

    print(f'[leads-export] Done: {len(rows)} contacts in '
          f'{len(set(r["domain"] for r in rows))} agencies  ({skipped} leads skipped)', flush=True)
    return rows

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

TIER_COLORS = {
    1: 'C00000', 2: 'FF0000', 3: 'FF9900', 4: 'FFFF00', 5: 'D9D9D9',
}
TIER_TEXT = {1: 'FFFFFF', 2: 'FFFFFF', 3: '000000', 4: '000000', 5: '000000'}


def _build_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: All Contacts ───────────────────────────────────────────
    ws = wb.active
    ws.title = 'Contacts'

    COLS = [
        ('Approved',       'approved',        10),
        ('Tier',           'tier_label',     20),
        ('Score',          'reseller_score',  8),
        ('Priority',       'priority',       12),
        ('AI Potential',   'ai_potential',   12),
        ('Domain',         'domain',         28),
        ('Company',        'company',        24),
        ('Email',          'email',          30),
        ('Website',       'website',        35),
        ('Website',       'website',           35),
        ('Name',           'name',           22),
        ('Title',          'title',          20),
        ('Phone',          'phone',          14),
        ('LinkedIn',       'linkedin',       30),
        ('Email Type',     'email_type',     12),
        ('Contact Role',   'contact_type',   16),
        ('Outreach P',     'outreach_priority', 9),
        ('Country',        'country',         8),
        ('Sector',         'ai_sector',      16),
        ('Specialisation', 'specialisation', 30),
        ('Client Base',    'client_base',    12),
        ('Platform',       'platform',       14),
        ('Category',       'category_leads', 10),
        ('Mark',           'mark_leads',      14),
        ('Summary',        'summary',        50),
        ('Reasons',        'reasons',        35),
        ('Doc ID',         'doc_id',          28),
        ('Lead ID Leads',  'lead_id_leads',   28),
    ]

    write_contacts_sheet(
        ws, rows, COLS,
        sort_key  = lambda r: (int(r.get('tier') or 5), -int(r.get('reseller_score') or 0)),
        wrap_keys = {'summary', 'reasons', 'linkedin', 'specialisation'},
    )

    # ── Sheet 2: Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 10

    for ci, hdr in enumerate(['Tier', 'Agencies', 'Contacts'], 1):
        make_header_cell(ws2, 1, ci, hdr)

    tier_labels = {1:'1-Prime Reseller', 2:'2-Strong Prospect',
                   3:'3-Good Prospect', 4:'4-Possible', 5:'5-Low Priority'}
    tier_sites = Counter()
    tier_c     = Counter(r['tier'] for r in rows)
    for r in rows:
        tier_sites[(r['tier'], r['domain'])] = 1
    ts_count: dict[int, int] = {}
    for (t, _) in tier_sites:
        ts_count[t] = ts_count.get(t, 0) + 1

    for ri, t in enumerate(sorted(tier_labels), 2):
        bg, fg = TIER_COLORS.get(t, 'FFFFFF'), TIER_TEXT.get(t, '000000')
        for ci, val in enumerate([tier_labels[t], ts_count.get(t, 0), tier_c.get(t, 0)], 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.font = Font(name='Arial', size=10, color=fg)

    ws2.cell(8, 1, 'TOTAL').font = Font(name='Arial', bold=True)
    ws2.cell(8, 2, len(set(r['domain'] for r in rows))).font = Font(name='Arial', bold=True)
    ws2.cell(8, 3, len(rows)).font = Font(name='Arial', bold=True)

    # Specialisation breakdown
    ws2.cell(10, 1).value = 'Top Specialisations'
    ws2.cell(10, 1).font  = Font(name='Arial', bold=True)
    spec_c: Counter = Counter()
    for r in rows:
        for s in r['specialisation'].split(', '):
            if s.strip():
                spec_c[s.strip()] += 1
    for ri, (sp, cnt) in enumerate(spec_c.most_common(12), 11):
        ws2.cell(ri, 1, sp)
        ws2.cell(ri, 2, cnt)

    # Client base breakdown
    ws2.cell(10, 4).value = 'Client Base'
    ws2.cell(10, 4).font  = Font(name='Arial', bold=True)
    cb_c = Counter(r['client_base'] or 'unknown' for r in rows)
    for ri, (cb, cnt) in enumerate(cb_c.most_common(), 11):
        ws2.cell(ri, 4, cb)
        ws2.cell(ri, 5, cnt)

    # ── Sheet 3: WordPress focus ─────────────────────────────────────────
    ws3 = wb.create_sheet('WP vs Others')
    wp_rows = [r for r in rows if any(t in r['specialisation'].lower()
                                       for t in ['wordpress', 'woocommerce'])]
    ot_rows = [r for r in rows if r not in wp_rows]

    for ci, hdr in enumerate(['Group', 'Agencies', 'Contacts', 'Avg Score'], 1):
        make_header_cell(ws3, 1, ci, hdr)
    ws3.column_dimensions['A'].width = 22
    for col in ['B', 'C', 'D']: ws3.column_dimensions[col].width = 12

    for ri, (label, grp) in enumerate([
        ('WordPress / WooCommerce', wp_rows),
        ('Other specialisations', ot_rows),
    ], 2):
        sites   = len(set(r['domain'] for r in grp))
        avg_scr = int(sum(r['reseller_score'] for r in grp) / len(grp)) if grp else 0
        for ci, val in enumerate([label, sites, len(grp), avg_scr], 1):
            ws3.cell(ri, ci, val).font = Font(name='Arial', size=10)

    save_workbook(wb, out_path, '[leads-export]')

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description='Tiered reseller prospect export from leads')
    p.add_argument('--countries', nargs='+', default=None, metavar='CC',
                   help='Country codes e.g. UK IN NO (comma or space separated)')
    p.add_argument('--min-score', type=int, default=0, metavar='N',
                   help='Minimum reseller_score (default 0)')
    p.add_argument('--outreach-priority', type=int, default=None, metavar='N',
                   help='Only include contacts with outreach_priority <= N (1=best)')
    p.add_argument('--collection', default='leads', metavar='NAME',
                   help='Firestore collection (default: leads)')
    p.add_argument('--out', default=None, metavar='PATH',
                   help='Output .xlsx path (default: exports/leads_prospects_<countries>_<ts>.xlsx)')
    p.add_argument('--write-contacts', action='store_true',
                   help='Write contacts to email_contacts Firestore collection')
    p.add_argument('--campaign', default=None, metavar='NAME',
                   help='Campaign tag written to email_contacts (e.g. UK_resellers_jun02)')
    p.add_argument('--dry-run-contacts', action='store_true',
                   help='Print what would be written to email_contacts without writing')
    args = p.parse_args(argv)

    countries = None
    if args.countries:
        raw = []
        for t in args.countries:
            raw.extend(c.strip().upper() for c in t.split(',') if c.strip())
        countries = raw or None

    fb_key = _load_secrets()
    db     = _init_firestore(fb_key)

    rows = _load_data(db, countries, args.min_score, args.outreach_priority, args.collection)
    if not rows:
        print('[leads-export] No contacts found.')
        return

    if not args.out:
        ts  = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')
        cc  = '_'.join(countries) if countries else 'all'
        out = Path('exports') / f'leads_prospects_{cc}_{ts}.xlsx'
    else:
        out = Path(args.out)

    _build_excel(rows, out)

    tc = Counter(r['tier_label'] for r in rows)
    print('\n[leads-export] Tier breakdown:')
    for label, count in sorted(tc.items()):
        print(f'  {label}: {count} contacts')

    # ── Write to email_contacts if requested ──────────────────────────────
    if args.write_contacts or args.dry_run_contacts:
        tag = '[DRY RUN] ' if args.dry_run_contacts else ''
        print(f'\n[leads-export] {tag}Writing {len(rows)} contacts to email_contacts…', flush=True)
        n = _write_email_contacts(
            db, rows,
            campaign  = args.campaign,
            dry_run   = args.dry_run_contacts,
        )
        print(f'[leads-export] {tag}{n} contacts written to email_contacts', flush=True)


if __name__ == '__main__':
    main()
