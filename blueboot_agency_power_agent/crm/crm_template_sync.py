"""
crm_template_sync.py -- Read the CRM template Google Sheet and sync to Firestore.

Reads the 'Outreach' tab from the crm_template sheet, maps Norwegian column
headers to field names, and upserts each row into:

    crm / crm_template / items / {doc_id}

Doc ID is derived from the Nettside (website) column using lead_id_from_url,
falling back to a slug of the Bedrift (company) name if no website is set.

Usage:
    python crm/crm_template_sync.py
    python crm/crm_template_sync.py --tab Outreach
    python crm/crm_template_sync.py --dry-run
"""
from __future__ import annotations

import sys
import re
import threading
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "app"))
import _pathsetup  # noqa: F401,F811

from dotenv import load_dotenv
load_dotenv()

from functions.models import lead_id_from_url
from functions.utils import normalize_url
from functions.firebase_cred import get_firebase_cred
import firebase_admin
from firebase_admin import firestore
import firebase_admin.credentials as fb_creds

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build

# -- Config -------------------------------------------------------------------
from crm.config import CRM_TEMPLATE_ID

SHEET_ID      = CRM_TEMPLATE_ID
TAB_NAME      = "Outreach"
SCOPES        = ["https://www.googleapis.com/auth/spreadsheets"]
TOKEN_PATH    = str(Path(__file__).resolve().parent.parent / "config" / "google_token.json")
CLIENT_SECRET = str(Path(__file__).resolve().parent.parent / "config" / "google_oauth_client.json")

CRM_COLLECTION = "crm"
CRM_DOC        = "crm_template"

# Map Norwegian header -> Firestore field name
HEADER_MAP = {
    "dato lagt i":      "created_date",
    "bedrift":          "company",
    "nettside":         "website",
    "bransje":          "sector",
    "størrelse":        "size",
    "beslutningstaker": "decision_maker",
    "rolle":            "role",
    "e-post":           "email",
    "telefon":          "phone",
    "score":            "score",
    "status":           "status",
    "selger":           "seller",
    "kommentar":        "comment",
    "tilbud":           "offer",
    "site_lead_id":     "site_lead_id",
}

# -- Auth ---------------------------------------------------------------------
_fb_lock = threading.Lock()


def _init_firestore():
    cred_obj = get_firebase_cred()
    cred = cred_obj if isinstance(cred_obj, fb_creds.Base) else fb_creds.Certificate(cred_obj)
    with _fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    return firestore.client()


def _sheets_service():
    creds = None
    if Path(TOKEN_PATH).exists():
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow  = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET, SCOPES)
            creds = flow.run_local_server(port=0)
        Path(TOKEN_PATH).write_text(creds.to_json())
    return build("sheets", "v4", credentials=creds)


# -- Sheet reading ------------------------------------------------------------

def _read_sheet(svc, tab: str) -> list[dict]:
    """Read the sheet tab and return rows as dicts with normalised field names."""
    print(f"[crm-template] Reading sheet tab '{tab}'...", flush=True)
    result = svc.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range=f"{tab}!A:ZZ"
    ).execute()
    rows = result.get("values", [])
    if not rows:
        print("[crm-template] Sheet is empty.")
        return []

    raw_headers = rows[0]
    fields = [HEADER_MAP.get(h.lower().strip(), h.lower().strip().replace(" ", "_"))
              for h in raw_headers]

    records = []
    for row in rows[1:]:
        rec = {}
        for i, field in enumerate(fields):
            rec[field] = row[i].strip() if i < len(row) else ""
        # Skip fully empty rows
        if not any(v for v in rec.values()):
            continue
        records.append(rec)

    print(f"[crm-template] {len(records)} rows read", flush=True)
    return records


# -- Doc ID generation --------------------------------------------------------

def _make_doc_id(rec: dict) -> str:
    """Generate a Firestore-safe doc ID from website or company name."""
    website = (rec.get("website") or "").strip()
    if website and website.startswith("http"):
        return lead_id_from_url(normalize_url(website))
    if website:
        return lead_id_from_url(normalize_url("https://" + website))
    company = (rec.get("company") or "").strip()
    if company:
        slug = re.sub(r"[^a-z0-9]+", "_", company.lower()).strip("_")
        return slug[:80]
    return ""


# -- Firestore upsert ---------------------------------------------------------

def _upsert_to_firestore(db, records: list[dict], dry_run: bool = False) -> int:
    col   = db.collection(CRM_COLLECTION).document(CRM_DOC).collection("items")
    BATCH_SIZE = 400
    count = 0
    total = len(records)
    pairs = []

    for rec in records:
        doc_id = _make_doc_id(rec)
        if not doc_id:
            print(f"  [skip] could not derive doc_id for row: {rec.get('company','?')}")
            continue
        rec["doc_id"] = doc_id
        pairs.append((doc_id, rec))

    if dry_run:
        print(f"[crm-template] DRY RUN — would upsert {len(pairs)} docs to crm/{CRM_DOC}/items")
        for doc_id, rec in pairs[:5]:
            print(f"  {doc_id}: {rec.get('company')} / {rec.get('email')}")
        if len(pairs) > 5:
            print(f"  ... and {len(pairs)-5} more")
        return len(pairs)

    for i in range(0, len(pairs), BATCH_SIZE):
        chunk = pairs[i:i + BATCH_SIZE]
        batch = db.batch()
        for doc_id, rec in chunk:
            batch.set(col.document(doc_id), rec, merge=True)
        batch.commit()
        count += len(chunk)
        print(f"[crm-template]   upserted {count}/{total}...", flush=True)

    print(f"[crm-template] Synced {count} docs -> crm/{CRM_DOC}/items")
    return count



# -- Site leads enrichment ----------------------------------------------------

def _ensure_site_lead_id_col(svc) -> int:
    """Ensure sheet has a site_lead_id column as the last column.
    Returns the 1-based column index."""
    result = svc.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range=f"{TAB_NAME}!1:1"
    ).execute()
    headers = result.get("values", [[]])[0]

    if "site_lead_id" in headers:
        return headers.index("site_lead_id") + 1

    # Append as new last column
    col_idx = len(headers) + 1
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col_idx)
    svc.spreadsheets().values().update(
        spreadsheetId=SHEET_ID,
        range=f"{TAB_NAME}!{col_letter}1",
        valueInputOption="USER_ENTERED",
        body={"values": [["site_lead_id"]]},
    ).execute()
    print(f"[crm-enrich] Added 'site_lead_id' column at {col_letter}", flush=True)
    return col_idx


def _write_site_lead_ids_to_sheet(svc, matches: dict[str, str], tab: str) -> None:
    """Write site_lead_id values back to the sheet for matched rows.

    matches: {website_raw: site_lead_id}
    """
    result = svc.spreadsheets().values().get(
        spreadsheetId=SHEET_ID, range=f"{tab}!A:ZZ"
    ).execute()
    rows = result.get("values", [])
    if not rows:
        return

    headers = rows[0]
    website_idx    = next((i for i, h in enumerate(headers) if h.lower().strip() in ("nettside", "website")), -1)
    site_lead_idx  = next((i for i, h in enumerate(headers) if h.lower().strip() == "site_lead_id"), -1)

    if website_idx < 0 or site_lead_idx < 0:
        print("[crm-enrich] Could not find website or site_lead_id column in sheet")
        return

    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(site_lead_idx + 1)
    updates = []

    for ri, row in enumerate(rows[1:], 2):
        website = row[website_idx].strip() if website_idx < len(row) else ""
        if not website:
            continue
        # Normalize for lookup
        ws_norm = website if website.startswith("http") else "https://" + website
        site_id = lead_id_from_url(normalize_url(ws_norm))
        if site_id in matches.values():
            # Find which raw website produced this site_id
            matched_id = matches.get(website) or matches.get(ws_norm)
            if matched_id:
                updates.append((ri, matched_id))

    # Write in one batch
    if updates:
        data = []
        # build sparse update using individual cell writes
        for row_num, site_id in updates:
            data.append({
                "range": f"{tab}!{col_letter}{row_num}",
                "values": [[site_id]],
            })
        svc.spreadsheets().values().batchUpdate(
            spreadsheetId=SHEET_ID,
            body={"valueInputOption": "USER_ENTERED", "data": data},
        ).execute()
        print(f"[crm-enrich] Wrote site_lead_id for {len(updates)} rows in sheet", flush=True)


def enrich_from_site_leads(db, svc=None, dry_run: bool = False) -> None:
    """For each crm_template item, look up matching doc in site_leads by
    normalized website URL and merge site_leads data into the crm_template item.
    Also writes site_lead_id back to the Google Sheet.
    """
    crm_col  = db.collection(CRM_COLLECTION).document(CRM_DOC).collection("items")
    site_col = db.collection("site_leads")

    print("[crm-enrich] Loading crm_template items...", flush=True)
    docs = list(crm_col.stream())
    print(f"[crm-enrich] {len(docs)} items to enrich", flush=True)

    found = 0
    not_found = 0
    BATCH_SIZE = 400
    updates = []
    # website_raw -> site_lead_id for sheet writeback
    sheet_matches: dict[str, str] = {}

    for doc in docs:
        crm_data = doc.to_dict() or {}
        website  = (crm_data.get("website") or "").strip()
        if not website:
            not_found += 1
            continue

        ws_norm = website if website.startswith("http") else "https://" + website
        site_id = lead_id_from_url(normalize_url(ws_norm))

        site_doc = site_col.document(site_id).get()
        if not site_doc.exists:
            print(f"  [miss] {site_id}", flush=True)
            not_found += 1
            continue

        site_data = site_doc.to_dict() or {}
        merged = {**site_data, **crm_data}
        merged["site_lead_id"]      = site_id
        merged["site_lead_matched"] = True

        updates.append((doc.id, merged))
        sheet_matches[website] = site_id
        found += 1
        print(f"  [match] {site_id} -> {crm_data.get('company','?')}", flush=True)

    print(f"[crm-enrich] Matched {found}, not found {not_found}", flush=True)

    if dry_run:
        print(f"[crm-enrich] DRY RUN -- would update {len(updates)} docs in Firestore")
        return

    # Write to Firestore
    count = 0
    total = len(updates)
    for i in range(0, total, BATCH_SIZE):
        chunk = updates[i:i + BATCH_SIZE]
        batch = db.batch()
        for doc_id, data in chunk:
            batch.set(crm_col.document(doc_id), data, merge=True)
        batch.commit()
        count += len(chunk)
        print(f"[crm-enrich]   updated {count}/{total}...", flush=True)

    # Write site_lead_id back to sheet
    if svc and sheet_matches:
        _ensure_site_lead_id_col(svc)
        _write_site_lead_ids_to_sheet(svc, sheet_matches, TAB_NAME)

    print(f"[crm-enrich] Done -- {count} items enriched from site_leads.")


def _update_site_leads_from_sheet(db, records: list[dict], dry_run: bool = False) -> int:
    """For each sheet row with a site_lead_id, patch site_leads with
    crm_status, crm_sales_person, crm_date — only non-blank values."""
    site_col = db.collection("site_leads")
    BATCH_SIZE = 400
    updates = []

    for rec in records:
        site_id = (rec.get("site_lead_id") or "").strip()
        if not site_id:
            continue
        patch = {}
        if rec.get("status"):
            patch["crm_status"] = rec["status"]
        if rec.get("seller"):
            patch["crm_sales_person"] = rec["seller"]
        if rec.get("created_date") or rec.get("dato_lagt_i"):
            patch["crm_date"] = rec.get("created_date") or rec.get("dato_lagt_i")
        if patch:
            updates.append((site_id, patch))

    print(f"[crm-template] Updating {len(updates)} site_leads docs with CRM fields...", flush=True)

    if dry_run:
        for sid, patch in updates[:5]:
            print(f"  {sid}: {patch}")
        if len(updates) > 5:
            print(f"  ... and {len(updates)-5} more")
        return len(updates)

    count = 0
    total = len(updates)
    for i in range(0, total, BATCH_SIZE):
        chunk = updates[i:i + BATCH_SIZE]
        batch = db.batch()
        for site_id, patch in chunk:
            batch.update(site_col.document(site_id), patch)
        batch.commit()
        count += len(chunk)
        print(f"[crm-template]   site_leads updated {count}/{total}...", flush=True)

    print(f"[crm-template] site_leads CRM fields updated for {count} docs")
    return count

# -- Main ---------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description="Sync CRM template sheet -> Firestore")
    p.add_argument("--tab",     default=TAB_NAME, metavar="TAB",
                   help=f"Sheet tab name (default: {TAB_NAME})")
    p.add_argument("--dry-run", action="store_true",
                   help="Print what would be written without touching Firestore")
    p.add_argument("--enrich",  action="store_true",
                   help="Match items to site_leads by website URL and merge data")
    args = p.parse_args(argv)

    db = _init_firestore()

    if args.enrich:
        svc = _sheets_service()
        enrich_from_site_leads(db, svc=svc, dry_run=args.dry_run)
        return

    svc     = _sheets_service()
    records = _read_sheet(svc, args.tab)

    if not records:
        print("[crm-template] Nothing to sync.")
        return

    _upsert_to_firestore(db, records, dry_run=args.dry_run)
    _update_site_leads_from_sheet(db, records, dry_run=args.dry_run)
    print("[crm-template] Done.")


if __name__ == "__main__":
    main()
