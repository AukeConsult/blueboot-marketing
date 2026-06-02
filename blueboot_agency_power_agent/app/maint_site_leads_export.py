"""site_leads_export.py -- Export site_leads + contacts to an Excel spreadsheet.

Reads site_leads from Firestore (with optional country / AI-sector filter),
fetches each lead's site_contacts sub-collection, and writes a single .xlsx
file where every lead is one row.  All contacts are folded into one "contacts"
column, one contact per line formatted as:

    Name | email@example.com | +47 123 456 78

Usage:
    python app/site_leads_export.py
    python app/site_leads_export.py --countries NO,SE
    python app/site_leads_export.py --countries NO --output exports/no_leads.xlsx
    python app/site_leads_export.py --sector technology --limit 500
    python app/site_leads_export.py --with-contacts-only
    python app/site_leads_export.py --dry-run
"""
from __future__ import annotations

import threading as _threading

# Guards firebase_admin.initialize_app against concurrent init
_local_fb_lock = _threading.Lock()
import argparse
import importlib.util
import os
from datetime import datetime, timezone
from pathlib import Path

import _pathsetup  # noqa: F401

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

COLLECTION_DEFAULT = "site_leads"

# Columns written to the sheet, in order.
# "contacts" is always appended at the end of the lead columns.
LEAD_COLUMNS: list[tuple[str, str, int]] = [
    # (firestore_key,   header_label,       col_width)
    ("domain",          "Domain",           28),
    ("website",         "Website",          34),
    ("company",         "Company",          28),
    ("country",         "Country",           8),
    ("country_name",    "Country Name",     16),
    ("query_category",  "Category",         16),
    ("page_count",      "Pages",             9),
    ("sitemap_type",    "Sitemap Type",     13),
    ("platform",        "Platform",         13),
    ("sitemap_oldest_date", "Oldest Content", 14),
    ("sitemap_newest_date", "Newest Content", 14),
    ("sitemaps",    "Sitemap URLs",     55),
    ("ai_sector",       "AI Sector",        18),
    ("ai_company_type", "AI Type",          12),
    ("ai_country",      "AI Country",       10),
    ("ai_confidence",   "AI Conf",           9),
    ("ai_keywords",     "AI Keywords",      45),
    ("ai_summary",      "AI Summary",       40),
    ("location",         "Location",         35),
    ("location_city",    "City",             18),
    ("location_country", "Loc Country",      10),
    ("location_confidence", "Loc Conf",       8),
    ("title",           "Title",            35),
    ("description",     "Description",      50),
    ("keywords",        "Keywords",         40),
    ("crawled_at",      "Crawled At",       20),
    ("source_query",    "Source Query",     30),
    ("lead_id",         "Lead ID",          36),
]

CONTACTS_COL_WIDTH = 60

# ---------------------------------------------------------------------------
# Secrets / Firestore (sync)
# ---------------------------------------------------------------------------

def _load_secrets():
    """Load Firebase credentials from env (FIREBASE_KEY_JSON or FIREBASE_CREDENTIALS)."""
    from dotenv import load_dotenv
    load_dotenv()
    from functions.firebase_cred import get_firebase_cred
    return get_firebase_cred()


def _init_firestore(fb_key_dict, collection: str):
    try:
        import firebase_admin
        from firebase_admin import firestore
        import firebase_admin.credentials as fb_creds
    except ImportError:
        raise RuntimeError("firebase-admin not installed — run: pip install firebase-admin")

    if isinstance(fb_key_dict, fb_creds.Certificate):
        cred = fb_key_dict
    elif fb_key_dict:
        cred = fb_creds.Certificate(fb_key_dict)
    else:
        cred = fb_creds.Certificate(os.getenv("FIREBASE_CREDENTIALS",
                                              "config/serviceAccountKey.json"))
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)

    db  = firestore.client()
    col = db.collection(collection)
    return db, col


# ---------------------------------------------------------------------------
# Firestore fetch
# ---------------------------------------------------------------------------

def _fetch_contacts(doc_ref) -> list[dict]:
    """Return list of contact dicts from the site_contacts sub-collection."""
    try:
        return [c.to_dict() for c in doc_ref.collection("site_contacts").stream()]
    except Exception as exc:
        print(f"  [export] contacts fetch error for {doc_ref.id}: {exc}")
        return []


def _format_contacts(contacts: list[dict]) -> str:
    """Fold contacts into a single multi-line string: Name | email | phone."""
    lines = []
    for c in contacts:
        name  = (c.get("name")  or "").strip()
        email = (c.get("email") or "").strip()
        phone = (c.get("phone") or "").strip()
        parts = [p for p in [name, email, phone] if p]
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines)


def _stream_leads(
    col,
    countries:          list[str] | None,
    sector:             str | None,
    category:           str | None,
    location:           str | None,
    with_contacts_only: bool,
    limit:              int | None,
    dry_run:            bool = False,
) -> list[dict]:
    """Stream site_leads; fetch sub-collection contacts per lead."""
    print("  [export] Streaming site_leads…", flush=True)
    rows: list[dict] = []
    scanned = skipped = 0

    for doc in col.stream():
        scanned += 1
        data = doc.to_dict() or {}

        # ── filters ────────────────────────────────────────────────────────
        if countries:
            c = (data.get("country") or "").upper()
            if c not in countries and c != "*":
                skipped += 1
                continue

        if sector:
            if (data.get("ai_sector") or "").lower() != sector.lower():
                skipped += 1
                continue

        if category:
            if (data.get("query_category") or "").lower() != category.lower():
                skipped += 1
                continue

        if location:
            loc_full = (data.get("location_full") or data.get("location") or "").lower()
            if location.lower() not in loc_full:
                skipped += 1
                continue

        # ── contacts ───────────────────────────────────────────────────────
        if dry_run:
            contacts = []   # skip sub-collection fetch during dry-run
        else:
            contacts = _fetch_contacts(doc.reference)

        if with_contacts_only and not contacts and not dry_run:
            skipped += 1
            continue

        data["_contacts_raw"] = contacts
        rows.append(data)

        if len(rows) % 100 == 0:
            print(f"  [export] ... {len(rows)} leads collected (scanned {scanned})", flush=True)

        if limit and len(rows) >= limit:
            break

    print(
        f"  [export] {scanned} scanned → {len(rows)} rows  "
        f"({skipped} skipped by filter)",
        flush=True,
    )
    return rows


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_xlsx(rows: list[dict], output_path: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "site_leads"

    # ── header style ────────────────────────────────────────────────────────
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill   = PatternFill("solid", start_color="2D5C8E")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_font     = Font(name="Arial", size=9)
    wrap_align    = Alignment(vertical="top", wrap_text=True)
    nowrap_align  = Alignment(vertical="top", wrap_text=False)
    thin_border   = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right =Side(style="thin", color="CCCCCC"),
    )

    # ── build column list: lead cols + contacts ──────────────────────────
    all_cols = LEAD_COLUMNS + [("contacts", "Contacts", CONTACTS_COL_WIDTH)]

    # ── write header row ────────────────────────────────────────────────────
    for col_idx, (_, label, _) in enumerate(all_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # ── row fill colours (alternating) ─────────────────────────────────────
    fill_white = PatternFill("solid", start_color="FFFFFF")
    fill_light = PatternFill("solid", start_color="F2F7FC")

    # ── write data rows ─────────────────────────────────────────────────────
    contacts_col_idx = len(all_cols)  # last column

    for row_idx, data in enumerate(rows, start=2):
        row_fill = fill_light if row_idx % 2 == 0 else fill_white

        for col_idx, (key, _, _) in enumerate(all_cols, start=1):
            if key == "contacts":
                value = _format_contacts(data.get("_contacts_raw") or [])
                align = wrap_align
            elif key == "sitemaps":
                # Format as: filename  [lastmod]  (one per line)
                raw = data.get("sitemaps") or []
                lines = []
                for s in raw:
                    if not isinstance(s, dict):
                        continue
                    name = s.get("filename") or s.get("url", "")
                    lm   = s.get("lastmod", "")
                    lm_n = s.get("lastmod_newest", "")
                    pc   = s.get("page_count")
                    pc_s = f"  ({pc:,} pages)" if isinstance(pc, int) and pc else ""
                    if lm_n and lm_n != lm:
                        lm_s = f"  [{lm} → {lm_n}]" if lm else f"  [→ {lm_n}]"
                    else:
                        lm_s = f"  [{lm}]" if lm else ""
                    lines.append(f"{name}{lm_s}{pc_s}")
                value = "\n".join(lines)
                align = wrap_align
            else:
                raw = data.get(key, "")
                if isinstance(raw, list):
                    value = ", ".join(str(x) for x in raw)
                elif isinstance(raw, float):
                    value = round(raw, 3)
                else:
                    value = raw if raw is not None else ""
                align = wrap_align if key in ("description", "ai_summary", "keywords") else nowrap_align

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = cell_font
            cell.fill      = row_fill
            cell.alignment = align
            cell.border    = thin_border

    # ── column widths ───────────────────────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(all_cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── auto row heights for wrap-text rows ─────────────────────────────────
    # Give contacts rows some extra height when there are multiple lines
    for row_idx, data in enumerate(rows, start=2):
        contacts = data.get("_contacts_raw") or []
        n_lines  = max(1, len(contacts))
        ws.row_dimensions[row_idx].height = min(15 * n_lines, 80)

    # ── summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    _build_summary(ws_sum, rows, header_font, header_fill, header_align, cell_font)

    wb.save(output_path)
    print(f"  [export] Saved → {output_path}")


def _build_summary(ws, rows, header_font, header_fill, header_align, cell_font):
    from openpyxl.styles import Font, Alignment
    from collections import Counter

    ws.title = "Summary"

    def _hdr(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    def _val(row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(name="Arial", size=9)

    total_leads    = len(rows)
    total_contacts = sum(len(r.get("_contacts_raw") or []) for r in rows)
    leads_w_contacts = sum(1 for r in rows if r.get("_contacts_raw"))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14

    r = 1
    _hdr(r, 1, "Metric");          _hdr(r, 2, "Value");  r += 1
    _val(r, 1, "Total leads");     _val(r, 2, total_leads);         r += 1
    _val(r, 1, "With contacts");   _val(r, 2, leads_w_contacts);    r += 1
    _val(r, 1, "Total contacts");  _val(r, 2, total_contacts);      r += 1
    _val(r, 1, "Exported at");
    _val(r, 2, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"));  r += 2

    # by country
    _hdr(r, 1, "Country");  _hdr(r, 2, "Leads");  r += 1
    for country, cnt in sorted(Counter(d.get("country","?") for d in rows).items(),
                               key=lambda x: -x[1]):
        _val(r, 1, country);  _val(r, 2, cnt);  r += 1
    r += 1

    # by category
    _hdr(r, 1, "Category");  _hdr(r, 2, "Leads");  r += 1
    for cat, cnt in sorted(Counter(d.get("query_category","—") for d in rows).items(),
                           key=lambda x: -x[1]):
        _val(r, 1, cat or "—");  _val(r, 2, cnt);  r += 1
    r += 1

    # by AI sector
    _hdr(r, 1, "AI Sector");  _hdr(r, 2, "Leads");  r += 1
    for sec, cnt in sorted(Counter(d.get("ai_sector","—") for d in rows).items(),
                           key=lambda x: -x[1]):
        _val(r, 1, sec or "—");  _val(r, 2, cnt);  r += 1


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def export_site_leads(
    collection:         str           = COLLECTION_DEFAULT,
    countries:          list[str] | None = None,
    sector:             str | None    = None,
    category:           str | None    = None,
    location:           str | None    = None,
    limit:              int | None    = None,
    output:             str | None    = None,
    with_contacts_only: bool          = False,
    dry_run:            bool          = False,
) -> str:
    """Export site_leads to .xlsx.  Returns the output file path."""
    fb_key = _load_secrets()
    _, col = _init_firestore(fb_key, collection)

    rows = _stream_leads(col, countries, sector, category, location, with_contacts_only, limit, dry_run=dry_run)
    if not rows:
        print("  [export] No leads matched the filters — nothing to export.")
        return ""

    if dry_run:
        print(f"  [export] dry-run — would export {len(rows)} leads, skipping file write.", flush=True)
        return ""

    # ── default output path ─────────────────────────────────────────────────
    if not output:
        ts         = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        suffix     = ("_" + "_".join(c.upper() for c in countries)) if countries else ""
        suffix    += ("_" + location.replace(" ", "_").replace(",", "")) if location else ""
        suffix    += ("_" + category)      if category           else ""
        suffix    += ("_" + sector)        if sector             else ""
        suffix    += "_contacts_only"      if with_contacts_only else ""
        out_dir    = Path(__file__).parent.parent / "exports"
        out_dir.mkdir(exist_ok=True)
        output     = str(out_dir / f"site_leads{suffix}_{ts}.xlsx")

    total_contacts = sum(len(r.get("_contacts_raw") or []) for r in rows)
    print(f"\n  [export] Leads      : {len(rows)}")
    print(f"  [export] Contacts   : {total_contacts}")
    print(f"  [export] Output     : {output}")
    _build_xlsx(rows, output)
    return output


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> None:
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    p = argparse.ArgumentParser(
        description="Export site_leads + contacts to Excel (.xlsx)"
    )
    p.add_argument("--collection",   default=COLLECTION_DEFAULT, metavar="NAME",
                   help=f"Firestore collection  (default: {COLLECTION_DEFAULT})")
    p.add_argument("--countries",    default=None, metavar="CODES",
                   help="Space or comma-separated country codes e.g. --countries NO SE UK")
    p.add_argument("--sector",       default=None, metavar="NAME",
                   help="Filter by ai_sector  e.g. technology, ecommerce")
    p.add_argument("--category",     default=None, metavar="NAME",
                   help="Filter by query_category  e.g. real_estate, company")
    p.add_argument("--location",     default=None, metavar="TEXT",
                   help="Filter by location_full keyword e.g. London, Pune, Manchester")
    p.add_argument("--limit",        type=int, default=None, metavar="N",
                   help="Max leads to export")
    p.add_argument("--out",       default=None, metavar="PATH",
                   help="Output .xlsx path  (default: exports/site_leads_<ts>.xlsx)")
    p.add_argument("--with-contacts-only", action="store_true",
                   help="Only export leads that have at least one contact")
    p.add_argument("--dry-run",      action="store_true",
                   help="Count leads and contacts without writing a file")

    args = p.parse_args(argv)
    countries = None
    if args.countries:
        countries = [c.strip().upper() for c in args.countries.split(",") if c.strip()]

    export_site_leads(
        collection         = args.collection,
        countries          = countries,
        sector             = args.sector,
        category           = args.category,
        location           = args.location,
        limit              = args.limit,
        output             = args.out,
        with_contacts_only = args.with_contacts_only,
        dry_run            = args.dry_run,
    )


if __name__ == "__main__":
    main()
