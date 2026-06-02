"""email_contacts_export.py — Export from email_contacts Firestore collection to Excel.

Reads directly from the unified email_contacts collection (written by both
site_smart_export --write-contacts and leads_smart_export --write-contacts).

Filters: country, campaign, status, pipeline mark (site/leads/both).

Usage:
    python app/email_contacts_export.py --countries NO
    python app/email_contacts_export.py --countries UK NO --status pending
    python app/email_contacts_export.py --campaign NO_resellers_jun02
    python app/email_contacts_export.py --mark site   # only SITE_LEADS contacts
    python app/email_contacts_export.py --mark leads  # only LEADS contacts
    python app/email_contacts_export.py --mark both   # contacts in both pipelines
"""
from __future__ import annotations

import threading as _threading
_local_fb_lock = _threading.Lock()

import sys
import argparse
import importlib.util
import os
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path as _Path
sys.path.insert(0, str(_Path(__file__).parent))
from functions.utils import clean_str, resolve_country
from functions.excel_builder import write_contacts_sheet, make_header_cell, save_workbook, TIER_COLORS, TIER_TEXT
from pathlib import Path

# ---------------------------------------------------------------------------
# Firestore bootstrap  (same pattern as other export scripts)
# ---------------------------------------------------------------------------

def _load_secrets():
    p = Path(__file__).parent.parent / 'blueboot_secrets.py'
    if not p.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location('blueboot_secrets', p)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return getattr(mod, 'fireBaseAdminKey', None)
    except Exception:
        return None


def _init_firestore(fb_key):
    import firebase_admin
    from firebase_admin import firestore
    import firebase_admin.credentials as creds
    cred = creds.Certificate(fb_key) if fb_key else creds.Certificate(
        os.getenv('FIREBASE_CREDENTIALS', 'config/serviceAccountKey.json'))
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    return firestore.client()

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_data(db,
               countries:  list[str] | None = None,
               campaign:   str | None        = None,
               status:     str | None        = None,
               mark:       str | None        = None,
               collection: str               = 'email_contacts') -> list[dict]:
    """Stream email_contacts, apply filters, return list of row dicts."""
    from google.cloud.firestore_v1.base_query import FieldFilter
    from concurrent.futures import ThreadPoolExecutor
    BATCH_SIZE = 400
    WORKERS    = 10

    col = db.collection(collection)

    # Build server-side filters where possible
    query = col
    if campaign:
        query = query.where(filter=FieldFilter('campaign', '==', campaign))
    if status:
        query = query.where(filter=FieldFilter('status', '==', status))
    if mark == 'site':
        query = query.where(filter=FieldFilter('mark_site_leads', '==', True))
    elif mark == 'leads':
        query = query.where(filter=FieldFilter('mark_leads', '==', True))

    print(f'[ec-export] Loading from {collection}…', flush=True)
    docs = list(query.stream())
    print(f'[ec-export] {len(docs)} docs fetched', flush=True)

    rows = []
    skipped = 0
    for doc in docs:
        d = doc.to_dict() or {}

        # mark=both: must have both flags
        if mark == 'both':
            if not (d.get('mark_site_leads') and d.get('mark_leads')):
                skipped += 1
                continue

        # country filter (client-side — country may be ISO or internal code)
        if countries:
            c = resolve_country(d)
            if c not in countries:
                skipped += 1
                continue

        rows.append(d)

    print(f'[ec-export] {len(rows)} rows after filters  ({skipped} skipped)', flush=True)
    return rows

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

COLS = [
    # Contact
    ('Approved',        'approved',          10),
    ('Tier',            'tier_label',        20),
    ('Outreach P',      'outreach_priority',  9),
    ('Status',          'status',            12),
    ('Email',           'email',             32),
    ('Name',            'name',              22),
    ('Title',           'title',             22),
    ('Phone',           'phone',             16),
    ('LinkedIn',        'linkedin',          30),
    ('Email Type',      'email_type',        12),
    ('Contact Role',    'contact_type',      16),
    # Source
    ('Domain',          'domain',            28),
    ('Website',         'website',           35),
    ('Company',         'company',           26),
    ('Country',         'country',            8),
    ('Location',        'location',          30),
    ('City',            'location_city',     18),
    ('Region',          'location_region',   16),
    # Classification
    ('Platform',        'ai_platform',       16),
    ('Sector',          'ai_sector',         16),
    ('Potential',       'ai_potential',      14),
    ('Client Base',     'ai_client_base',    14),
    ('Company Type',    'ai_company_type',   14),
    ('Confidence',      'ai_confidence',      9),
    ('Summary',         'ai_summary',        55),
    ('Keywords',        'keywords',          35),
    # Scoring
    ('Score',           'reseller_score',     8),
    # Origin
    ('Category Site',   'category_site',     18),
    ('Category Leads',  'category_leads',    14),
    ('Campaign',        'campaign',          18),
    # Pipeline marks
    ('Site Mark',       'mark_site_leads',   10),
    ('Leads Mark',      'mark_leads',        10),
    # IDs
    ('Doc ID',          'doc_id',            28),
    ('Lead ID Site',    'lead_id_site',      28),
    ('Lead ID Leads',   'lead_id_leads',     28),
    ('Contact ID',      'contact_id',        28),
    # Lifecycle
    ('Created',         'created_at',        20),
]


def _build_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    write_contacts_sheet(
        ws, rows, COLS,
        sort_key  = lambda r: (int(r.get('tier') or 9), (r.get('company') or '').lower()),
        wrap_keys = {'ai_summary', 'linkedin', 'keywords'},
    )

    # ── Sheet 2: Summary ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20


    # By tier
    make_header_cell(ws2, 1, 1, 'Tier');  _hdr2(1, 2, 'Domains');  _hdr2(1, 3, 'Contacts')
    tier_domains  = {}
    tier_contacts = Counter()
    for r in rows:
        t = int(r.get('tier') or 5)
        tier_domains.setdefault(t, set()).add(r.get('domain', ''))
        tier_contacts[t] += 1
    for ri2, t in enumerate(sorted(tier_contacts), 2):
        ws2.cell(ri2, 1, r.get('tier_label') or f'Tier {t}')
        ws2.cell(ri2, 2, len(tier_domains[t]))
        ws2.cell(ri2, 3, tier_contacts[t])

    # By country
    make_header_cell(ws2, 1, 5, 'Country');  _hdr2(1, 6, 'Contacts')
    cc = Counter(r.get('country', '') for r in rows)
    for ri2, (country, cnt) in enumerate(cc.most_common(), 2):
        ws2.cell(ri2, 5, country)
        ws2.cell(ri2, 6, cnt)

    # By pipeline mark
    make_header_cell(ws2, 1, 8,  'Pipeline');  _hdr2(1, 9, 'Contacts')
    site_only  = sum(1 for r in rows if r.get('mark_site_leads') and not r.get('mark_leads'))
    leads_only = sum(1 for r in rows if r.get('mark_leads') and not r.get('mark_site_leads'))
    both_mark  = sum(1 for r in rows if r.get('mark_site_leads') and r.get('mark_leads'))
    for ri2, (label, cnt) in enumerate([('SITE_LEADS only', site_only),
                                        ('LEADS only',      leads_only),
                                        ('Both pipelines',  both_mark)], 2):
        ws2.cell(ri2, 8,  label)
        ws2.cell(ri2, 9,  cnt)

    # Status breakdown
    make_header_cell(ws2, 1, 11, 'Status');  _hdr2(1, 12, 'Contacts')
    sc = Counter(r.get('status', '') for r in rows)
    for ri2, (s, cnt) in enumerate(sc.most_common(), 2):
        ws2.cell(ri2, 11, s)
        ws2.cell(ri2, 12, cnt)

    save_workbook(wb, out_path, '[ec-export]')

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description='Export email_contacts collection to Excel')
    p.add_argument('--countries', nargs='+', default=None, metavar='CC',
                   help='Country codes e.g. UK NO IN (comma or space separated)')
    p.add_argument('--campaign', default=None, metavar='NAME',
                   help='Filter by campaign tag')
    p.add_argument('--status', default=None, metavar='STATUS',
                   help='Filter by status (pending / approved / sent)')
    p.add_argument('--mark', default=None, choices=['site', 'leads', 'both'],
                   help='Filter by pipeline: site=SITE_LEADS only, leads=LEADS only, both=in both')
    p.add_argument('--collection', default='email_contacts', metavar='NAME',
                   help='Firestore collection (default: email_contacts)')
    p.add_argument('--out', default=None, metavar='PATH',
                   help='Output .xlsx path (default: exports/email_contacts_<cc>_<ts>.xlsx)')
    args = p.parse_args(argv)

    countries = None
    if args.countries:
        raw = []
        for t in args.countries:
            raw.extend(c.strip().upper() for c in t.split(',') if c.strip())
        countries = raw or None

    fb_key = _load_secrets()
    db     = _init_firestore(fb_key)

    rows = _load_data(db,
                      countries  = countries,
                      campaign   = args.campaign,
                      status     = args.status,
                      mark       = args.mark,
                      collection = args.collection)
    if not rows:
        print('[ec-export] No contacts found.')
        return

    if not args.out:
        ts   = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')
        cc   = '_'.join(countries) if countries else 'all'
        tag  = f'_{args.campaign}' if args.campaign else ''
        mark = f'_{args.mark}'     if args.mark     else ''
        out  = Path('exports') / f'email_contacts_{cc}{tag}{mark}_{ts}.xlsx'
    else:
        out = Path(args.out)

    _build_excel(rows, out)

    tc = Counter(r.get('tier_label', 'unknown') for r in rows)
    print('\n[ec-export] Tier breakdown:')
    for label, count in sorted(tc.items()):
        print(f'  {label}: {count} contacts')


if __name__ == '__main__':
    main()
