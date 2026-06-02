"""
gmail_outreach.py — Gmail-based outreach for BlueBoot Lead Agent.

Campaign files live in a single input folder, named:
    {campaign_id}_campaign.xlsx   e.g.  BB-2025-NO-01_campaign.xlsx

Each file has a "Contacts" sheet with the columns below.
Sending updates the file in-place (status, sent_at, thread_id).
Reply-checking scans Gmail and writes back to the matching campaign file.

Setup (one-time):
    pip install google-auth google-auth-oauthlib google-auth-httplib2 google-api-python-client

    1. https://console.cloud.google.com -> enable Gmail API
    2. Create OAuth2 credentials (Desktop app) -> save as config/gmail_credentials.json
    3. First run opens browser for consent -> token saved to config/gmail_token.json

Environment (.env):
    GMAIL_SENDER      leif@blueboot.no
    CAMPAIGN_LABEL    BlueBoot-Outreach
    OUTREACH_DELAY    3.0
    CAMPAIGNS_DIR     campaigns              (folder with *_campaign.xlsx files)

Contacts sheet columns:
    email           required
    company         agency name
    contact_name    person name (optional)
    domain          agency domain
    country         NO/SE/DK/DE/UK/FR/ES
    campaign_id     e.g. BB-2025-NO-01  (auto-set from filename on send)
    subject         email subject
    body            plain-text body
    status          New | Contacted | Replied | Skipped
    sent_at         auto-filled when sent
    thread_id       auto-filled when sent
    notes           free text
"""

from __future__ import annotations

import base64
import os
import time
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Paths / constants
# ---------------------------------------------------------------------------

CREDENTIALS_PATH = Path("config/gmail_credentials.json")
TOKEN_PATH       = Path("config/gmail_token.json")
CONTACTS_SHEET   = "Contacts"
CAMPAIGN_SUFFIX  = "_campaign.xlsx"

SCOPES = [
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.modify",
    "https://www.googleapis.com/auth/gmail.labels",
]


# ---------------------------------------------------------------------------
# Campaign file helpers
# ---------------------------------------------------------------------------

def campaigns_dir() -> Path:
    d = Path(os.getenv("CAMPAIGNS_DIR", "campaigns"))
    d.mkdir(parents=True, exist_ok=True)
    return d


def campaign_id_from_path(p: Path) -> str:
    """BB-2025-NO-01_campaign.xlsx  ->  BB-2025-NO-01"""
    return p.name.removesuffix(CAMPAIGN_SUFFIX)


def campaign_path(campaign_id: str) -> Path:
    return campaigns_dir() / f"{campaign_id}{CAMPAIGN_SUFFIX}"


def discover_campaigns(campaign_filter: str = "") -> list[Path]:
    """Return all *_campaign.xlsx files in the campaigns folder, optionally filtered."""
    folder = campaigns_dir()
    files  = sorted(folder.glob(f"*{CAMPAIGN_SUFFIX}"))
    if campaign_filter:
        files = [f for f in files if campaign_id_from_path(f) == campaign_filter]
    if not files:
        hint = folder / f"{campaign_filter or 'BB-2025-XX-01'}{CAMPAIGN_SUFFIX}"
        print(f"  No campaign files found in {folder}/")
        print(f"  Expected filename pattern: {hint.name}")
    return files


# ---------------------------------------------------------------------------
# Read / write contacts sheet
# ---------------------------------------------------------------------------

def _read_contacts(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=CONTACTS_SHEET, dtype=str)
    return df.fillna("")


def _write_contacts(df: pd.DataFrame, path: Path) -> None:
    """Overwrite the Contacts sheet, preserving all other sheets."""
    wb = load_workbook(path)
    if CONTACTS_SHEET in wb.sheetnames:
        idx = wb.sheetnames.index(CONTACTS_SHEET)
        del wb[CONTACTS_SHEET]
        ws = wb.create_sheet(CONTACTS_SHEET, idx)
    else:
        ws = wb.create_sheet(CONTACTS_SHEET, 0)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    wb.save(path)


# ---------------------------------------------------------------------------
# Gmail auth
# ---------------------------------------------------------------------------

def get_gmail_service():
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError:
        raise ImportError(
            "Run: pip install google-auth google-auth-oauthlib "
            "google-auth-httplib2 google-api-python-client"
        )

    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_PATH.exists():
                raise FileNotFoundError(
                    f"Missing {CREDENTIALS_PATH}\n"
                    "Download OAuth2 credentials from Google Cloud Console."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_PATH), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_PATH.write_text(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def get_or_create_label(service, label_name: str) -> str:
    labels = service.users().labels().list(userId="me").execute().get("labels", [])
    for lbl in labels:
        if lbl["name"].lower() == label_name.lower():
            return lbl["id"]
    result = service.users().labels().create(
        userId="me",
        body={"name": label_name,
              "labelListVisibility": "labelShow",
              "messageListVisibility": "show"},
    ).execute()
    print(f"  Created Gmail label: {label_name}")
    return result["id"]


# ---------------------------------------------------------------------------
# Build MIME message
# ---------------------------------------------------------------------------

def _build_raw(sender: str, to: str, subject: str, body: str,
               campaign_id: str, lead_domain: str) -> dict:
    msg = MIMEMultipart("alternative")
    msg["From"]       = sender
    msg["To"]         = to
    msg["Subject"]    = subject
    msg["X-Campaign"] = campaign_id
    msg["X-Lead"]     = lead_domain
    msg.attach(MIMEText(body, "plain", "utf-8"))
    return {"raw": base64.urlsafe_b64encode(msg.as_bytes()).decode()}


# ---------------------------------------------------------------------------
# send_campaign  — processes one campaign file
# ---------------------------------------------------------------------------

def send_campaign(
    campaign_file: Path,
    dry_run: bool = True,
    delay: float = 3.0,
) -> None:
    """Send to all New rows in one campaign file. Updates file after each send."""
    load_dotenv()
    sender      = os.getenv("GMAIL_SENDER", "leif@blueboot.no")
    label_name  = os.getenv("CAMPAIGN_LABEL", "BlueBoot-Outreach")
    campaign_id = campaign_id_from_path(campaign_file)

    df = _read_contacts(campaign_file)

    # Stamp campaign_id from filename into any blank campaign_id cells
    df.loc[df["campaign_id"].str.strip() == "", "campaign_id"] = campaign_id

    mask = (
        (df["status"].str.strip() == "New") &
        (df["email"].str.strip() != "") &
        (df["subject"].str.strip() != "") &
        (df["body"].str.strip() != "")
    )
    eligible = df[mask]

    print(f"\n{'='*60}")
    print(f"Campaign file : {campaign_file.name}")
    print(f"Campaign ID   : {campaign_id}")
    print(f"Eligible      : {len(eligible)} of {len(df)} contacts")

    if dry_run:
        print("\n[DRY RUN — no emails will be sent]")
        for _, row in eligible.iterrows():
            print(f"  -> {str(row['company']):<25}  <{row['email']}>  {str(row['subject'])[:55]}")
        print(f"\nAdd --send to actually send {len(eligible)} email(s).")
        return

    service  = get_gmail_service()
    label_id = get_or_create_label(service, label_name)
    sent = errors = 0

    for idx, row in eligible.iterrows():
        recipient = str(row["email"]).split(",")[0].strip()
        subject   = str(row["subject"]).strip()
        body      = str(row["body"]).strip()
        domain    = str(row["domain"]).strip()
        company   = str(row["company"]).strip() or domain

        try:
            raw    = _build_raw(sender, recipient, subject, body, campaign_id, domain)
            result = service.users().messages().send(userId="me", body=raw).execute()
            thread = result.get("threadId", "")

            service.users().messages().modify(
                userId="me", id=result["id"],
                body={"addLabelIds": [label_id]},
            ).execute()

            df.at[idx, "status"]      = "Contacted"
            df.at[idx, "campaign_id"] = campaign_id
            df.at[idx, "sent_at"]     = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            df.at[idx, "thread_id"]   = thread

            print(f"  [SENT]  {company:<25} <{recipient}>  thread={thread}")
            sent += 1
            _write_contacts(df, campaign_file)   # persist after every send
            time.sleep(delay)

        except Exception as exc:
            print(f"  [ERROR] {company} <{recipient}>: {exc}")
            errors += 1

    print(f"\nCampaign {campaign_id}: Sent {sent}, Errors {errors}")


# ---------------------------------------------------------------------------
# check_replies  — scans Gmail and routes replies to the right campaign file
# ---------------------------------------------------------------------------

def check_replies(campaign_filter: str = "") -> None:
    """
    Fetch all labelled threads from Gmail.
    For each thread with >= 2 messages (i.e. a reply exists):
      - Identify campaign_id from the X-Campaign header
      - Load the matching *_campaign.xlsx file
      - Mark the matching row as "Replied"
    """
    load_dotenv()
    label_name = os.getenv("CAMPAIGN_LABEL", "BlueBoot-Outreach")

    service  = get_gmail_service()
    label_id = get_or_create_label(service, label_name)

    # Load all campaign files into memory: campaign_id -> (path, df)
    all_files = discover_campaigns(campaign_filter)
    if not all_files:
        return

    campaign_data: dict[str, tuple[Path, pd.DataFrame]] = {}
    for f in all_files:
        cid = campaign_id_from_path(f)
        campaign_data[cid] = (f, _read_contacts(f))

    # Build thread_id -> (campaign_id, df_index) lookup
    thread_map: dict[str, tuple[str, int]] = {}
    for cid, (_, df) in campaign_data.items():
        for idx, row in df.iterrows():
            tid = str(row.get("thread_id", "")).strip()
            if tid:
                thread_map[tid] = (cid, idx)

    # Fetch labelled threads from Gmail
    label_slug = label_name.replace(" ", "-")
    query      = f"label:{label_slug}"
    threads: list[dict] = []
    page_token = None
    while True:
        kw: dict = {"userId": "me", "q": query, "maxResults": 100}
        if page_token:
            kw["pageToken"] = page_token
        result     = service.users().threads().list(**kw).execute()
        threads   += result.get("threads", [])
        page_token = result.get("nextPageToken")
        if not page_token:
            break

    print(f"Found {len(threads)} labelled thread(s) in Gmail")
    dirty: set[str] = set()   # campaign_ids whose files need saving

    for stub in threads:
        thread = service.users().threads().get(
            userId="me", id=stub["id"], format="metadata",
            metadataHeaders=["X-Campaign", "X-Lead", "From", "To"],
        ).execute()

        messages = thread.get("messages", [])
        if len(messages) < 2:
            continue   # no reply yet

        tid = stub["id"]

        # -- Match by thread_id (most reliable) --
        if tid in thread_map:
            cid, idx = thread_map[tid]
            _, df = campaign_data[cid]
            if df.at[idx, "status"] != "Replied":
                df.at[idx, "status"] = "Replied"
                df.at[idx, "notes"]  = str(df.at[idx, "notes"]).rstrip() + " | REPLIED"
                print(f"  Reply [{cid}]: {df.at[idx, 'company']} <{df.at[idx, 'email']}>")
                dirty.add(cid)
            continue

        # -- Fallback: match by X-Campaign + X-Lead headers --
        hdrs = {h["name"]: h["value"]
                for h in messages[0].get("payload", {}).get("headers", [])}
        x_campaign = hdrs.get("X-Campaign", "").strip()
        x_lead     = hdrs.get("X-Lead", "").strip()

        if x_campaign not in campaign_data:
            continue

        _, df = campaign_data[x_campaign]
        if x_lead:
            mask = df["domain"].str.lower() == x_lead.lower()
