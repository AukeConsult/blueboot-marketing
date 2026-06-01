"""site_contact_export.py -- Export site_contacts to Excel, enriched with site data.

Reads every document from the site_contacts collectionGroup
(site_leads/{lead_id}/site_contacts/{contact_id}), joins in key fields from
the parent site_leads document, and writes one Excel row per contact.

Each row contains:
  Contact:  name, email, phone, title, occupation, company, linkedin, twitter, facebook
  Site:     domain, website, country, page_count, ai_sector, ai_company_type,
            ai_platform, ai_hosting, ai_summary, ai_confidence, query_category,
            ai_keywords, crawled_at, found_on

This makes it easy to filter by country, sector, platform, or any other field
to find the most relevant contacts for outreach.

Usage:
    python app/site_contact_export.py
    python app/site_contact_export.py --countries NO,SE
    python app/site_contact_export.py --countries NO --sector ecommerce
    python app/site_contact_export.py --countries NO --with-email-only
    python app/site_contact_export.py --output exports/contacts_no.xlsx
    python app/site_contact_export.py --limit 2000
"""
from __future__ import annotations

import argparse
import importlib.util
import os
from datetime import datetime, timezone
from pathlib import Path

import _pathsetup  # noqa: F401

# ---------------------------------------------------------------------------
# Page size buckets (shared with statistics.py)
# ---------------------------------------------------------------------------

PAGE_SIZE_BUCKETS = {
    "micro":   (1,      50),
    "small":   (51,     500),
    "medium":  (501,    3000),
    "large":   (3001,   10000),
    "huge":    (10001,  100000),
    "ultra":   (100001, 999999999),
}

def _page_count_bucket(page_count) -> str:
    """Return the bucket name for a page_count value."""
    try:
        pc = int(page_count or 0)
    except (TypeError, ValueError):
        pc = 0
    if pc == 0:
        return "unknown"
    for name, (lo, hi) in PAGE_SIZE_BUCKETS.items():
        if lo <= pc <= hi:
            return name
    return "ultra"

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

LEADS_COLLECTION    = "site_leads"
CONTACTS_COLLECTION = "site_contacts"

# Columns in output order: (contact_field, header, width)
CONTACT_COLUMNS: list[tuple[str, str, int]] = [
    ("_doc_id",          "Doc ID",           28),
    ("_site_doc_id",     "Site Doc ID",      28),
    ("name",             "Name",             28),
    ("email",            "Email",            32),
    ("phone",            "Phone",            18),
    ("title",            "Title (scraped)",  28),
    ("occupation",       "Occupation",       28),
    ("company",          "Company",          28),
    ("linkedin",         "LinkedIn",         40),
    ("twitter",          "Twitter",          30),
    ("facebook",         "Facebook",         30),
    ("ai_country",       "AI Country",       12),
]

LEAD_COLUMNS: list[tuple[str, str, int]] = [
    ("domain",           "Domain",           28),
    ("website",          "Website",          34),
    ("country",          "Country",           8),
    ("country_name",     "Country Name",     16),
    ("ai_country",       "AI Country",       12),
    ("query_category",   "Category",         16),
    ("page_count",       "Pages",             9),
    ("ai_sector",        "AI Sector",        18),
    ("ai_company_type",  "AI Type",          14),
    ("ai_platform",      "Platform",         16),
    ("ai_hosting",       "Hosting",          16),
    ("ai_confidence",    "AI Conf",           9),
    ("ai_summary",       "AI Summary",       45),
    ("ai_keywords",      "AI Keywords",      45),
    ("crawled_at",       "Crawled At",       20),
    ("found_on",         "Found On",         30),
]

# ---------------------------------------------------------------------------
# Secrets / Firestore
# ---------------------------------------------------------------------------

def _load_secrets():
    secrets_path = Path(__file__).parent.parent / "blueboot_secrets.py"
    if not secrets_path.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location("blueboot_secrets", secrets_path)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return getattr(mod, "fireBaseAdminKey", None)
    except Exception as e:
        print(f"  [contact-export] could not load blueboot_secrets: {e}")
        return None


def _init_firestore(fb_key_dict):
    try:
        import firebase_admin
        from firebase_admin import firestore
        import firebase_admin.credentials as fb_creds
    except ImportError:
        raise RuntimeError("firebase-admin not installed — run: pip install firebase-admin")
    cred = (fb_creds.Certificate(fb_key_dict) if fb_key_dict
            else fb_creds.Certificate(os.getenv("FIREBASE_CREDENTIALS",
                                                "config/serviceAccountKey.json")))
    if not firebase_admin._apps:
        firebase_admin.initialize_app(cred)
    return firestore.client()


# ---------------------------------------------------------------------------
# Firestore fetch
# ---------------------------------------------------------------------------

def _load_leads_index(db, countries: list[str] | None) -> dict[str, dict]:
    """Load all site_leads into a dict keyed by lead_id for fast join."""
    print("  [contact-export] Loading site_leads index…", flush=True)
    col   = db.collection(LEADS_COLLECTION)
    index = {}
    for doc in col.stream():
        data = doc.to_dict() or {}
        if countries:
            ai_c = (data.get("ai_country") or "").upper()
            if ai_c not in countries:
                continue
        index[doc.id] = data
    print(f"  [contact-export] {len(index)} leads loaded into index", flush=True)
    return index


def _stream_contacts(
    db,
    leads_index:     dict[str, dict],
    with_email_only: bool,
    sector:          str | None,
    category:        str | None,
    location:        str | None,
    limit:           int | None,
    countries:       list[str] | None = None,
    page_count:       str | None       = None,
) -> tuple[list[dict], dict]:
    """Stream site_contacts collectionGroup, join with parent lead data.

    Returns (rows, used_leads) where used_leads contains only the site_leads
    that have at least one contact in the result set.
    """
    print(f"  [contact-export] Scanning collectionGroup '{CONTACTS_COLLECTION}'…", flush=True)
    col        = db.collection_group(CONTACTS_COLLECTION)
    rows       = []
    used_leads: dict[str, dict] = {}
    scanned = skipped = 0

    for doc in col.stream():
        scanned += 1
        contact = doc.to_dict() or {}

        # Must have a name
        if not (contact.get("name") or "").strip():
            skipped += 1
            continue

        if with_email_only and not (contact.get("email") or "").strip():
            skipped += 1
            continue

        # Get parent lead_id from the doc path: site_leads/{lead_id}/site_contacts/{id}
        lead_id = doc.reference.parent.parent.id
        lead    = leads_index.get(lead_id, {})

        # If countries filter active, skip contacts whose parent lead isn't in index
        if not lead and leads_index:
            skipped += 1
            continue

        # Contact-level country filter: match on lead's ai_country only
        if countries:
            ai_c = (lead.get("ai_country") or "").upper()
            if ai_c not in countries:
                skipped += 1
                continue

        # Sector filter
        if sector and (lead.get("ai_sector") or "").lower() != sector.lower():
            skipped += 1
            continue

        # Category filter
        if category and (lead.get("query_category") or "").lower() != category.lower():
            skipped += 1
            continue

        # Page size filter
        if page_count:
            bucket = _page_count_bucket(lead.get("page_count"))
            if bucket != page_count.lower():
                skipped += 1
                continue

        # Location filter (keyword search in location_full)
        if location:
            loc_full = (lead.get("location_full") or lead.get("location") or "").lower()
            if location.lower() not in loc_full:
                skipped += 1
                continue

        # Merge contact + lead fields into one flat row
        row = dict(contact)
        row["_doc_id"]      = doc.id
        row["_site_doc_id"] = lead_id
        row["_lead_id"]     = lead_id
        # Copy lead fields (prefix collision avoided — lead fields go under their own keys)
        for key in [c[0] for c in LEAD_COLUMNS]:
            if key not in row:  # don't overwrite contact's own field
                row[key] = lead.get(key, "")
        # ai_country always comes from the parent site_lead, never from the contact doc
        row["ai_country"] = (lead.get("ai_country") or "").upper()
        # found_on may be on contact, not lead
        if "found_on" not in contact:
            row["found_on"] = ""

        # Normalise ai_keywords list → comma string
        kw = row.get("ai_keywords")
        if isinstance(kw, list):
            row["ai_keywords"] = ", ".join(kw)

        rows.append(row)
        used_leads[lead_id] = lead

        if len(rows) % 200 == 0:
            print(f"  [contact-export] … {len(rows)} contacts collected (scanned {scanned})", flush=True)

        if limit and len(rows) >= limit:
            break

    print(
        f"  [contact-export] {scanned} scanned → {len(rows)} rows  "
        f"({skipped} skipped by filter)  "
        f"({len(used_leads)} unique sites)",
        flush=True,
    )
    return rows, used_leads


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

def _build_xlsx(rows: list[dict], output_path: str, leads_index: dict | None = None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    wb = Workbook()

    # ── Contacts sheet ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Contacts"

    all_columns = CONTACT_COLUMNS + LEAD_COLUMNS
    headers     = [h for _, h, _ in all_columns]
    widths      = [w for _, _, w in all_columns]
    fields      = [f for f, _, _ in all_columns]

    # Header style
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", start_color="1F497D")
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin        = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write headers
    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Alternating row fill
    fill_even = PatternFill("solid", start_color="EEF2F7")
    fill_odd  = PatternFill("solid", start_color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    data_align_wrap = Alignment(vertical="top", wrap_text=True)
    data_align      = Alignment(vertical="top", wrap_text=False)

    # Contact section separator — light blue left border for first lead column
    lead_start_col = len(CONTACT_COLUMNS) + 1
    sep_border = Border(left=Side(style="medium", color="1F497D"),
                        right=thin, top=thin, bottom=thin)

    for row_idx, row in enumerate(rows, start=2):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        for col_idx, field in enumerate(fields, start=1):
            val = row.get(field, "") or ""
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            if isinstance(val, float):
                val = round(val, 3)
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = data_align_wrap if col_idx in (2, 16) else data_align
            cell.border    = sep_border if col_idx == lead_start_col else cell_border

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_columns))}1"

    # ── Summary sheet ───────────────────────────────────────────────────────
    ws2          = wb.create_sheet("Summary")
    ws2["A1"]    = "Site Contact Export"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)
    summary_rows = [
        ("Generated at",  datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Total contacts", len(rows)),
        ("With email",    sum(1 for r in rows if r.get("email"))),
        ("With LinkedIn", sum(1 for r in rows if r.get("linkedin"))),
        ("With phone",    sum(1 for r in rows if r.get("phone"))),
        ("Enriched (Brave)", sum(1 for r in rows if r.get("brave_enriched_at"))),
    ]
    for i, (label, value) in enumerate(summary_rows, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
        ws2.cell(row=i, column=2, value=value).font  = Font(name="Arial", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 30

    # Country breakdown — by ai_country from site_lead
    ws2.cell(row=10, column=1, value="By AI Country").font = Font(name="Arial", bold=True, size=10)
    country_counts: dict[str, int] = {}
    for r in rows:
        c = (r.get("ai_country") or "?").upper()
        country_counts[c] = country_counts.get(c, 0) + 1
    for i, (c, cnt) in enumerate(sorted(country_counts.items()), start=11):
        ws2.cell(row=i, column=1, value=c).font   = Font(name="Arial", size=9)
        ws2.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=9)

    # Sector breakdown
    sector_start = 11 + len(country_counts) + 2
    ws2.cell(row=sector_start, column=1, value="By AI Sector").font = Font(name="Arial", bold=True, size=10)
    sector_counts: dict[str, int] = {}
    for r in rows:
        s = r.get("ai_sector") or "unknown"
        sector_counts[s] = sector_counts.get(s, 0) + 1
    for i, (s, cnt) in enumerate(sorted(sector_counts.items(), key=lambda x: -x[1]), start=sector_start + 1):
        ws2.cell(row=i, column=1, value=s).font   = Font(name="Arial", size=9)
        ws2.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=9)

    # ── Sites sheet ─────────────────────────────────────────────────────────
    if leads_index:
        SITE_COLUMNS: list[tuple[str, str, int]] = [
            ("_doc_id",          "Doc ID",           28),
            ("domain",           "Domain",           28),
            ("website",          "Website",          34),
            ("country",          "Country",           8),
            ("country_name",     "Country Name",     16),
            ("ai_country",       "AI Country",       12),
            ("query_category",   "Category",         16),
            ("page_count",       "Pages",             9),
            ("ai_sector",        "AI Sector",        18),
            ("ai_company_type",  "AI Type",          14),
            ("ai_platform",      "Platform",         16),
            ("ai_hosting",       "Hosting",          16),
            ("ai_confidence",    "AI Conf",           9),
            ("ai_summary",       "AI Summary",       45),
            ("ai_keywords",      "AI Keywords",      45),
            ("crawled_at",       "Crawled At",       20),
        ]
        ws_sites = wb.create_sheet("Sites")
        site_fields  = [f for f, _, _ in SITE_COLUMNS]
        site_headers = [h for _, h, _ in SITE_COLUMNS]
        site_widths  = [w for _, _, w in SITE_COLUMNS]

        for col_idx, (header, width) in enumerate(zip(site_headers, site_widths), start=1):
            cell = ws_sites.cell(row=1, column=col_idx, value=header)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            cell.border    = cell_border
            ws_sites.column_dimensions[get_column_letter(col_idx)].width = width
        ws_sites.row_dimensions[1].height = 20
        ws_sites.freeze_panes = "A2"

        for row_idx, (lead_id, lead) in enumerate(sorted(leads_index.items()), start=2):
            fill = fill_even if row_idx % 2 == 0 else fill_odd
            lead_row = dict(lead)
            lead_row["_doc_id"] = lead_id
            kw = lead_row.get("ai_keywords")
            if isinstance(kw, list):
                lead_row["ai_keywords"] = ", ".join(kw)
            for col_idx, field in enumerate(site_fields, start=1):
                val = lead_row.get(field, "") or ""
                if isinstance(val, list):
                    val = ", ".join(str(v) for v in val)
                if isinstance(val, float):
                    val = round(val, 3)
                cell           = ws_sites.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = data_align_wrap if field == "ai_summary" else data_align
                cell.border    = cell_border
        ws_sites.auto_filter.ref = f"A1:{get_column_letter(len(SITE_COLUMNS))}1"

        # ── Sites Summary sheet ──────────────────────────────────────────────
        ws_ss = wb.create_sheet("Sites Summary")
        ws_ss["A1"]       = "Site Leads Export"
        ws_ss["A1"].font  = Font(name="Arial", bold=True, size=14)
        ws_ss.column_dimensions["A"].width = 22
        ws_ss.column_dimensions["B"].width = 30

        site_summary_rows = [
            ("Generated at",     datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total sites",      len(leads_index)),
            ("With AI analysis", sum(1 for l in leads_index.values() if l.get("ai_classified_at"))),
        ]
        for i, (label, value) in enumerate(site_summary_rows, start=3):
            ws_ss.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
            ws_ss.cell(row=i, column=2, value=value).font = Font(name="Arial", size=10)

        # Country breakdown (sites) — by ai_country only
        ws_ss.cell(row=8, column=1, value="By AI Country").font = Font(name="Arial", bold=True, size=10)
        site_country: dict[str, int] = {}
        for l in leads_index.values():
            c = (l.get("ai_country") or "?").upper()
            site_country[c] = site_country.get(c, 0) + 1
        for i, (c, cnt) in enumerate(sorted(site_country.items()), start=9):
            ws_ss.cell(row=i, column=1, value=c).font   = Font(name="Arial", size=9)
            ws_ss.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=9)

        # Sector breakdown (sites)
        site_sector_start = 9 + len(site_country) + 2
        ws_ss.cell(row=site_sector_start, column=1, value="By AI Sector").font = Font(name="Arial", bold=True, size=10)
        site_sector: dict[str, int] = {}
        for l in leads_index.values():
            s = l.get("ai_sector") or "unknown"
            site_sector[s] = site_sector.get(s, 0) + 1
        for i, (s, cnt) in enumerate(sorted(site_sector.items(), key=lambda x: -x[1]), start=site_sector_start + 1):
            ws_ss.cell(row=i, column=1, value=s).font   = Font(name="Arial", size=9)
            ws_ss.cell(row=i, column=2, value=cnt).font = Font(name="Arial", size=9)

    wb.save(output_path)
    print(f"  [contact-export] Saved → {output_path}", flush=True)


# ---------------------------------------------------------------------------
# Campaign save
# ---------------------------------------------------------------------------

SITE_CAMPAIGNS_COLLECTION = "site_campaigns"

def _load_taken_sites(db) -> dict[str, str]:
    """Query site_campaign_sites collectionGroup to find all lead_ids already in a campaign.

    Returns {lead_id: campaign_name} for every site already claimed.
    One upfront scan — no per-site reads needed.
    """
    print("  [campaign] Scanning site_campaign_sites for already-claimed sites…", flush=True)
    taken: dict[str, str] = {}
    for doc in db.collection_group("site_campaign_sites").stream():
        lead_id  = doc.id
        campaign = doc.reference.parent.parent.id
        taken[lead_id] = campaign
    print(f"  [campaign] {len(taken)} sites already in campaigns", flush=True)
    return taken


def _save_campaign(
    db,
    campaign:        str,
    rows:            list[dict],
    used_leads:      dict,
    countries:       list[str] | None = None,
    sector:          str | None       = None,
    category:        str | None       = None,
    with_email_only: bool             = False,
    limit:           int | None       = None,
    force:           bool             = False,
) -> None:
    """Copy only the filtered sites and contacts into site_campaigns/{campaign}/.

    Checks site_campaign_sites collectionGroup first — any site already in another
    campaign is skipped (unless --force is set).

    Structure:
      site_campaigns/{campaign}/
          site_campaign_sites/{lead_id}/
              site_campaign_contacts/{contact_id}
    """

    total_sites    = len(used_leads)
    total_contacts = len(rows)

    print(
        f"  [campaign] Saving '{campaign}' → "
        f"{total_sites} sites  {total_contacts} contacts…",
        flush=True,
    )

    # ── Check which sites are already claimed ────────────────────────────────
    if not force:
        taken = _load_taken_sites(db)
        # Remove this campaign's own entries so re-runs don't self-block
        taken = {lid: c for lid, c in taken.items() if c != campaign}
    else:
        taken = {}
        print("  [campaign] --force: skipping duplicate check", flush=True)

    # Build contact lookup: site_doc_id → [contact rows]
    contacts_by_site: dict[str, list[dict]] = {}
    for row in rows:
        sid = row.get("_site_doc_id", "unknown")
        contacts_by_site.setdefault(sid, []).append(row)

    camp_ref = db.collection(SITE_CAMPAIGNS_COLLECTION).document(campaign)

    # ── Campaign summary doc ─────────────────────────────────────────────────
    camp_ref.set({
        "campaign_id":    campaign,
        "created_at":     datetime.now(timezone.utc).isoformat(),
        "site_count":     total_sites,
        "contact_count":  total_contacts,
        "filters": {
            "countries":       countries or [],
            "sector":          sector or "",
            "category":        category or "",
            "with_email_only": with_email_only,
            "limit":           limit,
        },
    }, merge=True)

    # ── Write sites + contacts ───────────────────────────────────────────────
    sites_written    = 0
    sites_skipped    = 0
    contacts_written = 0

    for lead_id, lead in used_leads.items():
        # Skip sites already in another campaign
        if lead_id in taken:
            sites_skipped += 1
            continue

        data = {k: v for k, v in lead.items() if not k.startswith("_")}
        kw = data.get("ai_keywords")
        if isinstance(kw, list):
            data["ai_keywords"] = ", ".join(kw)

        site_ref = camp_ref.collection("site_campaign_sites").document(lead_id)
        site_ref.set(data, merge=True)
        sites_written += 1

        for row in contacts_by_site.get(lead_id, []):
            contact_data = {k: v for k, v in row.items() if not k.startswith("_")}
            doc_id = row.get("_doc_id") or "unknown"
            site_ref.collection("site_campaign_contacts").document(doc_id).set(contact_data, merge=True)
            contacts_written += 1

        if sites_written % 50 == 0:
            print(
                f"  [campaign] sites {sites_written}/{total_sites - sites_skipped}  "
                f"contacts {contacts_written}/{total_contacts}",
                flush=True,
            )

    if sites_skipped:
        # Summarise which campaigns the skipped sites belong to
        from collections import Counter
        skip_sources = Counter(
            taken[lid] for lid in used_leads if lid in taken
        )
        skip_detail = "  ".join(f"{c}: {n}" for c, n in skip_sources.most_common())
        print(
            f"  [campaign] {sites_skipped} sites skipped (already in other campaigns): {skip_detail}",
            flush=True,
        )
    print(
        f"  [campaign] Done → site_campaigns/{campaign}  "
        f"{sites_written} sites saved  {sites_skipped} skipped  {contacts_written} contacts",
        flush=True,
    )


def _load_campaign_data(db, campaign: str) -> tuple[list[dict], dict]:
    """Read sites and contacts back from a saved campaign.

    Returns (rows, used_leads) in the same shape as _stream_contacts so
    _build_xlsx can consume them directly.
    """
    print(f"  [campaign] Loading data from site_campaigns/{campaign}…", flush=True)

    camp_ref   = db.collection(SITE_CAMPAIGNS_COLLECTION).document(campaign)
    sites_col  = camp_ref.collection("site_campaign_sites")

    used_leads: dict[str, dict] = {}
    rows:       list[dict]      = []

    for site_doc in sites_col.stream():
        lead_id  = site_doc.id
        lead     = site_doc.to_dict() or {}
        used_leads[lead_id] = lead

        for contact_doc in site_doc.reference.collection("site_campaign_contacts").stream():
            contact = contact_doc.to_dict() or {}
            # Re-attach internal keys so _build_xlsx columns resolve correctly
            contact["_doc_id"]      = contact_doc.id
            contact["_site_doc_id"] = lead_id
            contact["_lead_id"]     = lead_id
            # Ensure ai_country is present
            if "ai_country" not in contact:
                contact["ai_country"] = lead.get("ai_country", "")
            rows.append(contact)

    print(
        f"  [campaign] Loaded {len(used_leads)} sites  {len(rows)} contacts "
        f"from campaign '{campaign}'",
        flush=True,
    )
    return rows, used_leads


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def export_contacts(
    countries:       list[str] | None = None,
    sector:          str | None       = None,
    category:        str | None       = None,
    location:        str | None       = None,
    with_email_only: bool             = False,
    limit:           int | None       = None,
    output_path:     str | None       = None,
    campaign:        str | None       = None,
    force:           bool             = False,
    page_count:       str | None       = None,
) -> str:
    fb_key = _load_secrets()
    db     = _init_firestore(fb_key)

    # ── Step 1: filter from site_leads + site_contacts ───────────────────────
    leads_index = _load_leads_index(db, countries)
    rows, used_leads = _stream_contacts(
        db, leads_index, with_email_only, sector, category, location, limit,
        countries=countries, page_count=page_count
    )
    if not rows:
        print("  [contact-export] No contacts found matching filters.")
        return ""

    # ── Step 2: if campaign set, save first then reload from campaign ─────────
    if campaign:
        _save_campaign(
            db, campaign, rows, used_leads,
            countries       = countries,
            sector          = sector,
            category        = category,
            with_email_only = with_email_only,
            limit           = limit,
            force           = force,
        )
        # Reload rows and leads from what was actually saved to the campaign
        # (some sites may have been skipped as duplicates)
        rows, used_leads = _load_campaign_data(db, campaign)
        if not rows:
            print(f"  [contact-export] No data in campaign '{campaign}' to export.")
            return ""

    # ── Step 3: build output path ─────────────────────────────────────────────
    if not output_path:
        ts      = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
        parts   = []
        if campaign:
            parts.append(campaign)
        else:
            if countries:
                parts.append("_".join(countries))
            if sector:
                parts.append(sector)
            if category:
                parts.append(category)
            if with_email_only:
                parts.append("email")
        filter_str  = ("_" + "_".join(parts)) if parts else ""
        output_path = str(
            Path(__file__).parent.parent / "exports" / f"site_contacts{filter_str}_{ts}.xlsx"
        )

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # ── Step 4: export to Excel ───────────────────────────────────────────────
    print(f"\n  [contact-export] Building Excel — {len(rows)} contacts…", flush=True)
    _build_xlsx(rows, output_path, leads_index=used_leads)

    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> None:
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    p = argparse.ArgumentParser(
        description="Export site_contacts to Excel, enriched with site_leads data"
    )
    p.add_argument("--countries",       default=None, metavar="CODES",
                   help="Comma-separated country codes  e.g. NO,SE")
    p.add_argument("--sector",          default=None, metavar="SECTOR",
                   help="Filter by ai_sector  e.g. ecommerce, technology")
    p.add_argument("--category",        default=None, metavar="CATEGORY",
                   help="Filter by query_category  e.g. real_estate, healthcare")
    p.add_argument("--location",        default=None, metavar="TEXT",
                   help="Filter by location keyword e.g. London, Pune, Manchester")
    p.add_argument("--with-email-only", action="store_true",
                   help="Only include contacts that have an email address")
    p.add_argument("--limit",           type=int, default=None, metavar="N",
                   help="Max contacts to export")
    p.add_argument("--output",          default=None, metavar="FILE",
                   help="Output .xlsx path  (default: exports/contacts_<country>_<ts>.xlsx)")
    p.add_argument("--campaign",        default=None, metavar="NAME",
                   help="Save selection to site_campaigns/<NAME> in Firestore (sites + contacts)")
    p.add_argument("--force",            action="store_true",
                   help="Re-assign sites already in another campaign (skips duplicate check)")
    p.add_argument("--page-count",       default=None, metavar="BUCKET",
                   help="Filter by page count bucket: micro/small/medium/large/huge/ultra/unknown")

    args = p.parse_args(argv)

    countries = None
    if args.countries:
        countries = [c.strip().upper() for c in args.countries.split(",") if c.strip()]

    path = export_contacts(
        countries       = countries,
        sector          = args.sector,
        category        = args.category,
        location        = args.location,
        with_email_only = args.with_email_only,
        limit           = args.limit,
        output_path     = args.output,
        campaign        = args.campaign,
        force           = args.force,
        page_count       = args.page_count,
    )
    if path:
        print(f"\n  Done → {path}")


if __name__ == "__main__":
    main()
