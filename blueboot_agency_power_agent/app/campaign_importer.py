# app/campaign_importer.py

from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

from app.firestore_client import get_firestore


LEAD_UPDATABLE_FIELDS = {
    "company",
    "priority",
    "status",
    "notes",
    "suggested_angle",
}

CONTACT_UPDATABLE_FIELDS = {
    "name",
    "title",
    "phone",
    "linkedin",
}


LEAD_HEADER_MAP = {
    "Company": "company",
    "Priority": "priority",
    "Status": "status",
    "Notes": "notes",
    "Suggested Angle": "suggested_angle",
}

CONTACT_HEADER_MAP = {
    "Name": "name",
    "Title": "title",
    "Phone": "phone",
    "LinkedIn": "linkedin",
}


def _normalize(value):
    if value is None:
        return ""

    return str(value).strip()


def _load_workbook(excel_file):
    return load_workbook(
        excel_file,
        data_only=True,
    )


def _validate_campaign(wb, campaign_id):
    ws = wb["Campaign"]

    headers = [c.value for c in ws[1]]
    values = [c.value for c in ws[2]]
    row = dict(zip(headers, values))

    workbook_campaign_id = _normalize(row.get("Extract ID"))

    if workbook_campaign_id != campaign_id:
        raise ValueError(
            f"Campaign mismatch. "
            f"Workbook={workbook_campaign_id} "
            f"Argument={campaign_id}"
        )


def _import_leads(wb, campaign_ref, dry_run):
    ws = wb["Leads"]

    headers = [c.value for c in ws[1]]

    lead_updates = 0
    lead_skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        lead_id = _normalize(data.get("Lead ID"))

        if not lead_id:
            continue

        lead_ref = (campaign_ref.collection("leads_extracted").document(lead_id))
        snap = lead_ref.get()

        if not snap.exists:
            lead_skipped += 1
            continue

        firestore_data = (snap.to_dict() or {})
        updates = {}

        for excel_col, field in (LEAD_HEADER_MAP.items()):
            if field not in LEAD_UPDATABLE_FIELDS:
                continue

            excel_value = _normalize(data.get(excel_col))

            firestore_value = _normalize(firestore_data.get(field))

            if excel_value != firestore_value:
                updates[field] = excel_value

        if updates:
            print(f"Lead update: {lead_id} -> {list(updates.keys())}")
            if not dry_run:
                lead_ref.update(updates)

            lead_updates += 1

    return (
        lead_updates,
        lead_skipped,
    )


def _import_contacts(wb, campaign_ref, dry_run):
    ws = wb["Contacts"]

    headers = [c.value for c in ws[1]]
    contact_updates = 0
    contact_skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        lead_id = _normalize(data.get("Lead ID"))

        contact_id = _normalize(data.get("Contact ID"))

        if not lead_id:
            continue

        if not contact_id:
            continue

        contact_ref = (
            campaign_ref
            .collection("leads_extracted")
            .document(lead_id)
            .collection(
                "contacts_extracted"
            )
            .document(contact_id)
        )

        snap = contact_ref.get()

        if not snap.exists:
            contact_skipped += 1
            continue

        firestore_data = (snap.to_dict() or {})

        updates = {}

        for excel_col, field in (CONTACT_HEADER_MAP.items()):
            if (
                field
                not in
                CONTACT_UPDATABLE_FIELDS
            ):
                continue

            excel_value = _normalize(data.get(excel_col))

            firestore_value = _normalize(firestore_data.get(field))

            if excel_value != firestore_value:
                updates[field] = excel_value

        if updates:
            print(f"Contact update: {contact_id} -> {list(updates.keys())}")
            if not dry_run:
                contact_ref.update(updates)

            contact_updates += 1

    return (
        contact_updates,
        contact_skipped,
    )


def import_campaign(campaign_id: str, excel_file: str, dry_run: bool = False):
    excel_file = Path(excel_file)
    print(f"Import file: {excel_file.resolve()}")

    if not excel_file.exists():
        raise FileNotFoundError(excel_file)

    wb = _load_workbook(excel_file)

    _validate_campaign(wb, campaign_id)

    db = get_firestore()

    campaign_ref = (db.collection("leads_extract").document(campaign_id))

    campaign_snap = (campaign_ref.get())

    if not campaign_snap.exists:
        raise ValueError(
            f"Campaign not found: "
            f"{campaign_id}"
        )

    (
        lead_updates,
        lead_skipped,
    ) = _import_leads(
        wb,
        campaign_ref,
        dry_run,
    )

    (
        contact_updates,
        contact_skipped,
    ) = _import_contacts(
        wb,
        campaign_ref,
        dry_run,
    )

    result = {
        "campaign_id": campaign_id,
        "lead_updates": lead_updates,
        "contact_updates": contact_updates,
        "lead_skipped": lead_skipped,
        "contact_skipped": contact_skipped,
        "dry_run": dry_run,
    }

    print(result)

    return result

if __name__ == "__main__":
    import_campaign(
        campaign_id="NO_high_score_may26",
        excel_file="output/NO_high_score_may26/campaign.xlsx",
        dry_run=True,
    )
