# app/campaign_exporter.py

from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from app.firestore_client import get_firestore

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
CAMPAIGN_COLUMNS = [
    ("extract_id", "Extract ID", 30),
    ("created_at", "Created At", 24),
    ("lead_count", "Lead Count", 14),
    ("contact_count", "Contact Count", 14),
    ("countries", "Countries", 30),
    ("keywords", "Keywords", 40),
    ("min_score", "Min Score", 12),
    ("max_score", "Max Score", 12),
]

LEAD_COLUMNS = [
    ("lead_id", "Lead ID", 24),
    ("domain", "Domain", 28),
    ("company", "Company", 28),
    ("website", "Website", 36),
    ("country", "Country", 10),
    ("priority", "Priority", 18),
    ("status", "Status", 18),
    ("notes", "Notes", 50),
    ("suggested_angle", "Suggested Angle", 60),
    ("source_query", "Source Query", 30),
    ("reseller_score", "Reseller Score", 14),
]

CONTACT_COLUMNS = [
    ("contact_id", "Contact ID", 16),
    ("lead_id", "Lead ID", 24),
    ("email", "Email", 35),
    ("name", "Name", 24),
    ("title", "Title", 30),
    ("phone", "Phone", 20),
    ("linkedin", "LinkedIn", 35),
    ("website", "Website", 35),
]

# ---------------------------------------------------------------------------
# Firestore
# ---------------------------------------------------------------------------
def _load_campaign(campaign_id: str):
    db = get_firestore()

    campaign_ref = (db.collection("leads_extract").document(campaign_id))
    snap = campaign_ref.get()

    if not snap.exists:
        raise ValueError(f"Campaign not found: {campaign_id}")

    return campaign_ref, snap.to_dict()


def _load_leads_and_contacts(campaign_ref):
    leads = []
    contacts = []

    lead_docs = (campaign_ref.collection("leads_extracted").stream())
    for lead_doc in lead_docs:
        lead = lead_doc.to_dict() or {}
        lead["lead_id"] = (lead.get("lead_id") or lead_doc.id)
        leads.append(lead)
        contact_docs = (lead_doc.reference.collection("contacts_extracted").stream())

        for contact_doc in contact_docs:
            contact = contact_doc.to_dict() or {}
            contact["contact_id"] = contact_doc.id
            contact["lead_id"] = (contact.get("lead_id") or lead["lead_id"])
            contacts.append(contact)

    leads.sort(key=lambda x: x.get("lead_id", ""))

    contacts.sort(key=lambda x: (x.get("lead_id", ""), x.get("contact_id", "")))

    return leads, contacts

# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------
def _json_default(obj):
    try:
        return obj.isoformat()
    except Exception:
        return str(obj)


def _write_json(output_dir: Path, campaign: dict, leads: list, contacts: list):
    json_file = output_dir / f"campaign.json"
    print (json_file)
    payload = {
        "schema_version": 1,
        "campaign": campaign,
        "leads": leads,
        "contacts": contacts,
    }

    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=_json_default)

    return json_file

# ---------------------------------------------------------------------------
# XLSX helpers
# ---------------------------------------------------------------------------
def _styles():
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)

    return {
        "header_font": Font(
            name="Arial",
            bold=True,
            color="FFFFFF",
            size=10,
        ),
        "header_fill": PatternFill(
            "solid",
            start_color="2D5C8E",
        ),
        "header_align": Alignment(
            horizontal="center",
            vertical="center",
        ),
        "cell_font": Font(
            name="Arial",
            size=9,
        ),
        "wrap": Alignment(
            vertical="top",
            wrap_text=True,
        ),
        "nowrap": Alignment(
            vertical="top",
            wrap_text=False,
        ),
        "fill_white": PatternFill(
            "solid",
            start_color="FFFFFF",
        ),
        "fill_light": PatternFill(
            "solid",
            start_color="F2F7FC",
        ),
        "border": Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    }


def _write_header(ws, columns, styles):
    for idx, (_, label, _) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=idx, value=label)

        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20


def _set_widths(ws, columns):
    from openpyxl.utils import get_column_letter

    for idx, (_, _, width) in enumerate(columns, start=1):
        ws.column_dimensions[
            get_column_letter(idx)
        ].width = width

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
def _build_summary_sheet(ws, campaign_id, campaign, leads, contacts, styles):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    r = 1

    def write_metric(name, value):
        nonlocal r

        c1 = ws.cell(r, 1, name)
        c2 = ws.cell(r, 2, value)

        c1.font = styles["cell_font"]
        c2.font = styles["cell_font"]

        r += 1

    def _hdr(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = styles["header_font"]
        c.fill = styles["header_fill"]
        c.alignment = styles["header_align"]

    write_metric("Campaign ID", campaign_id)
    write_metric("Campaign Name", campaign.get("name") or campaign.get("extract_id") or "")
    write_metric("Created At", campaign.get("created_at", ""))
    write_metric("Lead Count", len(leads))
    write_metric("Contact Count", len(contacts))
    write_metric("Exported At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    write_metric("Schema Version", 1)

    r += 2

    priority_counter = Counter(
        l.get("priority", "Unknown")
        for l in leads
    )

    _hdr(r, 1, "Priority")
    _hdr(r, 2, "Count")
    r += 1

    for k, v in priority_counter.items():
        write_metric(k, v)

    r += 1

    status_counter = Counter(
        l.get("status", "Unknown")
        for l in leads
    )

    _hdr(r, 1, "Status")
    _hdr(r, 2, "Count")
    r += 1

    for k, v in status_counter.items():
        write_metric(k, v)

    r += 1

    country_counter = Counter(
        l.get("country", "Unknown")
        for l in leads
    )

    _hdr(r, 1, "Country")
    _hdr(r, 2, "Count")
    r += 1

    for k, v in country_counter.items():
        write_metric(k, v)

# ---------------------------------------------------------------------------
# Campaign Sheet
# ---------------------------------------------------------------------------
def _build_campaign_sheet(ws, campaign, leads, contacts, styles):
    _write_header(ws, CAMPAIGN_COLUMNS, styles)

    filters = campaign.get("filters", {})

    values = [
        campaign.get("extract_id", ""),
        campaign.get("created_at", ""),
        len(leads),
        len(contacts),
        ", ".join(filters.get("countries") or []),
        ", ".join(filters.get("keywords") or []),
        filters.get("min_score", ""),
        filters.get("max_score", ""),
    ]

    for col, value in enumerate(values, start=1):
        ws.cell(2, col, value)

    _set_widths(ws, CAMPAIGN_COLUMNS)

# ---------------------------------------------------------------------------
# Leads Sheet
# ---------------------------------------------------------------------------
def _build_leads_sheet(ws, leads, styles):
    _write_header(ws, LEAD_COLUMNS, styles)

    for row_idx, lead in enumerate(leads, start=2):
        fill = (
            styles["fill_light"]
            if row_idx % 2 == 0
            else styles["fill_white"]
        )

        for col_idx, (key, _, _) in enumerate(LEAD_COLUMNS, start=1,):
            value = lead.get(key, "")

            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill = fill
            c.font = styles["cell_font"]
            c.border = styles["border"]

            if key in ("notes", "suggested_angle"):
                c.alignment = styles["wrap"]
            else:
                c.alignment = styles["nowrap"]

        if lead.get("notes") or lead.get("suggested_angle"):
            ws.row_dimensions[row_idx].height = 45

    _set_widths(ws, LEAD_COLUMNS)

# ---------------------------------------------------------------------------
# Contacts Sheet
# ---------------------------------------------------------------------------
def _build_contacts_sheet(ws, contacts, styles):
    _write_header(ws, CONTACT_COLUMNS, styles)

    for row_idx, contact in enumerate(contacts, start=2):
        fill = (
            styles["fill_light"]
            if row_idx % 2 == 0
            else styles["fill_white"]
        )

        for col_idx, (key, _, _) in enumerate(CONTACT_COLUMNS, start=1):
            value = contact.get(key, "")

            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.fill = fill
            c.font = styles["cell_font"]
            c.border = styles["border"]
            c.alignment = styles["nowrap"]

    _set_widths(ws, CONTACT_COLUMNS)

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
def _build_xlsx(output_dir, campaign_id, campaign, leads, contacts):
    from openpyxl import Workbook
    wb = Workbook()
    styles = _styles()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    _build_summary_sheet(ws_summary, campaign_id, campaign, leads, contacts,styles,)
    ws_campaign = wb.create_sheet("Campaign")

    _build_campaign_sheet(ws_campaign, campaign, leads, contacts, styles)
    ws_leads = wb.create_sheet("Leads")

    _build_leads_sheet(ws_leads, leads, styles)
    ws_contacts = wb.create_sheet("Contacts")

    _build_contacts_sheet(ws_contacts, contacts, styles)
    xlsx_file = output_dir / "campaign.xlsx"
    wb.save(xlsx_file)

    return xlsx_file

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def export_campaign(campaign_id: str, output_dir: str | None = None):
    campaign_ref, campaign = (_load_campaign(campaign_id))
    leads, contacts = (_load_leads_and_contacts(campaign_ref))
    output_dir = Path(output_dir) if output_dir else (Path("output") / campaign_id)
    output_dir.mkdir(parents=True, exist_ok=True)

    if campaign.get("extract_id"):
        assert (campaign["extract_id"] == campaign_id), \
            (
            f"Campaign mismatch: "
            f"{campaign['extract_id']} != {campaign_id}"
        )

    json_file = _write_json(output_dir, campaign, leads, contacts)
    xlsx_file = _build_xlsx(output_dir, campaign_id, campaign, leads, contacts)

    print(
        f"[campaign_export] "
        f"{campaign_id} "
        f"→ {len(leads)} leads / "
        f"{len(contacts)} contacts"
    )

    return {
        "campaign_id": campaign_id,
        "lead_count": len(leads),
        "contact_count": len(contacts),
        "json_file": str(json_file),
        "xlsx_file": str(xlsx_file),
    }

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> None:
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass

    import argparse
    p = argparse.ArgumentParser(
        description="Export a leads_extract campaign to Excel + JSON"
    )
    p.add_argument(
        "campaign_id",
        nargs="?",
        default=None,
        metavar="CAMPAIGN_ID",
        help="Firestore document ID under leads_extract/  e.g. NO_high_score_may26",
    )
    p.add_argument(
        "--list",
        action="store_true",
        help="List all available campaign IDs and exit",
    )
    p.add_argument(
        "--output",
        default=None,
        metavar="DIR",
        help="Output directory  (default: output/<campaign_id>/)",
    )

    args = p.parse_args(argv)

    if args.list:
        from app.firestore_client import get_firestore
        db  = get_firestore()
        ids = [doc.id for doc in db.collection("leads_extract").stream()]
        if ids:
            print("Available campaigns:")
            for cid in sorted(ids):
                print(f"  {cid}")
        else:
            print("No campaigns found in leads_extract collection.")
        return

    if not args.campaign_id:
        p.error("campaign_id is required (or use --list to see available campaigns)")

    result = export_campaign(args.campaign_id, output_dir=args.output)

    print(f"\n  Done → {result['xlsx_file']}")
    print(f"         {result['json_file']}")
    print(f"  Leads: {result['lead_count']}  Contacts: {result['contact_count']}")


if __name__ == "__main__":
    main()