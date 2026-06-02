"""leads_smart_export.py — Smart tiered export of leads (web agencies) by reseller potential.

Reads leads + contacts from Firestore, keeps only leads with valid email contacts,
scores and groups into 5 tiers based on reseller fit for BlueSearch (selling SEO
services through agency resellers).

Tier 1 — Prime Reseller  : ai_reseller_potential=high  + score >= 75 (A priority)
Tier 2 — Strong Prospect : ai_reseller_potential=high  + score 55-74  OR  WordPress+SMB
Tier 3 — Good Prospect   : ai_reseller_potential=medium + score >= 55
Tier 4 — Possible        : ai_reseller_potential=medium + score < 55
Tier 5 — Low Priority    : ai_reseller_potential=low   or no AI classification

Bonus signals that lift a tier:
  - WordPress/WooCommerce specialisation → +1 tier
  - SMB client base → +1 tier
  - care_plan / hosting in specialisation → +1 tier

Usage:
    python app\\leads_smart_export.py --countries UK
    python app\\leads_smart_export.py --countries UK IN NO --out exports\\leads_resellers.xlsx
    python app\\leads_smart_export.py --countries NO SE DK --min-score 50
"""
from __future__ import annotations

import threading as _threading

# Guards firebase_admin.initialize_app against concurrent init
_local_fb_lock = _threading.Lock()
import argparse
import importlib.util
import os
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Email validation
# ---------------------------------------------------------------------------

_EMAIL_RE  = re.compile(r'^[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-]+(?:\.[a-zA-Z0-9\-]+)+$')
_EMAIL_BAD = re.compile(
    r'(noreply|no-reply|donotreply|example|localhost|invalid|test@|dummy|placeholder)',
    re.IGNORECASE
)

def _valid_email(email: str) -> bool:
    if not email or '@' not in email:
        return False
    local = email.split('@')[0]
    if not _EMAIL_RE.match(email):
        return False
    if _EMAIL_BAD.search(email):
        return False
    if len(local) >= 16 and re.fullmatch(r'[0-9a-f\-]+', local):
        return False
    return True

# ---------------------------------------------------------------------------
# Tier classification
# ---------------------------------------------------------------------------

WP_TAGS    = {'wordpress', 'woocommerce', 'elementor', 'divi', 'wp'}
SMB_TAGS   = {'smb', 'local', 'small business'}
CARE_TAGS  = {'care_plan', 'hosting', 'maintenance'}


def _score_lead(lead: dict, email_count: int) -> tuple[int, str]:
    potential  = (lead.get('ai_reseller_potential') or '').lower()
    score      = int(float(lead.get('reseller_score') or 0))
    priority   = (lead.get('priority') or '').upper()
    specs      = [s.lower() for s in (lead.get('ai_specialisation') or [])]
    client_base = (lead.get('ai_client_base') or '').lower()

    is_wp   = any(t in specs for t in WP_TAGS)
    is_smb  = client_base in ('smb', 'local') or any(t in specs for t in SMB_TAGS)
    is_care = any(t in specs for t in CARE_TAGS)

    # Base tier from AI potential + score
    if potential == 'high' and score >= 75:
        tier = 1
    elif potential == 'high':
        tier = 2
    elif potential == 'medium' and score >= 55:
        tier = 3
    elif potential == 'medium':
        tier = 4
    else:
        tier = 5

    # Bonus: bump up one tier for strong reseller signals
    bonuses = sum([is_wp, is_smb, is_care, email_count >= 2])
    if bonuses >= 2 and tier > 1:
        tier -= 1
    elif bonuses == 1 and tier > 2:
        tier -= 1

    labels = {
        1: '1 - Prime Reseller',
        2: '2 - Strong Prospect',
        3: '3 - Good Prospect',
        4: '4 - Possible',
        5: '5 - Low Priority',
    }
    return tier, labels[tier]

# ---------------------------------------------------------------------------
# Firestore
# ---------------------------------------------------------------------------

def _load_secrets():
    p = Path(__file__).parent.parent / 'blueboot_secrets.py'
    if not p.exists():
        return None
    try:
        spec = importlib.util.spec_from_file_location('blueboot_secrets', p)
        mod  = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        return getattr(mod, 'fireBaseAdminKey', None)
    except Exception:
        return None


def _init_firestore(fb_key):
    import firebase_admin
    from firebase_admin import firestore
    import firebase_admin.credentials as creds
    c = creds.Certificate(fb_key) if fb_key else creds.Certificate(
        os.getenv('FIREBASE_CREDENTIALS', 'config/serviceAccountKey.json'))
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    return firestore.client()

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_data(db, countries: list[str] | None, min_score: int,
               outreach_priority: int | None = None,
               collection: str = 'leads') -> list[dict]:
    from google.cloud.firestore_v1.base_query import FieldFilter

    col   = db.collection(collection)
    rows: list[dict] = []
    scanned = skipped = 0

    if countries and len(countries) <= 10:
        stream = col.where(filter=FieldFilter('country', 'in', countries)).stream()
        print(f'[leads-export] Firestore query: country in {countries}', flush=True)
    else:
        stream = col.stream()
        print(f'[leads-export] Streaming all {collection}…', flush=True)

    for lead_doc in stream:
        scanned += 1
        lead = lead_doc.to_dict() or {}

        if lead.get('country') == '*':
            skipped += 1
            continue

        if countries:
            c = (lead.get('country') or '').upper()
            if c not in countries:
                skipped += 1
                continue

        score = int(float(lead.get('reseller_score') or 0))
        if score < min_score:
            skipped += 1
            continue

        # Load contacts
        contacts_raw = list(lead_doc.reference.collection('contacts').stream())
        valid = [
            c.to_dict() or {} for c in contacts_raw
            if _valid_email((c.to_dict() or {}).get('email', ''))
        ]
        if not valid:
            skipped += 1
            continue

        # Outreach priority filter on contacts
        if outreach_priority is not None:
            valid = [c for c in valid
                     if c.get("outreach_priority") is None
                     or int(c.get("outreach_priority", 4)) <= outreach_priority]
            if not valid:
                skipped += 1
                continue

        tier, tier_label = _score_lead(lead, len(valid))

        specs = lead.get('ai_specialisation') or []
        base = {
            'tier':           tier,
            'tier_label':     tier_label,
            'domain':         lead.get('domain', ''),
            'website':        lead.get('website', ''),
            'company':        lead.get('company', ''),
            'country':        (lead.get('country') or '').upper(),
            'reseller_score': score,
            'priority':       lead.get('priority', ''),
            'ai_potential':   lead.get('ai_reseller_potential', ''),
            'ai_sector':      lead.get('ai_sector', ''),
            'specialisation': ', '.join(specs[:6]),
            'client_base':    lead.get('ai_client_base', ''),
            'platform':       lead.get('ai_platform', ''),
            'summary':        (lead.get('ai_summary') or lead.get('description') or '')[:120],
            'reasons':        (lead.get('reasons') or '')[:100],
            'email_count':    len(valid),
            'source':         'catalog' if lead.get('found_by_catalog') == 'yes' else 'search',
        }

        for contact in valid:
            rows.append({
                **base,
                'email':             contact.get('email', ''),
                'name':              contact.get('name', ''),
                'title':             contact.get('title', ''),
                'phone':             contact.get('phone', ''),
                'linkedin':          contact.get('linkedin', ''),
                'email_type':        contact.get('email_type', ''),
                'contact_type':      contact.get('contact_type', ''),
                'outreach_priority': contact.get('outreach_priority', ''),
            })

        if scanned % 500 == 0:
            print(f'[leads-export] {scanned} leads scanned, {len(rows)} contacts…', flush=True)

    print(f'[leads-export] {scanned} scanned, {skipped} filtered, '
          f'{len(rows)} contacts in {len(set(r["domain"] for r in rows))} agencies', flush=True)
    return rows

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

TIER_COLORS = {
    1: 'C00000', 2: 'FF0000', 3: 'FF9900', 4: 'FFFF00', 5: 'D9D9D9',
}
TIER_TEXT = {1: 'FFFFFF', 2: 'FFFFFF', 3: '000000', 4: '000000', 5: '000000'}


def _build_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    THIN   = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL = PatternFill('solid', start_color='1F497D')
    HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    DATA_FONT = Font(name='Arial', size=10)
    WRAP  = Alignment(wrap_text=True, vertical='top')
    NOWRAP = Alignment(vertical='top')

    def _hdr(ws, row, col, val, w=None):
        c = ws.cell(row, col, val)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
        if w:
            ws.column_dimensions[get_column_letter(col)].width = w
        return c

    # ── Sheet 1: All Contacts ───────────────────────────────────────────
    ws = wb.active
    ws.title = 'Contacts'

    COLS = [
        ('Tier',           'tier_label',     20),
        ('Score',          'reseller_score',  8),
        ('Priority',       'priority',       12),
        ('AI Potential',   'ai_potential',   12),
        ('Domain',         'domain',         28),
        ('Company',        'company',        24),
        ('Email',          'email',          30),
        ('Name',           'name',           22),
        ('Title',          'title',          20),
        ('Phone',          'phone',          14),
        ('LinkedIn',       'linkedin',       30),
        ('Email Type',     'email_type',     12),
        ('Contact Role',   'contact_type',   16),
        ('Outreach P',     'outreach_priority', 9),
        ('Country',        'country',         8),
        ('Sector',         'ai_sector',      16),
        ('Specialisation', 'specialisation', 30),
        ('Client Base',    'client_base',    12),
        ('Platform',       'platform',       14),
        ('Source',         'source',         10),
        ('Summary',        'summary',        50),
        ('Reasons',        'reasons',        35),
    ]

    for ci, (hdr, _, w) in enumerate(COLS, 1):
        _hdr(ws, 1, ci, hdr, w)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    sorted_rows = sorted(rows, key=lambda r: (r['tier'], -r['reseller_score']))

    for ri, row in enumerate(sorted_rows, 2):
        tier = row['tier']
        bg, fg = TIER_COLORS.get(tier, 'FFFFFF'), TIER_TEXT.get(tier, '000000')
        tier_fill = PatternFill('solid', start_color=bg)
        tier_font = Font(name='Arial', size=10, color=fg, bold=tier <= 2)

        for ci, (_, key, _) in enumerate(COLS, 1):
            val  = row.get(key, '')
            cell = ws.cell(ri, ci, val)
            cell.border = BORDER
            if ci == 1:
                cell.fill = tier_fill
                cell.font = tier_font
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif key in ('summary', 'reasons', 'linkedin', 'specialisation'):
                cell.font = DATA_FONT
                cell.alignment = WRAP
            else:
                cell.font = DATA_FONT
                cell.alignment = NOWRAP
        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

    # ── Sheet 2: Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 10

    for ci, hdr in enumerate(['Tier', 'Agencies', 'Contacts'], 1):
        _hdr(ws2, 1, ci, hdr)

    tier_labels = {1:'1-Prime Reseller', 2:'2-Strong Prospect',
                   3:'3-Good Prospect', 4:'4-Possible', 5:'5-Low Priority'}
    tier_sites = Counter()
    tier_c     = Counter(r['tier'] for r in rows)
    for r in rows:
        tier_sites[(r['tier'], r['domain'])] = 1
    ts_count: dict[int, int] = {}
    for (t, _) in tier_sites:
        ts_count[t] = ts_count.get(t, 0) + 1

    for ri, t in enumerate(sorted(tier_labels), 2):
        bg, fg = TIER_COLORS.get(t, 'FFFFFF'), TIER_TEXT.get(t, '000000')
        for ci, val in enumerate([tier_labels[t], ts_count.get(t, 0), tier_c.get(t, 0)], 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.font = Font(name='Arial', size=10, color=fg)
            cell.border = BORDER

    ws2.cell(8, 1, 'TOTAL').font = Font(name='Arial', bold=True)
    ws2.cell(8, 2, len(set(r['domain'] for r in rows))).font = Font(name='Arial', bold=True)
    ws2.cell(8, 3, len(rows)).font = Font(name='Arial', bold=True)

    # Specialisation breakdown
    ws2.cell(10, 1).value = 'Top Specialisations'
    ws2.cell(10, 1).font  = Font(name='Arial', bold=True)
    spec_c: Counter = Counter()
    for r in rows:
        for s in r['specialisation'].split(', '):
            if s.strip():
                spec_c[s.strip()] += 1
    for ri, (sp, cnt) in enumerate(spec_c.most_common(12), 11):
        ws2.cell(ri, 1, sp)
        ws2.cell(ri, 2, cnt)

    # Client base breakdown
    ws2.cell(10, 4).value = 'Client Base'
    ws2.cell(10, 4).font  = Font(name='Arial', bold=True)
    cb_c = Counter(r['client_base'] or 'unknown' for r in rows)
    for ri, (cb, cnt) in enumerate(cb_c.most_common(), 11):
        ws2.cell(ri, 4, cb)
        ws2.cell(ri, 5, cnt)

    # ── Sheet 3: WordPress focus ─────────────────────────────────────────
    ws3 = wb.create_sheet('WP vs Others')
    wp_rows = [r for r in rows if any(t in r['specialisation'].lower()
                                       for t in ['wordpress', 'woocommerce'])]
    ot_rows = [r for r in rows if r not in wp_rows]

    for ci, hdr in enumerate(['Group', 'Agencies', 'Contacts', 'Avg Score'], 1):
        _hdr(ws3, 1, ci, hdr)
    ws3.column_dimensions['A'].width = 22
    for col in ['B', 'C', 'D']: ws3.column_dimensions[col].width = 12

    for ri, (label, grp) in enumerate([
        ('WordPress / WooCommerce', wp_rows),
        ('Other specialisations', ot_rows),
    ], 2):
        sites   = len(set(r['domain'] for r in grp))
        avg_scr = int(sum(r['reseller_score'] for r in grp) / len(grp)) if grp else 0
        for ci, val in enumerate([label, sites, len(grp), avg_scr], 1):
            ws3.cell(ri, ci, val).font = Font(name='Arial', size=10)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'[leads-export] Saved → {out_path}', flush=True)

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description='Tiered reseller prospect export from leads')
    p.add_argument('--countries', nargs='+', default=None, metavar='CC',
                   help='Country codes e.g. UK IN NO (comma or space separated)')
    p.add_argument('--min-score', type=int, default=0, metavar='N',
                   help='Minimum reseller_score (default 0)')
    p.add_argument('--outreach-priority', type=int, default=None, metavar='N',
                   help='Only include contacts with outreach_priority <= N (1=best)')
    p.add_argument('--collection', default='leads', metavar='NAME',
                   help='Firestore collection (default: leads)')
    p.add_argument('--out', default=None, metavar='PATH',
                   help='Output .xlsx path (default: exports/leads_prospects_<countries>_<ts>.xlsx)')
    args = p.parse_args(argv)

    countries = None
    if args.countries:
        raw = []
        for t in args.countries:
            raw.extend(c.strip().upper() for c in t.split(',') if c.strip())
        countries = raw or None

    fb_key = _load_secrets()
    db     = _init_firestore(fb_key)

    rows = _load_data(db, countries, args.min_score, args.outreach_priority, args.collection)
    if not rows:
        print('[leads-export] No contacts found.')
        return

    if not args.out:
        ts  = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')
        cc  = '_'.join(countries) if countries else 'all'
        out = Path('exports') / f'leads_prospects_{cc}_{ts}.xlsx'
    else:
        out = Path(args.out)

    _build_excel(rows, out)

    tc = Counter(r['tier_label'] for r in rows)
    print('\n[leads-export] Tier breakdown:')
    for label, count in sorted(tc.items()):
        print(f'  {label}: {count} contacts')


if __name__ == '__main__':
    main()
