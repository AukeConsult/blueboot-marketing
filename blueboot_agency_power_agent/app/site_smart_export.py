"""site_smart_export.py — Smart tiered export of site_leads with valid email contacts.

Reads site_leads + site_contacts from Firestore, keeps only contacts with valid emails,
scores and groups leads into 5 tiers based on page count, platform and sector fit for
BlueSearch (SEO/search visibility services), then exports to Excel.

Tier 1 — Enterprise  : >10,000 pages
Tier 2 — Hot         : 500-10,000 pages + WordPress/WooCommerce OR priority sectors
Tier 3 — Good        : 100-500 pages, any platform
Tier 4 — Warm        :  50-100 pages
Tier 5 — Cold        : <50 pages

Bonus signals that can bump a site up one tier:
  - WordPress/WooCommerce platform detected
  - Sector is ecommerce, technology, or media
  - 3+ valid email contacts found

Usage:
    python app\\site_smart_export.py --countries UK
    python app\\site_smart_export.py --countries UK IN --out exports\\smart_uk_in.xlsx
    python app\\site_smart_export.py --countries NO --min-pages 50
"""
from __future__ import annotations

import threading as _threading

# Guards firebase_admin.initialize_app against concurrent init
_local_fb_lock = _threading.Lock()
import sys
import argparse
import importlib.util
import os
import re
from pathlib import Path as _Path
sys.path.insert(0, str(_Path(__file__).parent))  # make functions/ importable
from functions.utils import clean_str, resolve_country, ISO_TO_CC, email_matches_name, normalize_url
from functions.excel_builder import write_contacts_sheet, make_header_cell, save_workbook, TIER_COLORS, TIER_TEXT
from functions.config import cfg
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Email validation
# ---------------------------------------------------------------------------

_EMAIL_RE = re.compile(r'^[a-zA-Z0-9_.+\-]+@[a-zA-Z0-9\-]+(?:\.[a-zA-Z0-9\-]+)+$')
_EMAIL_BAD = re.compile(
    r'('
    r'noreply|no-reply|donotreply|do-not-reply|unsubscribe|optout|opt-out|'
    r'example|localhost|invalid|dummy|placeholder|fake|sample|'
    r'test@|testmail|mailtest|'
    r'guided-selling|guided_selling|automated|automailer|autorespond|'
    r'besttemplate|template@|campaign@|mailer-daemon|'
    r'bounce@|bounced@|ndr@|postmaster|mailer@|'
    r'my-orders|my-packages|order-notify|orders-packages|'
    r'tracking@|shipment@|invoice@|notification@|'
    r'vtex\\.|system@|robot@|bot@|no_reply|_noreply'
    r')',
    re.IGNORECASE
)

_VALID_TLD_RE = re.compile(r'\.[a-zA-Z]{2,}$')  # requires 2+ letter TLD
_NUMERIC_DOMAIN_RE = re.compile(r'@[0-9]+\.')    # block @0., @1., @123. domains




def _valid_email(email: str) -> bool:
    # Reject strings with control chars or JSON-artifact characters
    # e.g. 'partner","slug":"anne-sofie...' leaked from JSON parsing
    if any(ord(c) < 32 or ord(c) == 127 for c in email):
        return False
    if any(c in email for c in ('"', '{', '}', '\\', '<', '>')):
        return False
    if not email or '@' not in email:
        return False
    local, _, domain = email.partition('@')
    if not local or not domain:
        return False
    if not _EMAIL_RE.match(email):
        return False
    # Reject single-char TLDs (.x, .y) and numeric-only TLDs (.1, .123)
    if not _VALID_TLD_RE.search(domain):
        return False
    # Reject numeric domain prefix like @0., @1., @123.
    if _NUMERIC_DOMAIN_RE.search(email):
        return False
    if _EMAIL_BAD.search(email):
        return False
    # Reject hex/UUID local parts (automated system addresses)
    if len(local) >= 16 and re.fullmatch(r'[0-9a-f\-]+', local):
        return False
    # Reject all-digit local parts
    if re.fullmatch(r'[0-9\-]+', local):
        return False
    return True


# ---------------------------------------------------------------------------
# Tier classification
# ---------------------------------------------------------------------------

HOT_SECTORS   = {'ecommerce', 'technology', 'media', 'company', 'finance'}
HOT_PLATFORMS = {'wordpress', 'woocommerce', 'elementor', 'divi'}


def _score_lead(lead: dict, email_count: int) -> tuple[int, str]:
    """Return (tier 1-6, tier_label) for a lead."""
    pages = int(lead.get('page_count') or 0)
    platform = (lead.get('ai_platform') or lead.get('platform') or '').lower()
    sector   = (lead.get('ai_sector') or '').lower()

    is_wp     = any(p in platform for p in HOT_PLATFORMS)
    is_hot_s  = sector in HOT_SECTORS
    many_contacts = email_count >= 3

    # Base tier from page count — two enterprise tiers
    if pages > 100_000:
        tier = 1   # Ultra Enterprise
    elif pages > 10_000:
        tier = 2   # Enterprise
    elif pages >= 500:
        tier = 3   # Hot
    elif pages >= 100:
        tier = 4   # Good
    elif pages >= 50:
        tier = 5   # Warm
    else:
        tier = 6   # Cold

    # Bonus signals bump up one tier (enterprise tiers are immune — already top)
    bonuses = sum([is_wp, is_hot_s, many_contacts])
    if bonuses >= 2 and tier > 2:
        tier -= 1
    elif bonuses == 1 and tier > 3:
        tier -= 1

    labels = {
        1: '1 - Ultra Enterprise',
        2: '2 - Enterprise',
        3: '3 - Hot',
        4: '4 - Good',
        5: '5 - Warm',
        6: '6 - Cold',
    }
    return tier, labels[tier]


# ---------------------------------------------------------------------------
# Firestore
# ---------------------------------------------------------------------------

def _load_secrets():
    """Load Firebase credentials from env (FIREBASE_KEY_JSON or FIREBASE_CREDENTIALS)."""
    from dotenv import load_dotenv
    load_dotenv()
    from functions.firebase_cred import get_firebase_cred
    return get_firebase_cred()


def _init_firestore(fb_key):
    import firebase_admin
    from firebase_admin import firestore
    import firebase_admin.credentials as creds
    # fb_key may already be a Certificate object from get_firebase_cred()
    if isinstance(fb_key, creds.Certificate):
        cred = fb_key
    elif fb_key:
        cred = (fb_key if isinstance(fb_key, creds.Base) else creds.Certificate(fb_key))
    else:
        cred = creds.Certificate(
            cfg.FIREBASE_CREDENTIALS or "config/serviceAccountKey.json")
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    return firestore.client()


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_data(db, countries: list[str] | None, min_pages: int,
               outreach_priority: int | None = None) -> list[dict]:
    """Load contacts via collectionGroup, then batch-fetch parent leads. Fast."""
    from google.cloud.firestore_v1.base_query import FieldFilter
    from concurrent.futures import ThreadPoolExecutor

    PAGE_SIZE    = 500
    LEAD_BATCH   = 500   # Firestore in() limit
    WORKERS      = 20
    SCAN_PARTITIONS = 32

    print('[smart-export] Step 1: scanning site_contacts in parallel partitions...', flush=True)

    # Step 1: collect all valid contacts + their lead IDs (parallel partitions)
    cg = db.collection_group('site_contacts')

    def _scan_partition(query):
        local: dict[str, list[dict]] = {}
        for doc in query.stream():
            c = doc.to_dict() or {}
            if not _valid_email((c.get('email') or '').strip()):
                continue
            # Use lead_id field from contact doc (authoritative) or fall back to path
            # Path structure: site_leads/{lead_id}/site_contacts/{contact_id}
            parts = doc.reference.path.split('/')
            lead_id = c.get('lead_id') or (parts[1] if len(parts) >= 4 else None)
            if lead_id:
                local.setdefault(lead_id, []).append(c)
        return local

    try:
        queries = [pt.query() for pt in cg.get_partitions(SCAN_PARTITIONS)]
    except Exception as exc:
        print(f'[smart-export]   partitioning unavailable ({exc}); single-pass scan', flush=True)
        queries = [cg.order_by('__name__')]

    contact_map: dict[str, list[dict]] = {}   # lead_id -> [contact, ...]
    total_contacts = 0
    done_parts = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        for local in pool.map(_scan_partition, queries):
            for lid, lst in local.items():
                contact_map.setdefault(lid, []).extend(lst)
                total_contacts += len(lst)
            done_parts += 1
            print(f'[smart-export]   partition {done_parts}/{len(queries)} done  '
                  f'{total_contacts} valid contacts in {len(contact_map)} leads', flush=True)

    print(f'[smart-export] Step 1 done: {total_contacts} valid contacts in {len(contact_map)} leads', flush=True)

    # ── Step 2: batch-fetch parent lead docs ────────────────────────────
    print('[smart-export] Step 2: batch-fetching parent site_leads…', flush=True)
    lead_ids = list(contact_map.keys())
    leads_col = db.collection('site_leads')
    lead_data: dict[str, dict] = {}

    def _fetch_batch(batch_ids):
        refs = [leads_col.document(lid) for lid in batch_ids]
        return db.get_all(refs)

    batches = [lead_ids[i:i+LEAD_BATCH] for i in range(0, len(lead_ids), LEAD_BATCH)]
    total_batches = len(batches)
    print(f'[smart-export]   {len(lead_ids)} leads across {total_batches} batches of ≤{LEAD_BATCH} ({WORKERS} parallel)…', flush=True)
    done_batches = 0
    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = [pool.submit(_fetch_batch, b) for b in batches]
        for fut in futures:
            for doc in fut.result():
                if doc.exists:
                    lead_data[doc.id] = doc.to_dict() or {}
            done_batches += 1
            if done_batches % 5 == 0 or done_batches == total_batches:
                print(f'[smart-export]   batch {done_batches}/{total_batches}  ({len(lead_data)} leads fetched so far)', flush=True)

    print(f'[smart-export] Step 2 done: {len(lead_data)}/{len(lead_ids)} lead docs fetched', flush=True)

    # ── Step 3: merge + filter ───────────────────────────────────────────
    print('[smart-export] Step 3: merging and filtering…', flush=True)
    rows: list[dict] = []
    skipped = 0

    for lead_id, contacts in contact_map.items():
        lead = lead_data.get(lead_id)
        if not lead:
            skipped += 1
            continue

        if lead.get('country') == '*':
            skipped += 1
            continue

        if countries:
            # Priority: ai_country (GPT-classified, most reliable) → country (TLD detection)
            ai_c   = (lead.get('ai_country') or '').upper().strip()
            raw_c  = (lead.get('country') or '').upper().strip()
            # Normalise both ISO → internal code  (GB → UK)
            best   = ISO_TO_CC.get(ai_c, ai_c) or ISO_TO_CC.get(raw_c, raw_c)
            if best not in countries:
                skipped += 1
                continue

        pages = int(lead.get('page_count') or 0)
        if pages < min_pages:
            skipped += 1
            continue

        # Outreach priority filter on contacts
        if outreach_priority is not None:
            contacts = [c for c in contacts
                        if c.get("outreach_priority") is None
                        or int(c.get("outreach_priority", 4)) <= outreach_priority]
            if not contacts:
                skipped += 1
                continue

        tier, tier_label = _score_lead(lead, len(contacts))

        base = {
            'tier':         tier,
            'tier_label':   tier_label,
            'domain':       lead.get('domain', ''),
            'website':      normalize_url(lead.get('website', '') or ''),
            'country':      (lead.get('ai_country') or lead.get('country') or '').upper(),
            'ai_country':   (lead.get('ai_country') or '').upper(),
            'raw_country':  (lead.get('country') or '').upper(),
            'pages':        pages,
            'platform':        lead.get('ai_platform') or lead.get('platform') or '',
            'sector':          lead.get('ai_sector') or '',
            'ai_company_type': lead.get('ai_company_type') or '',
            'ai_confidence':   lead.get('ai_confidence') or 0,
            'ai_summary':      (lead.get('ai_summary') or '')[:200],
            'keywords':        lead.get('ai_keywords') or lead.get('keywords') or [],
            'location':             lead.get('location') or lead.get('location_full') or '',
            'location_city':        lead.get('location_city') or '',
            'location_region':      lead.get('location_region') or '',
            'location_country':     lead.get('location_country') or '',
            'location_confidence':  lead.get('location_confidence') or '',
            'location_source':      lead.get('location_source') or '',
            'email_count':     len(contacts),
            'sitemap_type':    lead.get('sitemap_type') or '',
            'sitemap_oldest':  lead.get('sitemap_oldest_date') or '',
        }

        for contact in contacts:
            rows.append({
                **base,
                'lead_id_site':        lead_id,
                'mark_site_leads':        True,
                'contact_id':         contact.get('contact_id', ''),
                'email':              contact.get('email', ''),
                'name':               clean_str(contact.get('name', ''))
                                      if email_matches_name(contact.get('email',''), clean_str(contact.get('name',''))) else '',
                'title':              clean_str(contact.get('title', '') or contact.get('occupation', '')),
                'phone':              contact.get('phone', ''),
                'linkedin':           contact.get('linkedin', ''),
                'found_on':           contact.get('found_on', ''),
                'email_type':         contact.get('email_type', ''),
                'contact_type':       contact.get('contact_type', ''),
                'outreach_priority':  contact.get('outreach_priority', ''),
            })

    print(f'[smart-export] {len(rows)} contact rows in '
          f'{len(set(r["domain"] for r in rows))} sites  ({skipped} leads skipped)', flush=True)
    return rows

# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

TIER_COLORS = {
    1: '7030A0',  # Purple    — Ultra Enterprise (>100k pages)
    2: 'C00000',  # Dark red  — Enterprise       (>10k pages)
    3: 'FF0000',  # Red       — Hot
    4: 'FF9900',  # Orange    — Good
    5: 'FFFF00',  # Yellow    — Warm
    6: 'D9D9D9',  # Grey      — Cold
}
TIER_TEXT = {1: 'FFFFFF', 2: 'FFFFFF', 3: 'FFFFFF', 4: '000000', 5: '000000', 6: '000000'}


def _build_excel(rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _THIN  = Side(style='thin', color='CCCCCC')
    BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    COLS = [
        # Contact
        ('Approved',      'approved',           10),
        ('Tier',          'tier_label',        20),
        ('Outreach P',    'outreach_priority',  9),
        ('Email',         'email',             32),
        ('Website',       'website',           35),
        ('Name',          'name',              22),
        ('Title',         'title',             22),
        ('Phone',         'phone',             16),
        ('LinkedIn',      'linkedin',          30),
        ('Email Type',    'email_type',        12),
        ('Contact Role',  'contact_type',      16),
        # Site
        ('Domain',        'domain',            28),
        ('Company',       'company',           26),
        ('Country',       'country',            8),
        ('AI Country',    'ai_country',         10),
        ('Raw Country',   'raw_country',         8),
        ('Location',      'location',          30),
        ('City',          'location_city',     18),
        ('Region',        'location_region',   16),
        ('Pages',         'pages',             10),
        ('Sitemap',       'sitemap_type',      10),
        ('Site Since',    'sitemap_oldest',    12),
        # Classification
        ('Platform',      'platform',          16),
        ('Sector',        'sector',            16),
        ('Company Type',  'ai_company_type',   14),
        ('Confidence',    'ai_confidence',      9),
        ('Summary',       'ai_summary',        55),
        ('Keywords',      'keywords',          35),
        # Origin
        ('Category',      'category_site',     18),
        ('Doc ID',        'doc_id',            28),
        ('Lead ID Site',  'lead_id_site',      28),
        ('Contact ID',    'contact_id',        28),
        ('Mark',          'mark_site_leads',         14),
        ('Email Count',   'email_count',       10),
        ('Found on',      'found_on',          30),
    ]

    write_contacts_sheet(
        ws, rows, COLS,
        sort_key  = lambda r: (int(r.get('tier') or 5), -int(r.get('pages') or 0)),
        wrap_keys = {'ai_summary', 'linkedin', 'found_on', 'keywords'},
    )

    # ── Sheet 2: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 20

    def _hdr2(cell, val):
        make_header_cell(ws2, cell.row, cell.column, val)

    # By tier
    _hdr2(ws2.cell(1, 1), 'Tier')
    _hdr2(ws2.cell(1, 2), 'Sites')
    _hdr2(ws2.cell(1, 3), 'Contacts (emails)')

    from collections import Counter
    tier_sites    = Counter(r['domain'] + '|' + str(r['tier']) for r in rows)
    tier_s: dict[int, int] = {}
    tier_c: dict[int, int] = Counter(r['tier'] for r in rows)
    for k in tier_sites:
        t = int(k.split('|')[1])
        tier_s[t] = tier_s.get(t, 0) + 1

    tier_labels = {1:'1-Ultra Enterprise', 2:'2-Enterprise', 3:'3-Hot', 4:'4-Good', 5:'5-Warm', 6:'6-Cold'}
    for ri, t in enumerate(sorted(tier_labels), 2):  # tiers 1-6
        bg = TIER_COLORS.get(t, 'FFFFFF')
        fg = TIER_TEXT.get(t, '000000')
        for ci, val in enumerate([tier_labels[t], tier_s.get(t, 0), tier_c.get(t, 0)], 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.font = Font(name='Arial', size=10, color=fg)
            cell.border = BORDER
    ws2.cell(9, 1, 'TOTAL').font = Font(name='Arial', bold=True)
    ws2.cell(9, 2, len(set(r['domain'] for r in rows))).font = Font(name='Arial', bold=True)
    ws2.cell(9, 3, len(rows)).font = Font(name='Arial', bold=True)

    # By sector
    ws2.cell(10, 1).value = 'By Sector'
    ws2.cell(10, 1).font  = Font(name='Arial', bold=True, size=11)
    sector_c = Counter(r['sector'] or 'unknown' for r in rows)
    for ri, (sec, cnt) in enumerate(sector_c.most_common(15), 11):
        ws2.cell(ri, 1, sec)
        ws2.cell(ri, 2, cnt)

    # By platform
    ws2.cell(10, 3).value = 'By Platform'
    ws2.cell(10, 3).font  = Font(name='Arial', bold=True, size=11)
    plat_c = Counter((r.get('platform') or 'unknown').lower() for r in rows)
    for ri, (plat, cnt) in enumerate(plat_c.most_common(10), 11):
        ws2.cell(ri, 3, plat)
        ws2.cell(ri, 4, cnt)
    ws2.column_dimensions['D'].width = 10

    # ── Sheet 3: By Platform ─────────────────────────────────────────────
    ws3 = wb.create_sheet('WordPress vs Others')
    wp_rows    = [r for r in rows if any(p in (r.get('platform') or '').lower() for p in HOT_PLATFORMS)]
    other_rows = [r for r in rows if r not in wp_rows]

    _hdr2(ws3.cell(1, 1), 'Group')
    _hdr2(ws3.cell(1, 2), 'Sites')
    _hdr2(ws3.cell(1, 3), 'Contacts')
    _hdr2(ws3.cell(1, 4), 'Avg Pages')

    for ri, (label, grp) in enumerate([('WordPress/WooCommerce', wp_rows), ('Other Platforms', other_rows)], 2):
        sites = len(set(r['domain'] for r in grp))
        avg_p = int(sum(r['pages'] for r in grp) / len(grp)) if grp else 0
        for ci, val in enumerate([label, sites, len(grp), avg_p], 1):
            ws3.cell(ri, ci, val).font = Font(name='Arial', size=10)
    ws3.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D']:
        ws3.column_dimensions[col].width = 12

    save_workbook(wb, out_path, '[smart-export]')


# ---------------------------------------------------------------------------
# CLI

# ---------------------------------------------------------------------------
# Country normalisation — ISO → internal code
# ---------------------------------------------------------------------------


def _doc_id(email: str) -> str:
    """Firestore document ID: email with special chars replaced by underscore."""
    import re as _re
    return _re.sub(r'[^a-zA-Z0-9_-]', '_', email.lower())


def _derive_name_from_email(email: str) -> tuple[str, str]:
    """Derive first/full name from a personal email local part.

    john.smith@  →  ("John", "John Smith")
    maria.garcia@ →  ("Maria", "Maria Garcia")
    john@         →  ("John", "John")
    j.smith@      →  ("Smith", "Smith")   # single letter dropped
    info@         →  ("", "")             # generic, not a name
    Returns ("", "") if local part doesn't look like a real name.
    """
    import re as _re
    local = email.split('@')[0].lower()
    # Remove common email suffixes that aren't names
    _GENERIC = {'info', 'hello', 'contact', 'support', 'sales', 'admin',
                'webmaster', 'mail', 'team', 'office', 'enquiries', 'enquiry',
                'help', 'billing', 'accounts', 'marketing', 'press', 'media'}
    if local in _GENERIC:
        return '', ''
    parts = _re.split(r'[._\-+0-9]+', local)
    parts = [p.capitalize() for p in parts if len(p) >= 2 and p.isalpha()]
    if not parts:
        return '', ''
    full_name = ' '.join(parts)
    first_name = parts[0]
    return first_name, full_name


def _write_email_contacts(db, rows: list[dict], campaign: str | None,
                           dry_run: bool = False) -> int:
    """Write filtered contacts to email_contacts Firestore collection.

    - Document ID = _doc_id(email)  →  natural deduplication by email.
    - Data fields always overwritten (merge=True in batches of 400).
    - Lifecycle fields (status, created_at) only set for NEW documents.
    - Uses parallel pre-scan + batch writes — no per-doc round trips.
    Returns count of docs written.
    """
    from datetime import datetime, timezone as _tz
    from concurrent.futures import ThreadPoolExecutor as _TPE

    BATCH_SIZE = 400   # Firestore max is 500, leave headroom
    WORKERS    = 10
    col        = db.collection('email_contacts')
    now_ts     = datetime.now(_tz.utc).isoformat()

    # Build doc dicts
    valid_rows = []
    for row in rows:
        email = (row.get('email') or '').strip()
        if not email:
            continue
        name  = clean_str((row.get('name') or '').strip())
        # Clear name if it doesn't correspond to the email address
        if name and not email_matches_name(email, name):
            name = ''
        # For personal emails with no known name, derive from email local part
        if not name and (row.get('email_type') or '') == 'personal':
            first, name = _derive_name_from_email(email)
        else:
            first = (name.split() or [''])[0]
        doc_id = _doc_id(email)
        valid_rows.append((doc_id, email, name, first, row))

    if not valid_rows:
        return 0

    total = len(valid_rows)
    print(f'  [write] {total} contacts to write…', flush=True)

    if dry_run:
        from collections import Counter as _Ctr
        tier_c = _Ctr(row.get('tier', 0) for _, _, _, _, row in valid_rows)
        prio_c = _Ctr(str(row.get('outreach_priority', '?')) for _, _, _, _, row in valid_rows)
        print(f'  [DRY] Would write {total} contacts to email_contacts', flush=True)
        print(f'  [DRY] Tiers:    { dict(sorted(tier_c.items())) }', flush=True)
        print(f'  [DRY] Priority: { dict(sorted(prio_c.items())) }', flush=True)
        return total

    # ── Step A: pre-scan which docs already exist ─────────────────────────
    print(f'  [write] Step A: scanning {total} existing docs…', flush=True)
    doc_ids = [r[0] for r in valid_rows]
    existing_ids: set[str] = set()

    def _check_batch(ids):
        refs = [col.document(i) for i in ids]
        snaps = db.get_all(refs)
        return {s.id for s in snaps if s.exists}

    id_batches = [doc_ids[i:i+BATCH_SIZE] for i in range(0, len(doc_ids), BATCH_SIZE)]
    with _TPE(max_workers=WORKERS) as pool:
        for result in pool.map(_check_batch, id_batches):
            existing_ids |= result
    new_count = total - len(existing_ids)
    print(f'  [write] Step A done: {len(existing_ids)} existing, {new_count} new', flush=True)

    # ── Step B: build data docs (no lifecycle fields) ─────────────────────
    data_docs: list[tuple[str, dict]] = []
    lifecycle_docs: list[tuple[str, dict]] = []   # only for new docs

    lifecycle = {'status': 'pending', 'created_at': now_ts}

    for doc_id, email, name, first, row in valid_rows:
        doc = {
            # Identity
            'doc_id':             doc_id,
            'approved':           '',
            'lead_id_site':       row.get('lead_id_site', ''),
            'contact_id':         row.get('contact_id', ''),
            'mark_site_leads':    True,
            'email':              email,
            'name':               name,
            'title':              clean_str(row.get('title', '')),
            'phone':              row.get('phone', ''),
            'linkedin':           row.get('linkedin', ''),
            # Source site
            'domain':             row.get('domain', ''),
            'website':            normalize_url(row.get('website', '') or ''),
            'company':            row.get('company', ''),
            'country':            resolve_country(row),
            'location':             row.get('location', ''),
            'location_city':        row.get('location_city', ''),
            'location_region':      row.get('location_region', ''),
            'location_country':     row.get('location_country', ''),
            'location_confidence':  row.get('location_confidence', ''),
            'location_source':      row.get('location_source', ''),
            # Classification
            'ai_sector':          row.get('sector', ''),
            'ai_company_type':    row.get('ai_company_type', ''),
            'ai_platform':        row.get('platform', ''),
            'ai_confidence':      row.get('ai_confidence', 0),
            'ai_summary':         row.get('ai_summary', ''),
            'keywords':           row.get('keywords', []),
            # Site size & origin
            'page_count':         row.get('pages', 0),
            'sitemap_type':       row.get('sitemap_type', ''),
            'sitemap_oldest':     row.get('sitemap_oldest', ''),
            'category_site':      row.get('source_query', ''),
            # Contact scoring
            'tier':               row.get('tier', 0),
            'tier_label':         row.get('tier_label', ''),
            'email_type':         row.get('email_type', ''),
            'contact_type':       row.get('contact_type', ''),
            'outreach_priority':  row.get('outreach_priority', 4),
            # Pipeline metadata
            # Mail-merge
            'personalisation': {
                'name':      first,
                'full_name': name,
            },
        }

        # Lifecycle fields — only written on first creation, never on updates
        lifecycle = {
            'status':     'pending',
            'created_at': now_ts,
            'campaign':   campaign or '',
        }

        data_docs.append((doc_id, doc))
        if doc_id not in existing_ids:
            lifecycle_docs.append((doc_id, lifecycle))

    # ── Step C: batch write data fields ───────────────────────────────────
    total_batches = (len(data_docs) + BATCH_SIZE - 1) // BATCH_SIZE
    print(f'  [write] Step C: writing data fields in {total_batches} batches…', flush=True)
    done = 0
    for i in range(0, len(data_docs), BATCH_SIZE):
        chunk = data_docs[i:i+BATCH_SIZE]
        batch = db.batch()
        for doc_id, doc in chunk:
            batch.set(col.document(doc_id), doc, merge=True)
        batch.commit()
        done += len(chunk)
        print(f'  [write]   {done}/{len(data_docs)} data docs written', flush=True)

    # ── Step D: batch write lifecycle fields (new docs only) ─────────────
    if lifecycle_docs:
        print(f'  [write] Step D: setting lifecycle on {len(lifecycle_docs)} new docs…', flush=True)
        done = 0
        for i in range(0, len(lifecycle_docs), BATCH_SIZE):
            chunk = lifecycle_docs[i:i+BATCH_SIZE]
            batch = db.batch()
            for doc_id, lc in chunk:
                batch.set(col.document(doc_id), lc, merge=True)
            batch.commit()
            done += len(chunk)
        print(f'  [write]   {done} lifecycle docs written', flush=True)

    print(f'  [write] Done: {len(data_docs)} contacts written ({len(lifecycle_docs)} new, {len(existing_ids)} updated)', flush=True)
    return len(data_docs)

# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description='Tiered BlueSearch prospect export from site_leads')
    p.add_argument('--countries', nargs='+', default=None, metavar='CC',
                   help='Country codes e.g. UK IN NO (comma or space separated)')
    p.add_argument('--min-pages', type=int, default=0, metavar='N',
                   help='Minimum page count (default 0)')
    p.add_argument('--outreach-priority', type=int, default=None, metavar='N',
                   help='Only include contacts with outreach_priority <= N (1=best only, 2=top two, etc.)')
    p.add_argument('--write-contacts', action='store_true',
                   help='Write contacts to email_contacts Firestore collection')
    p.add_argument('--campaign', default=None, metavar='NAME',
                   help='Campaign tag written to email_contacts (e.g. UK_tier2_jun02)')
    p.add_argument('--dry-run-contacts', action='store_true',
                   help='Print what would be written to email_contacts without writing')
    p.add_argument('--out', default=None, metavar='PATH',
                   help='Output .xlsx path (default: exports/smart_<countries>_<ts>.xlsx)')
    args = p.parse_args(argv)

    countries = None
    if args.countries:
        raw = []
        for t in args.countries:
            raw.extend(c.strip().upper() for c in t.split(',') if c.strip())
        countries = raw or None

    fb_key = _load_secrets()
    db     = _init_firestore(fb_key)

    rows = _load_data(db, countries, args.min_pages, args.outreach_priority)
    if not rows:
        print('[smart-export] No contacts found matching filters.')
        return

    if not args.out:
        ts  = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')
        cc  = '_'.join(countries) if countries else 'all'
        out = Path('exports') / f'site_prospects_{cc}_{ts}.xlsx'
    else:
        out = Path(args.out)

    _build_excel(rows, out)

    # Write to email_contacts if requested
    if args.write_contacts or args.dry_run_contacts:
        print(f'\n[smart-export] Writing {len(rows)} contacts to email_contacts…', flush=True)
        n = _write_email_contacts(db, rows,
                                   campaign=args.campaign,
                                   dry_run=args.dry_run_contacts)
        tag = '[DRY RUN] ' if args.dry_run_contacts else ''
        print(f'[smart-export] {tag}{n} contacts written to email_contacts', flush=True)

    # Tier summary
    from collections import Counter
    tc = Counter(r['tier_label'] for r in rows)
    print('\n[smart-export] Tier breakdown:')
    for label, count in sorted(tc.items()):
        print(f'  {label}: {count} contacts')


if __name__ == '__main__':
    main()
