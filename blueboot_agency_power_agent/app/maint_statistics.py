"""Statistics aggregation -- reads Firestore leads and writes summary docs + Excel.

Usage:
    python app/statistics.py              # Firestore write + Excel export
    python app/statistics.py --no-excel   # Firestore write only
    python app/statistics.py --excel-only # Excel from existing Firestore data

Firestore structure written:

    statistics/priority-pr-country                    <- head document (grand totals + priority summary)
    statistics/priority-pr-country/countries/NO       <- one sub-doc per country
    statistics/priority-pr-country/countries/SE
    ...

Head document includes a by_priority summary across ALL countries:
    {
        "generated_at":   "...",
        "total_leads":    999,
        "total_contacts": 2345,
        "country_codes":  ["DE", "NO", "SE"],
        "by_priority": {
            "A":     {"leads": 50,  "contacts": 120},
            "B":     {"leads": 200, "contacts": 500},
            "unset": {"leads": 749, "contacts": 1725}
        }
    }
"""
from __future__ import annotations

import threading as _threading

# Guards firebase_admin.initialize_app against concurrent init
_local_fb_lock = _threading.Lock()
import _pathsetup  # noqa: F401 — adds project root, app/, app/functions/, app/collect-functions/ to sys.path
import os
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from functions.config import cfg


# ---------------------------------------------------------------------------
# Credential loader
# ---------------------------------------------------------------------------

def _get_credentials():
    try:
        import firebase_admin.credentials as fb_creds
    except ImportError:
        raise SystemExit("firebase-admin not installed. Run: pip install firebase-admin")

    from dotenv import load_dotenv; load_dotenv()
    from functions.firebase_cred import get_firebase_cred
    return get_firebase_cred()


def _init_firebase():
    import firebase_admin
    from firebase_admin import firestore
    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(_get_credentials())
    return firestore.client()


# ---------------------------------------------------------------------------
# Aggregation + Firestore write
# ---------------------------------------------------------------------------

def summarise_country_pr_priority(
    leads_collection=None,
    stats_collection="statistics",
    head_doc_id="priority-pr-country",
):
    """Aggregate leads + contacts per (country, priority) and write to Firestore.

    Head doc  : statistics/priority-pr-country
                Includes a grand by_priority summary across all countries.
    Sub-docs  : statistics/priority-pr-country/countries/{ISO_code}

    Returns dict with keys:
        "head"      -> head document dict
        "countries" -> {ISO_code: country_doc_dict}
    """
    db = _init_firebase()
    col_name  = leads_collection or cfg.FIRESTORE_COLLECTION
    leads_col = db.collection(col_name)

    # ------------------------------------------------------------------
    # Pass 1: stream all lead docs
    # ------------------------------------------------------------------
    print(f"  [stats] streaming leads from '{col_name}' ...")
    lead_meta = {}

    gen = leads_col.select(["country", "country_name", "priority", "lead_id"]).stream()
    while True:
        try:
            doc = next(gen)
        except StopIteration:
            break
        except (ValueError, AttributeError):
            continue
        try:
            data = doc.to_dict() or {}
        except Exception:
            continue
        if not data:
            continue
        lid          = data.get("lead_id") or doc.id
        country      = (data.get("country")      or "").strip().upper() or "XX"
        country_name = (data.get("country_name") or "").strip() or country
        priority     = (data.get("priority")     or "").strip()
        lead_meta[lid] = {
            "country":      country,
            "country_name": country_name,
            "priority":     priority,
        }

    print(f"  [stats] {len(lead_meta)} lead docs loaded.")

    # ------------------------------------------------------------------
    # Pass 2: count contacts per lead via collection_group
    # ------------------------------------------------------------------
    print("  [stats] counting contacts via collection_group('contacts') ...")
    contacts_per_lead = defaultdict(int)

    for doc in db.collection_group("contacts").select(["lead_id"]).stream():
        data = doc.to_dict()
        lid  = data.get("lead_id") or doc.reference.parent.parent.id
        if lid:
            contacts_per_lead[lid] += 1

    print(f"  [stats] {sum(contacts_per_lead.values())} contact docs counted.")

    # ------------------------------------------------------------------
    # Aggregate per (country, priority)
    # NOTE: Firestore rejects empty-string map keys -> use "unset"
    # ------------------------------------------------------------------
    agg           = defaultdict(lambda: defaultdict(lambda: {"leads": 0, "contacts": 0}))
    country_names = {}

    for lid, meta in lead_meta.items():
        code     = meta["country"]
        priority = meta["priority"] or "unset"
        country_names[code] = meta["country_name"]
        agg[code][priority]["leads"]    += 1
        agg[code][priority]["contacts"] += contacts_per_lead.get(lid, 0)

    # ------------------------------------------------------------------
    # Roll up a grand by_priority summary across all countries
    # ------------------------------------------------------------------
    grand_by_priority = defaultdict(lambda: {"leads": 0, "contacts": 0})
    for by_prio in agg.values():
        for prio, counts in by_prio.items():
            grand_by_priority[prio]["leads"]    += counts["leads"]
            grand_by_priority[prio]["contacts"] += counts["contacts"]

    # ------------------------------------------------------------------
    # Write to Firestore
    # ------------------------------------------------------------------
    generated_at  = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")
    head_ref      = db.collection(stats_collection).document(head_doc_id)
    countries_col = head_ref.collection("countries")

    country_results = {}
    grand_leads     = 0
    grand_contacts  = 0

    for code, by_prio in sorted(agg.items()):
        total_leads    = sum(v["leads"]    for v in by_prio.values())
        total_contacts = sum(v["contacts"] for v in by_prio.values())
        grand_leads    += total_leads
        grand_contacts += total_contacts

        country_doc = {
            "country":          code,
            "country_name":     country_names.get(code, code),
            "generated_at":     generated_at,
            "leads_collection": col_name,
            "total_leads":      total_leads,
            "total_contacts":   total_contacts,
            "by_priority":      dict(sorted(by_prio.items())),
        }
        countries_col.document(code).set(country_doc)
        country_results[code] = country_doc
        print(
            f"  [stats] written -> {stats_collection}/{head_doc_id}/countries/{code}"
            f"  ({total_leads} leads, {total_contacts} contacts)"
        )

    head_doc = {
        "generated_at":     generated_at,
        "leads_collection": col_name,
        "total_leads":      grand_leads,
        "total_contacts":   grand_contacts,
        "country_codes":    sorted(agg.keys()),
        "by_priority":      dict(sorted(grand_by_priority.items())),
    }
    head_ref.set(head_doc)
    print(
        f"  [stats] written -> {stats_collection}/{head_doc_id}"
        f"  (grand total: {grand_leads} leads, {grand_contacts} contacts)"
    )

    return {"head": head_doc, "countries": country_results}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(results, outdir=None):
    """Write statistics to an Excel file with three sheets.

    Sheets:
        Summary         -- grand totals + by_priority across all countries
        By Country      -- one row per country (totals)
        Country x Prio  -- one row per (country, priority) combination

    Parameters
    ----------
    results : dict
        Return value of summarise_country_pr_priority().
    outdir : str or Path or None
        Output directory.  Defaults to <project_root>/output.

    Returns
    -------
    Path to the written Excel file.
    """
    import pandas as pd

    head      = results["head"]
    countries = results["countries"]

    outdir = Path(outdir) if outdir else Path(__file__).parent.parent / "output"
    outdir.mkdir(parents=True, exist_ok=True)
    out_path = outdir / "statistics.xlsx"

    # --- Sheet 1: Summary ---
    summary_rows = [
        {"metric": "Generated at",   "value": head["generated_at"]},
        {"metric": "Leads collection","value": head["leads_collection"]},
        {"metric": "Total leads",    "value": head["total_leads"]},
        {"metric": "Total contacts", "value": head["total_contacts"]},
        {"metric": "Countries",      "value": ", ".join(head["country_codes"])},
    ]
    summary_df = pd.DataFrame(summary_rows)

    prio_rows = [
        {"priority": prio, "leads": v["leads"], "contacts": v["contacts"]}
        for prio, v in sorted(head["by_priority"].items())
    ]
    prio_df = pd.DataFrame(prio_rows)

    # --- Sheet 2: By Country ---
    country_rows = [
        {
            "country":      data["country"],
            "country_name": data["country_name"],
            "total_leads":  data["total_leads"],
            "total_contacts": data["total_contacts"],
        }
        for data in countries.values()
    ]
    country_df = pd.DataFrame(country_rows)

    # --- Sheet 3: Country x Priority ---
    detail_rows = []
    for data in countries.values():
        for prio, counts in data["by_priority"].items():
            detail_rows.append({
                "country":      data["country"],
                "country_name": data["country_name"],
                "priority":     prio,
                "leads":        counts["leads"],
                "contacts":     counts["contacts"],
            })
    detail_df = pd.DataFrame(detail_rows)

    # --- Write ---
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary",        index=False)
        prio_df.to_excel(   writer, sheet_name="By Priority",    index=False, startrow=len(summary_df) + 2)
        country_df.to_excel(writer, sheet_name="By Country",     index=False)
        detail_df.to_excel( writer, sheet_name="Country x Prio", index=False)

        # Auto-fit columns on all sheets
        for sheet in writer.book.worksheets:
            for col in sheet.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    print(f"  [stats] Excel written -> {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Pretty-print helper
# ---------------------------------------------------------------------------

def _print_summary(results):
    head      = results["head"]
    countries = results["countries"]
    print()
    print("=" * 60)
    print("  Statistics: priority x country")
    print(f"  Generated   : {head['generated_at']}")
    print(f"  Grand total : {head['total_leads']} leads, {head['total_contacts']} contacts")
    print()
    print("  By priority (all countries):")
    for prio, v in head["by_priority"].items():
        print(f"    {prio:20s}  leads: {v['leads']:4d}   contacts: {v['contacts']:4d}")
    print("=" * 60)
    for code, data in countries.items():
        print(
            f"\n  [{code}] {data['country_name']}"
            f"  --  {data['total_leads']} leads, {data['total_contacts']} contacts"
        )
        for prio, counts in data["by_priority"].items():
            print(f"    {prio:20s}  leads: {counts['leads']:4d}   contacts: {counts['contacts']:4d}")
    print()


# ---------------------------------------------------------------------------
# Reasons count aggregation
# ---------------------------------------------------------------------------

def summarise_reasons_count(
    leads_collection=None,
    stats_collection="statistics",
    head_doc_id="reasons-count",
    writeback=True,
):
    """Aggregate reason strings per country and write to Firestore.

    Each lead's 'reasons' field is a semicolon-separated string of scored
    signals, e.g. "wordpress: site, plugins; agency language: web, design".
    This function splits them, counts occurrences per country, and writes:

        statistics/reasons-count                    <- head doc (global totals)
        statistics/reasons-count/countries/NO       <- one sub-doc per country
        statistics/reasons-count/countries/SE
        ...

    Reason counts are stored as a list of {reason, count} objects (sorted
    by count desc) because reason strings can contain characters that
    Firestore rejects as map keys (colons, slashes, etc.).

    Head doc shape:
    {
        "generated_at":     "...",
        "leads_collection": "leads",
        "total_leads":      999,
        "country_codes":    ["DE", "NO", "SE"],
        "reasons": [
            {"reason": "has services/customers/cases language", "count": 450},
            {"reason": "agency language: web, design, digital", "count": 320},
            ...
        ]
    }

    Country sub-doc shape:
    {
        "country":          "NO",
        "country_name":     "Norway",
        "generated_at":     "...",
        "leads_collection": "leads",
        "total_leads":      42,
        "reasons": [
            {"reason": "has services/customers/cases language", "count": 20},
            ...
        ]
    }

    Returns
    -------
    dict with keys "head" and "countries".
    """
    db = _init_firebase()
    col_name  = leads_collection or cfg.FIRESTORE_COLLECTION
    leads_col = db.collection(col_name)

    print(f"  [reasons] streaming leads from '{col_name}' ...")

    # country_reasons: country_code -> reason_str -> count
    country_reasons  = defaultdict(lambda: defaultdict(int))
    country_names    = {}
    country_leads    = defaultdict(int)
    global_reasons   = defaultdict(int)
    total_leads      = 0

    # lead_reasons_list: doc_id -> [reason, ...] (parsed labels, for writeback)
    lead_reasons_list = {}

    fields = ["country", "country_name", "lead_id", "reasons"]
    for doc in leads_col.select(fields).stream():
        data         = doc.to_dict()
        country      = (data.get("country")      or "").strip().upper() or "XX"
        country_name = (data.get("country_name") or "").strip() or country
        reasons_raw  = (data.get("reasons")      or "").strip()

        country_names[country] = country_name
        country_leads[country] += 1
        total_leads += 1

        parsed_reasons = []
        if reasons_raw:
            for part in reasons_raw.split(";"):
                segment = part.strip()
                if not segment:
                    continue
                # Take the label part (left of ":"), e.g.:
                #   "wordpress: site, plugins"  -> "wordpress"
                #   "NON-AGENCY penalty: -20"   -> "NON-AGENCY penalty"
                #   "has services/customers/cases language" -> (no colon, kept whole)
                label = segment.split(":")[0].strip()
                if not label:
                    continue
                # Further split on "/" to expand compound labels, e.g.:
                #   "has services/customers/cases language"
                #   -> ["has services", "customers", "cases language"]
                sub_parts = [s.strip() for s in label.split("/") if s.strip()]
                for reason in sub_parts:
                    parsed_reasons.append(reason)
                    country_reasons[country][reason] += 1
                    global_reasons[reason]           += 1

        lead_reasons_list[doc.id] = parsed_reasons

    print(f"  [reasons] {total_leads} leads read, "
          f"{len(global_reasons)} distinct reason strings found.")

    # ------------------------------------------------------------------
    # Write reasons-list back to each lead document (batch, optional)
    # ------------------------------------------------------------------
    if writeback:
        total_wb = len(lead_reasons_list)
        print(f"  [reasons] writing reasons-list back to {total_wb} lead docs ...")
        MAX_BATCH      = 400
        PROGRESS_EVERY = 100
        batch = db.batch()
        ops   = 0
        done  = 0

        for doc_id, r_list in lead_reasons_list.items():
            batch.update(leads_col.document(doc_id), {"reasons-list": r_list})
            ops  += 1
            done += 1
            if ops >= MAX_BATCH:
                batch.commit()
                batch = db.batch()
                ops   = 0
            if done % PROGRESS_EVERY == 0:
                print(f"  [reasons] {done}/{total_wb} docs updated…")

        if ops:
            batch.commit()
        print(f"  [reasons] reasons-list written to all {total_wb} lead docs.")
    else:
        print(f"  [reasons] writeback skipped (writeback=False).")

    # ------------------------------------------------------------------
    # Write to Firestore
    # ------------------------------------------------------------------
    generated_at  = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")
    head_ref      = db.collection(stats_collection).document(head_doc_id)
    countries_col = head_ref.collection("countries")

    country_results = {}

    for code in sorted(country_reasons):
        reason_list = [
            {"reason": r, "count": c}
            for r, c in sorted(country_reasons[code].items(), key=lambda x: -x[1])
        ]
        country_doc = {
            "country":          code,
            "country_name":     country_names.get(code, code),
            "generated_at":     generated_at,
            "leads_collection": col_name,
            "total_leads":      country_leads[code],
            "reasons":          reason_list,
        }
        countries_col.document(code).set(country_doc)
        country_results[code] = country_doc
        print(f"  [reasons] written -> {stats_collection}/{head_doc_id}/countries/{code}"
              f"  ({len(reason_list)} distinct reasons)")

    global_reason_list = [
        {"reason": r, "count": c}
        for r, c in sorted(global_reasons.items(), key=lambda x: -x[1])
    ]
    head_doc = {
        "generated_at":     generated_at,
        "leads_collection": col_name,
        "total_leads":      total_leads,
        "country_codes":    sorted(country_reasons.keys()),
        "reasons":          global_reason_list,
    }
    head_ref.set(head_doc)
    print(f"  [reasons] written -> {stats_collection}/{head_doc_id}"
          f"  (grand total: {total_leads} leads, {len(global_reason_list)} reasons)")

    return {"head": head_doc, "countries": country_results}


def export_reasons_to_excel(results, outdir=None):
    """Write reasons-count statistics to output/statistics_reasons.xlsx.

    Sheets:
        Global Reasons  -- all reasons sorted by count desc (across all countries)
        By Country      -- one row per (country, reason) combination

    Parameters
    ----------
    results : dict
        Return value of summarise_reasons_count().
    outdir : str or Path or None
        Output directory. Defaults to <project_root>/output.

    Returns
    -------
    Path to the written Excel file.
    """
    import pandas as pd

    head      = results["head"]
    countries = results["countries"]

    outdir = Path(outdir) if outdir else Path(__file__).parent.parent / "output"
    outdir.mkdir(parents=True, exist_ok=True)
    out_path = outdir / "statistics_reasons.xlsx"

    # Sheet 1: global reasons
    global_df = pd.DataFrame(head["reasons"])
    if global_df.empty:
        global_df = pd.DataFrame(columns=["reason", "count"])

    # Sheet 2: per country x reason
    detail_rows = []
    for data in countries.values():
        for entry in data["reasons"]:
            detail_rows.append({
                "country":      data["country"],
                "country_name": data["country_name"],
                "reason":       entry["reason"],
                "count":        entry["count"],
            })
    detail_df = pd.DataFrame(detail_rows) if detail_rows else pd.DataFrame(
        columns=["country", "country_name", "reason", "count"]
    )
    # Sort by country then count desc
    if not detail_df.empty:
        detail_df = detail_df.sort_values(["country", "count"], ascending=[True, False])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        global_df.to_excel(writer, sheet_name="Global Reasons", index=False)
        detail_df.to_excel(writer, sheet_name="By Country",     index=False)

        for sheet in writer.book.worksheets:
            for col in sheet.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    print(f"  [reasons] Excel written -> {out_path}")
    return out_path


# ---------------------------------------------------------------------------
# Collection overview statistics
# ---------------------------------------------------------------------------

OVERVIEW_COLLECTIONS = [
    # (collection_name, display_label, subcollections_to_count)
    ("leads",           "Leads (legacy)",        [("contacts", "contacts")]),
    ("leads_excluded",  "Leads Excluded",         []),
    ("site_leads",      "Site Leads",             [("site_contacts", "contacts")]),
    ("sites_excluded",  "Sites Excluded",         []),
    ("statistics",      "Statistics",             []),
]


def _count_by_field(db, col_name: str, field: str, limit: int = 5000) -> dict[str, int]:
    """Stream a collection (up to limit docs) and tally values of a field."""
    counts: dict[str, int] = {}
    for doc in db.collection(col_name).select([field]).limit(limit).stream():
        val = (doc.to_dict() or {}).get(field) or "?"
        if isinstance(val, str):
            val = val.strip().upper() or "?"
        counts[str(val)] = counts.get(str(val), 0) + 1
    return dict(sorted(counts.items(), key=lambda x: -x[1]))



def _stream_safe(query):
    """Stream a Firestore query, skipping docs that raise ValueError (e.g. _rowy_)."""
    gen = query.stream()
    while True:
        try:
            doc = next(gen)
        except StopIteration:
            break
        except (ValueError, AttributeError):
            continue
        try:
            yield doc.to_dict() or {}, doc
        except Exception:
            continue


def _stream_partitioned(query, partitions: int = 16, workers: int = 16):
    """Stream a collection/query using parallel partitions for speed.

    Falls back to single sequential stream if partitioning is unavailable.
    Returns an iterable of (dict, doc) tuples — same as _stream_safe.
    Uses the same pattern as load_leads_from_firebase in lead_agent.py.
    """
    from concurrent.futures import ThreadPoolExecutor

    try:
        partition_queries = [p.query() for p in query.get_partitions(partitions)]
    except Exception:
        partition_queries = [query]

    def _fetch(q):
        results = []
        for d, doc in _stream_safe(q):
            results.append((d, doc))
        return results

    all_results = []
    with ThreadPoolExecutor(max_workers=min(workers, len(partition_queries))) as pool:
        for batch in pool.map(_fetch, partition_queries):
            all_results.extend(batch)
    return all_results


def _parallel_stream(queries: list, workers: int = 2) -> list:
    """Run multiple Firestore streams in parallel, return list of (results_per_query).
    Each query is a tuple (label, query_obj, process_fn) where process_fn(d, doc) -> partial_result.
    Returns list of results in same order as queries.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _run(item):
        label, q, process_fn = item
        print(f"  [parallel] streaming {label}…", flush=True)
        result = process_fn(q)
        return result

    results = [None] * len(queries)
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_run, q): i for i, q in enumerate(queries)}
        for fut in as_completed(futures):
            results[futures[fut]] = fut.result()
    return results


def collection_overview(stats_collection: str = "statistics") -> dict:
    """Count documents in all major collections with country + dimension breakdowns."""
    import firebase_admin
    from firebase_admin import firestore

    cred = _get_credentials()
    if cred is None:
        print("  [overview] no credentials — skipping")
        return {}

    with _local_fb_lock:
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
    db = firestore.client()

    now_ts = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")
    results = {"generated_at": now_ts, "collections": {}}

    print("\n--- Collection Overview ---")
    print(f"  {'Collection':<30} {'Docs':>7}  Breakdowns")
    print("  " + "─" * 70)

    # ── leads (legacy) ──────────────────────────────────────────────────────
    def _leads_stats():
        count = sum(1 for _ in db.collection("leads").select([]).stream())
        by_country  = _count_by_field(db, "leads", "country")
        by_priority = _count_by_field(db, "leads", "priority")
        top_c = "  ".join(f"{k}:{v}" for k, v in list(by_country.items())[:5])
        top_p = "  ".join(f"{k}:{v}" for k, v in by_priority.items())
        print(f"  {'Leads (legacy)':<30} {count:>7}  country: {top_c}")
        print(f"  {'':30}         priority: {top_p}")
        return {"count": count, "by_country": by_country, "by_priority": by_priority}

    # ── leads_excluded ──────────────────────────────────────────────────────
    def _leads_excluded_stats():
        by_country: dict[str, int] = {}
        by_reason:  dict[str, int] = {}
        count = 0
        for d, _ in _stream_partitioned(
                db.collection("leads_excluded").select(["country","reason"])):
            count += 1
            c = (d.get("country") or "?").strip().upper() or "?"
            r = (d.get("reason")  or "?").strip()         or "?"
            by_country[c] = by_country.get(c, 0) + 1
            by_reason[r]  = by_reason.get(r, 0)  + 1
        by_country = dict(sorted(by_country.items(), key=lambda x: -x[1]))
        by_reason  = dict(sorted(by_reason.items(),  key=lambda x: -x[1]))
        top_c = "  ".join(f"{k}:{v}" for k, v in list(by_country.items())[:5])
        top_r = "  ".join(f"{k}:{v}" for k, v in list(by_reason.items())[:3])
        print(f"  {'Leads Excluded':<30} {count:>7}  country: {top_c or '—'}")
        print(f"  {'':<30}         reason: {top_r or '—'}")
        return {"count": count, "by_country": by_country, "by_reason": by_reason}

    # ── site_leads ──────────────────────────────────────────────────────────
    def _site_leads_stats():
        count = sum(1 for _ in db.collection("site_leads").select([]).stream())
        by_country    = _count_by_field(db, "site_leads", "country")
        by_ai_country = _count_by_field(db, "site_leads", "ai_country")
        by_ai_sector  = _count_by_field(db, "site_leads", "ai_sector")

        # Page count breakdown — read page_count field
        page_buckets: dict[str, int] = {
            "micro   (1–50)":       0,
            "small   (51–500)":     0,
            "medium  (501–3k)":     0,
            "large   (3k–10k)":     0,
            "huge    (10k–100k)":   0,
            "ultra   (100k+)":      0,
            "unknown (0/None)":     0,
        }
        for doc in db.collection("site_leads").select(["page_count"]).stream():
            pc = (doc.to_dict() or {}).get("page_count") or 0
            try:
                pc = int(pc)
            except (TypeError, ValueError):
                pc = 0
            if pc == 0:
                page_buckets["unknown (0/None)"] += 1
            elif pc <= 50:
                page_buckets["micro   (1–50)"] += 1
            elif pc <= 500:
                page_buckets["small   (51–500)"] += 1
            elif pc <= 3000:
                page_buckets["medium  (501–3k)"] += 1
            elif pc <= 10000:
                page_buckets["large   (3k–10k)"] += 1
            elif pc <= 100000:
                page_buckets["huge    (10k–100k)"] += 1
            else:
                page_buckets["ultra   (100k+)"] += 1

        top_c   = "  ".join(f"{k}:{v}" for k, v in list(by_country.items())[:5])
        top_aic = "  ".join(f"{k}:{v}" for k, v in list(by_ai_country.items())[:5])
        top_s   = "  ".join(f"{k}:{v}" for k, v in list(by_ai_sector.items())[:4])
        top_pg  = "  ".join(f"{k}:{v}" for k, v in page_buckets.items() if v > 0)
        print(f"  {'Site Leads':<30} {count:>7}  country: {top_c}")
        print(f"  {'':30}         ai_country: {top_aic}")
        print(f"  {'':30}         ai_sector: {top_s}")
        print(f"  {'':30}         page_size: {top_pg}")
        return {"count": count, "by_country": by_country,
                "by_ai_country": by_ai_country, "by_ai_sector": by_ai_sector,
                "by_page_size": page_buckets}

    # ── sites_excluded ──────────────────────────────────────────────────────
    def _sites_excluded_stats():
        by_country: dict[str, int] = {}
        by_reason:  dict[str, int] = {}
        count = 0
        for d, _ in _stream_partitioned(
                db.collection("sites_excluded").select(["country","reason"])):
            count += 1
            c = (d.get("country") or "?").strip().upper() or "?"
            r = (d.get("reason")  or "?").strip()         or "?"
            by_country[c] = by_country.get(c, 0) + 1
            by_reason[r]  = by_reason.get(r, 0)  + 1
        by_country = dict(sorted(by_country.items(), key=lambda x: -x[1]))
        by_reason  = dict(sorted(by_reason.items(),  key=lambda x: -x[1]))
        top_c = "  ".join(f"{k}:{v}" for k, v in list(by_country.items())[:5])
        top_r = "  ".join(f"{k}:{v}" for k, v in list(by_reason.items())[:3])
        print(f"  {'Sites Excluded':<30} {count:>7}  country: {top_c or '—'}")
        print(f"  {'':<30}         reason: {top_r or '—'}")
        return {"count": count, "by_country": by_country, "by_reason": by_reason}

    runners = [
        ("leads",          _leads_stats),
        ("leads_excluded", _leads_excluded_stats),
        ("site_leads",     _site_leads_stats),
        ("sites_excluded", _sites_excluded_stats),
    ]

    for col_name, fn in runners:
        try:
            results["collections"][col_name] = fn()
        except Exception as exc:
            print(f"  {col_name:<30} ERROR: {exc}")
            results["collections"][col_name] = {"error": str(exc)}

    # ── Exclusion rates ─────────────────────────────────────────────────────
    print("\n  EXCLUSION RATES")
    cols = results["collections"]
    for stored, excl, label in [
        ("leads",      "leads_excluded",  "Lead pipeline"),
        ("site_leads", "sites_excluded",  "Site pipeline"),
    ]:
        n_stored = cols.get(stored, {}).get("count", 0)
        n_excl   = cols.get(excl,   {}).get("count", 0)
        total    = n_stored + n_excl
        excl_pct = int(100 * n_excl / total) if total else 0
        print(f"  {label:<20} stored={n_stored:>6}  excluded={n_excl:>6}  "
              f"rejection rate={excl_pct}%")
        results[f"{excl}_rate"] = excl_pct

    # Write to Firestore
    try:
        # Strip large subcollection detail before writing (keep counts only)
        slim = {"generated_at": now_ts, "collections": {}}
        for k, v in results["collections"].items():
            slim["collections"][k] = {kk: vv for kk, vv in v.items()
                                       if not isinstance(vv, list) or len(str(vv)) < 2000}
        db.collection(stats_collection).document("collection-overview").set(slim, merge=True)
        print(f"\n  Written → {stats_collection}/collection-overview")
    except Exception as exc:
        print(f"  [overview] Firestore write error: {exc}")

    return results


def export_overview_to_excel(results: dict, outdir: str | None = None) -> str:
    """Write collection overview to Excel with country/dimension breakdowns."""
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [overview] pandas/openpyxl not installed — skipping Excel")
        return ""

    outdir = Path(outdir) if outdir else Path(__file__).parent.parent / "output"
    outdir.mkdir(parents=True, exist_ok=True)
    out_path = outdir / "collection_overview.xlsx"

    rows = []
    for col_name, entry in results.get("collections", {}).items():
        if "error" in entry:
            rows.append({"Collection": col_name, "Dimension": "ERROR",
                         "Key": entry["error"], "Count": 0})
            continue
        rows.append({"Collection": col_name, "Dimension": "total",
                     "Key": "(all)", "Count": entry.get("count", 0)})
        for dim in ("by_country", "by_ai_country", "by_ai_sector", "by_priority", "by_reason", "by_page_size"):
            for k, v in (entry.get(dim) or {}).items():
                rows.append({"Collection": col_name, "Dimension": dim.replace("by_", ""),
                              "Key": k, "Count": v})

    df = pd.DataFrame(rows)
    df.to_excel(out_path, index=False, sheet_name="Overview")
    print(f"  [overview] Excel written → {out_path}")
    return str(out_path)



# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# A. Site pipeline enrichment funnel
# ---------------------------------------------------------------------------

def site_pipeline_enrichment_funnel(stats_collection: str = "statistics") -> dict:
    """Report enrichment completion for site_leads and site_contacts — runs both streams in parallel."""
    from concurrent.futures import ThreadPoolExecutor

    db = _init_firebase()

    def _scan_leads():
        r = {"total":0,"ai":0,"loc":0,"both":0}
        for d, _ in _stream_partitioned(
                db.collection("site_leads")
                  .select(["ai_classified_at","location_enriched_at"])):
            r["total"] += 1
            ai  = bool((d.get("ai_classified_at")     or "").strip())
            loc = bool((d.get("location_enriched_at") or "").strip())
            if ai:       r["ai"]   += 1
            if loc:      r["loc"]  += 1
            if ai and loc: r["both"] += 1
        return r

    def _scan_contacts():
        r = {"total":0,"brave":0,"checked":0,"both":0,"no_name":0,"no_email":0}
        for d, _ in _stream_partitioned(
                db.collection_group("site_contacts")
                  .select(["brave_enriched_at","email_checked_at","name","email"])):
            r["total"] += 1
            brave = bool((d.get("brave_enriched_at") or "").strip())
            chkd  = bool((d.get("email_checked_at")  or "").strip())
            if brave:         r["brave"]   += 1
            if chkd:          r["checked"] += 1
            if brave and chkd: r["both"]   += 1
            if not (d.get("name")  or "").strip(): r["no_name"]  += 1
            if not (d.get("email") or "").strip(): r["no_email"] += 1
        return r

    print("  [funnel] streaming site_leads + site_contacts in parallel…")
    with ThreadPoolExecutor(max_workers=2) as pool:
        f_leads    = pool.submit(_scan_leads)
        f_contacts = pool.submit(_scan_contacts)
        L = f_leads.result()
        C = f_contacts.result()

    leads_total            = L["total"]
    leads_ai_classified    = L["ai"]
    leads_location_enriched= L["loc"]
    leads_both             = L["both"]
    contacts_total         = C["total"]
    contacts_brave         = C["brave"]
    contacts_email_checked = C["checked"]
    contacts_both          = C["both"]
    contacts_no_name       = C["no_name"]
    contacts_no_email      = C["no_email"]

    def pct(n, total):
        return f"{n:>6}  ({100*n//total if total else 0}%)"

    print(f"\n  SITE LEADS ({leads_total} total)")
    print(f"    AI classified:       {pct(leads_ai_classified, leads_total)}")
    print(f"    Location enriched:   {pct(leads_location_enriched, leads_total)}")
    print(f"    Fully enriched:      {pct(leads_both, leads_total)}")
    print(f"\n  SITE CONTACTS ({contacts_total} total)")
    print(f"    Brave enriched:      {pct(contacts_brave, contacts_total)}")
    print(f"    Email checked:       {pct(contacts_email_checked, contacts_total)}")
    print(f"    Fully ready:         {pct(contacts_both, contacts_total)}")
    print(f"    Missing name:        {pct(contacts_no_name, contacts_total)}")
    print(f"    Missing email:       {pct(contacts_no_email, contacts_total)}")

    result = {
        "leads_total": leads_total,
        "leads_ai_classified": leads_ai_classified,
        "leads_location_enriched": leads_location_enriched,
        "leads_both_enriched": leads_both,
        "contacts_total": contacts_total,
        "contacts_brave_enriched": contacts_brave,
        "contacts_email_checked": contacts_email_checked,
        "contacts_fully_ready": contacts_both,
        "contacts_no_name": contacts_no_name,
        "contacts_no_email": contacts_no_email,
    }
    db.collection(stats_collection).document("site-enrichment-funnel").set(
        {**result, "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")}, merge=True)
    return result


def lead_pipeline_enrichment_funnel(leads_collection=None, stats_collection="statistics"):
    from concurrent.futures import ThreadPoolExecutor
    db = _init_firebase()
    col_name = leads_collection or cfg.FIRESTORE_COLLECTION

    def _scan_leads():
        r = {"total":0,"ai":0,"no_email":0}
        for d, _ in _stream_partitioned(
                db.collection(col_name).select(["ai_classified_at","emails"])):
            r["total"] += 1
            if (d.get("ai_classified_at") or "").strip(): r["ai"] += 1
            if not (d.get("emails") or "").strip():       r["no_email"] += 1
        return r

    def _scan_contacts():
        r = {"total":0,"social":0,"checked":0,"both":0}
        for d, _ in _stream_partitioned(
                db.collection_group("contacts")
                  .select(["social_enriched_at","email_checked_at"])):
            r["total"] += 1
            s = bool((d.get("social_enriched_at") or "").strip())
            c = bool((d.get("email_checked_at")   or "").strip())
            if s: r["social"]  += 1
            if c: r["checked"] += 1
            if s and c: r["both"] += 1
        return r

    print(f"  [funnel] streaming {col_name} + contacts in parallel…")
    with ThreadPoolExecutor(max_workers=2) as pool:
        f_leads    = pool.submit(_scan_leads)
        f_contacts = pool.submit(_scan_contacts)
        L = f_leads.result()
        C = f_contacts.result()

    leads_total = L["total"]; leads_ai = L["ai"]; leads_no_email = L["no_email"]
    ct = C["total"]; cs = C["social"]; cc = C["checked"]; cb = C["both"]

    def pct(n, t): return f"{n:>6}  ({100*n//t if t else 0}%)"
    print(f"\n  LEADS ({leads_total}) | AI classified: {pct(leads_ai,leads_total)} | No email: {pct(leads_no_email,leads_total)}")
    print(f"  LEAD CONTACTS ({ct}) | Social: {pct(cs,ct)} | Email-checked: {pct(cc,ct)} | Both: {pct(cb,ct)}")
    result = {"leads_total":leads_total,"leads_ai_classified":leads_ai,"leads_no_email":leads_no_email,
              "contacts_total":ct,"contacts_social":cs,"contacts_email_checked":cc,"contacts_both":cb}
    db.collection(stats_collection).document("lead-enrichment-funnel").set(
        {**result,"generated_at":datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")}, merge=True)
    return result


def data_quality_report(stats_collection="statistics"):
    from concurrent.futures import ThreadPoolExecutor
    from functions.utils import email_matches_name
    db = _init_firebase()

    def _scan_leads():
        r = {"total":0,"no_sitemap":0,"zero_pages":0,"not_classified":0}
        for d, _ in _stream_partitioned(
                db.collection("site_leads")
                  .select(["sitemap_type","page_count","ai_classified_at"])):
            r["total"] += 1
            if (d.get("sitemap_type") or "") in ("none",""): r["no_sitemap"]      += 1
            if int(d.get("page_count") or 0) == 0:           r["zero_pages"]      += 1
            if not (d.get("ai_classified_at") or "").strip(): r["not_classified"]  += 1
        return r

    def _scan_contacts():
        r = {"total":0,"no_name":0,"mismatch":0}
        for d, _ in _stream_partitioned(
                db.collection_group("site_contacts").select(["email","name"])):
            r["total"] += 1
            name  = (d.get("name")  or "").strip()
            email = (d.get("email") or "").strip()
            if not name: r["no_name"] += 1
            elif not email_matches_name(email, name): r["mismatch"] += 1
        return r

    print("  [quality] scanning site_leads + site_contacts in parallel…")
    with ThreadPoolExecutor(max_workers=2) as pool:
        f_leads    = pool.submit(_scan_leads)
        f_contacts = pool.submit(_scan_contacts)
        L = f_leads.result()
        C = f_contacts.result()

    lt = L["total"]; lns = L["no_sitemap"]; lzp = L["zero_pages"]; lna = L["not_classified"]
    ct = C["total"]; cnn = C["no_name"]; cnm = C["mismatch"]

    def pct(n, t): return f"{n:>6}  ({100*n//t if t else 0}%)"
    print(f"\n  SITE LEADS quality ({lt}) | No sitemap: {pct(lns,lt)} | Zero pages: {pct(lzp,lt)} | Not classified: {pct(lna,lt)}")
    print(f"  SITE CONTACTS quality ({ct}) | No name: {pct(cnn,ct)} | Name mismatch: {pct(cnm,ct)}")
    result = {"leads_total":lt,"leads_no_sitemap":lns,"leads_zero_pages":lzp,"leads_not_classified":lna,
              "contacts_total":ct,"contacts_no_name":cnn,"contacts_name_mismatch":cnm}
    db.collection(stats_collection).document("data-quality").set(
        {**result,"generated_at":datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")}, merge=True)
    return result


def email_contacts_funnel(stats_collection="statistics"):
    db = _init_firebase()
    print("  [funnel] streaming email_contacts...")
    total = 0
    by_status = {}; by_pipeline = {"site_only":0,"leads_only":0,"both":0}
    by_email_type = {}; by_priority = {}; by_country = {}
    gen = db.collection("email_contacts").stream()
    while True:
        try: doc = next(gen)
        except StopIteration: break
        except (ValueError, AttributeError): continue
        try: d = doc.to_dict() or {}
        except Exception: continue
        total += 1
        st = (d.get("status") or "pending").strip()
        by_status[st] = by_status.get(st, 0) + 1
        site = bool(d.get("mark_site_leads")); leads = bool(d.get("mark_leads"))
        if site and leads: by_pipeline["both"] += 1
        elif site: by_pipeline["site_only"] += 1
        elif leads: by_pipeline["leads_only"] += 1
        et = (d.get("email_type") or "?").strip()
        by_email_type[et] = by_email_type.get(et, 0) + 1
        pr = str(d.get("outreach_priority") or "?")
        by_priority[pr] = by_priority.get(pr, 0) + 1
        cc = (d.get("country") or "?").strip().upper() or "?"
        by_country[cc] = by_country.get(cc, 0) + 1

    def pct(n): return f"{n:>6}  ({100*n//total if total else 0}%)"
    print(f"\n  EMAIL CONTACTS ({total} total)")
    print("  Status: " + "  ".join(f"{k}={v}" for k,v in sorted(by_status.items(),key=lambda x:-x[1])))
    print("  Pipeline: " + "  ".join(f"{k}={v}" for k,v in by_pipeline.items()))
    print("  Email type: " + "  ".join(f"{k}={v}" for k,v in sorted(by_email_type.items(),key=lambda x:-x[1])))
    print("  Priority: " + "  ".join(f"p{k}={v}" for k,v in sorted(by_priority.items())))
    result = {"total":total,"by_status":by_status,"by_pipeline":by_pipeline,
              "by_email_type":by_email_type,"by_priority":by_priority,
              "top_countries":dict(sorted(by_country.items(),key=lambda x:-x[1])[:20])}
    db.collection(stats_collection).document("email-contacts-funnel").set(
        {**result,"generated_at":datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")}, merge=True)
    return result


def pipeline_coverage(stats_collection="statistics"):
    db = _init_firebase()
    print("  [coverage] streaming email_contacts...")
    total = site_only = leads_only = both = 0
    by_country = {}
    gen = db.collection("email_contacts").select(["mark_site_leads","mark_leads","country"]).stream()
    while True:
        try: doc = next(gen)
        except StopIteration: break
        except (ValueError, AttributeError): continue
        try: d = doc.to_dict() or {}
        except Exception: continue
        total += 1
        site = bool(d.get("mark_site_leads")); leads = bool(d.get("mark_leads"))
        cc = (d.get("country") or "?").strip().upper() or "?"
        if cc not in by_country: by_country[cc] = {"total":0,"site":0,"leads":0,"both":0}
        by_country[cc]["total"] += 1
        if site and leads: both += 1; by_country[cc]["both"] += 1
        elif site: site_only += 1; by_country[cc]["site"] += 1
        elif leads: leads_only += 1; by_country[cc]["leads"] += 1

    def pct(n,t): return f"({100*n//t if t else 0}%)"
    print(f"\n  PIPELINE COVERAGE ({total}) | Site only:{site_only} {pct(site_only,total)} | Leads only:{leads_only} {pct(leads_only,total)} | Both:{both} {pct(both,total)}")
    for cc, c in sorted(by_country.items(), key=lambda x:-x[1]["total"])[:10]:
        print(f"    {cc:<5} total={c['total']:>5}  site={c['site']:>5}  leads={c['leads']:>5}  both={c['both']:>5}")
    result = {"total":total,"site_only":site_only,"leads_only":leads_only,"both_pipelines":both,"by_country":by_country}
    db.collection(stats_collection).document("pipeline-coverage").set(
        {**result,"generated_at":datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z")}, merge=True)
    return result


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def export_full_statistics(all_results: dict, outdir=None) -> str:
    """Write all statistics to a single Excel workbook with one sheet per aggregation.

    Parameters
    ----------
    all_results : dict   Keys: priority, reasons, overview, site_funnel, lead_funnel,
                               quality, email_funnel, coverage
    outdir      : path   Output directory (default: <root>/output)

    Returns path to written file.
    """
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    outdir = Path(outdir) if outdir else Path(__file__).parent.parent / "output"
    outdir.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    out_path = outdir / f"statistics_{date_str}.xlsx"

    HDR_FILL = PatternFill("solid", start_color="1F497D")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DATA_FONT = Font(name="Arial", size=10)

    def _style_sheet(ws):
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = DATA_FONT
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 55)

    def _df_to_sheet(writer, df, sheet_name):
        if df is None or df.empty: return
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        _style_sheet(writer.book[sheet_name])

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        # ── Sheet 1: Priority × Country ──────────────────────────────────
        r = all_results.get("priority")
        if r:
            rows = []
            for data in r.get("countries", {}).values():
                for prio, counts in data.get("by_priority", {}).items():
                    rows.append({"country": data["country"],
                                 "country_name": data["country_name"],
                                 "priority": prio,
                                 "leads": counts["leads"],
                                 "contacts": counts["contacts"]})
            _df_to_sheet(writer, pd.DataFrame(rows), "Priority × Country")

        # ── Sheet 2: Reasons ─────────────────────────────────────────────
        r = all_results.get("reasons")
        if r:
            rows = [{"reason": k, "count": v}
                    for k, v in sorted((r.get("reasons") or {}).items(), key=lambda x: -x[1])]
            _df_to_sheet(writer, pd.DataFrame(rows), "Reasons")

        # ── Sheet 3: Collection Overview ─────────────────────────────────
        r = all_results.get("overview")
        if r:
            rows = []
            for col_name, data in (r.get("collections") or {}).items():
                rows.append({"collection": col_name,
                             "count": data.get("count", 0),
                             "top_countries": ", ".join(
                                 f"{k}:{v}" for k,v in list(
                                     (data.get("by_country") or data.get("by_ai_country") or {})
                                     .items())[:5])})
            rows.append({"collection": "--- Exclusion rates ---", "count": "",
                         "top_countries": ""})
            rows.append({"collection": "Lead rejection rate",
                         "count": f"{r.get('leads_excluded_rate',0)}%", "top_countries": ""})
            rows.append({"collection": "Site rejection rate",
                         "count": f"{r.get('sites_excluded_rate',0)}%", "top_countries": ""})
            _df_to_sheet(writer, pd.DataFrame(rows), "Collection Overview")

        # ── Sheet 4: Site Enrichment Funnel ──────────────────────────────
        r = all_results.get("site_funnel")
        if r:
            lt = r.get("leads_total",0); ct = r.get("contacts_total",0)
            def pct(n,t): return f"{n} ({100*n//t if t else 0}%)"
            rows = [
                {"metric": "Site Leads — total",            "value": lt},
                {"metric": "  AI classified",               "value": pct(r.get("leads_ai_classified",0),lt)},
                {"metric": "  Location enriched",           "value": pct(r.get("leads_location_enriched",0),lt)},
                {"metric": "  Fully enriched (both)",       "value": pct(r.get("leads_both_enriched",0),lt)},
                {"metric": "", "value": ""},
                {"metric": "Site Contacts — total",         "value": ct},
                {"metric": "  Brave enriched",              "value": pct(r.get("contacts_brave_enriched",0),ct)},
                {"metric": "  Email checked",               "value": pct(r.get("contacts_email_checked",0),ct)},
                {"metric": "  Fully ready (both)",          "value": pct(r.get("contacts_fully_ready",0),ct)},
                {"metric": "  Missing name",                "value": pct(r.get("contacts_no_name",0),ct)},
                {"metric": "  Missing email",               "value": pct(r.get("contacts_no_email",0),ct)},
            ]
            _df_to_sheet(writer, pd.DataFrame(rows), "Site Enrichment")

        # ── Sheet 5: Lead Enrichment Funnel ──────────────────────────────
        r = all_results.get("lead_funnel")
        if r:
            lt = r.get("leads_total",0); ct = r.get("contacts_total",0)
            def pct(n,t): return f"{n} ({100*n//t if t else 0}%)"
            rows = [
                {"metric": "Leads — total",                "value": lt},
                {"metric": "  AI classified",              "value": pct(r.get("leads_ai_classified",0),lt)},
                {"metric": "  No email/contacts",          "value": pct(r.get("leads_no_email",0),lt)},
                {"metric": "", "value": ""},
                {"metric": "Lead Contacts — total",        "value": ct},
                {"metric": "  Social enriched",            "value": pct(r.get("contacts_social_enriched",0),ct)},
                {"metric": "  Email checked",              "value": pct(r.get("contacts_email_checked",0),ct)},
                {"metric": "  Fully ready (both)",         "value": pct(r.get("contacts_fully_ready",0),ct)},
            ]
            _df_to_sheet(writer, pd.DataFrame(rows), "Lead Enrichment")

        # ── Sheet 6: Data Quality ─────────────────────────────────────────
        r = all_results.get("quality")
        if r:
            lt = r.get("leads_total",0); ct = r.get("contacts_total",0)
            def pct(n,t): return f"{n} ({100*n//t if t else 0}%)"
            rows = [
                {"metric": "Site Leads — total",           "value": lt},
                {"metric": "  No sitemap found",           "value": pct(r.get("leads_no_sitemap",0),lt)},
                {"metric": "  Zero page count",            "value": pct(r.get("leads_zero_pages",0),lt)},
                {"metric": "  Not AI classified",          "value": pct(r.get("leads_not_classified",0),lt)},
                {"metric": "", "value": ""},
                {"metric": "Site Contacts — total",        "value": ct},
                {"metric": "  No name scraped",            "value": pct(r.get("contacts_no_name",0),ct)},
                {"metric": "  Name/email mismatch",        "value": pct(r.get("contacts_name_mismatch",0),ct)},
            ]
            _df_to_sheet(writer, pd.DataFrame(rows), "Data Quality")

        # ── Sheet 7: email_contacts Funnel ────────────────────────────────
        r = all_results.get("email_funnel")
        if r:
            total = r.get("total", 0)
            def pct(n): return f"{n} ({100*n//total if total else 0}%)"
            rows = [{"section":"Total","key":"","value":total}]
            for k,v in sorted((r.get("by_status") or {}).items(), key=lambda x:-x[1]):
                rows.append({"section":"Status","key":k,"value":pct(v)})
            for k,v in (r.get("by_pipeline") or {}).items():
                rows.append({"section":"Pipeline","key":k,"value":pct(v)})
            for k,v in sorted((r.get("by_email_type") or {}).items(), key=lambda x:-x[1]):
                rows.append({"section":"Email Type","key":k,"value":pct(v)})
            for k,v in sorted((r.get("by_priority") or {}).items()):
                rows.append({"section":"Priority","key":f"priority {k}","value":pct(v)})
            _df_to_sheet(writer, pd.DataFrame(rows), "email_contacts Funnel")

        # ── Sheet 8: Pipeline Coverage ────────────────────────────────────
        r = all_results.get("coverage")
        if r:
            total = r.get("total", 0)
            def pct(n): return f"{n} ({100*n//total if total else 0}%)"
            rows = [
                {"segment": "Total contacts",    "count": total,                         "pct": ""},
                {"segment": "Site only",         "count": r.get("site_only",0),          "pct": pct(r.get("site_only",0))},
                {"segment": "Leads only",        "count": r.get("leads_only",0),         "pct": pct(r.get("leads_only",0))},
                {"segment": "Both pipelines",    "count": r.get("both_pipelines",0),     "pct": pct(r.get("both_pipelines",0))},
            ]
            for cc, c in sorted((r.get("by_country") or {}).items(), key=lambda x:-x[1].get("total",0))[:15]:
                rows.append({"segment": f"  {cc}", "count": c["total"],
                             "pct": f"site={c['site']} leads={c['leads']} both={c['both']}"})
            _df_to_sheet(writer, pd.DataFrame(rows), "Pipeline Coverage")

    print(f"  [stats] Combined Excel → {out_path}")
    return str(out_path)


def _write_summary_doc(all_results: dict, stats_collection: str = "statistics") -> None:
    """Write all statistics to a single Firestore document statistics/summary-YYYY-MM-DD."""
    db = _init_firebase()
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    doc_id   = f"summary-{date_str}"
    doc = {
        "generated_at":   datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00","Z"),
        "date":           date_str,
        "priority":       all_results.get("priority",  {}).get("head", {}),
        "reasons":        all_results.get("reasons",   {}),
        "overview":       all_results.get("overview",  {}),
        "site_funnel":    all_results.get("site_funnel",  {}),
        "lead_funnel":    all_results.get("lead_funnel",  {}),
        "data_quality":   all_results.get("quality",      {}),
        "email_funnel":   all_results.get("email_funnel", {}),
        "coverage":       all_results.get("coverage",     {}),
    }
    db.collection(stats_collection).document(doc_id).set(doc, merge=True)
    print(f"  [stats] Summary doc → {stats_collection}/{doc_id}")


def main():
    from dotenv import load_dotenv
    load_dotenv()
    import argparse
    parser = argparse.ArgumentParser(
        description="Aggregate statistics across both pipelines.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("--leads-collection", default=None)
    parser.add_argument("--stats-collection", default="statistics")
    parser.add_argument("--output", default=None)
    parser.add_argument("--no-excel", action="store_true")
    parser.add_argument("--no-writeback", action="store_true")
    parser.add_argument("--no-overview", action="store_true")
    parser.add_argument("--only", default=None,
        choices=["priority","reasons","overview","site-funnel","lead-funnel",
                 "quality","email-funnel","coverage"])
    args = parser.parse_args()

    run_all = args.only is None
    all_results: dict = {}

    if run_all or args.only == "priority":
        print("\n--- Priority x Country ---")
        r = summarise_country_pr_priority(args.leads_collection, args.stats_collection)
        _print_summary(r); all_results["priority"] = r
    if run_all or args.only == "reasons":
        print("\n--- Reasons Count ---")
        r = summarise_reasons_count(args.leads_collection, args.stats_collection, writeback=not args.no_writeback)
        all_results["reasons"] = r
    if (run_all or args.only == "overview") and not args.no_overview:
        print("\n--- Collection Overview ---")
        r = collection_overview(args.stats_collection); all_results["overview"] = r
    if run_all or args.only == "site-funnel":
        print("\n--- Site Pipeline Enrichment Funnel ---")
        all_results["site_funnel"] = site_pipeline_enrichment_funnel(args.stats_collection)
    if run_all or args.only == "lead-funnel":
        print("\n--- Lead Pipeline Enrichment Funnel ---")
        all_results["lead_funnel"] = lead_pipeline_enrichment_funnel(args.leads_collection, args.stats_collection)
    if run_all or args.only == "quality":
        print("\n--- Data Quality Report ---")
        all_results["quality"] = data_quality_report(args.stats_collection)
    if run_all or args.only == "email-funnel":
        print("\n--- email_contacts Outreach Funnel ---")
        all_results["email_funnel"] = email_contacts_funnel(args.stats_collection)
    if run_all or args.only == "coverage":
        print("\n--- Pipeline Cross-Coverage ---")
        all_results["coverage"] = pipeline_coverage(args.stats_collection)

    # ── Combined outputs ──────────────────────────────────────────────────
    if all_results:
        if not args.no_excel:
            export_full_statistics(all_results, outdir=args.output)
        _write_summary_doc(all_results, stats_collection=args.stats_collection)


if __name__ == "__main__":
    main()
