"""Shared Excel building utilities for all contact export scripts."""
from __future__ import annotations
from pathlib import Path

TIER_COLORS = {1: 'C00000', 2: 'FF0000', 3: 'FF9900', 4: 'FFFF00', 5: 'D9D9D9'}
TIER_TEXT   = {1: 'FFFFFF', 2: 'FFFFFF', 3: '000000', 4: '000000', 5: '000000'}


def write_contacts_sheet(
    ws,
    rows:      list[dict],
    cols:      list[tuple],      # [(header, field_key, width), ...]
    sort_key,                    # callable(row) -> sort tuple
    wrap_keys: set[str] | None = None,
) -> None:
    """Write the Contacts sheet into an existing openpyxl worksheet.

    Handles: header row, tier colouring, bool→YES/'', list→csv,
    wrap_text for specified keys, auto-filter, freeze panes.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wrap_keys = wrap_keys or set()

    HDR_FILL  = PatternFill('solid', start_color='1F497D')
    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    DATA_FONT = Font(name='Arial', size=10)
    WRAP      = Alignment(wrap_text=True, vertical='top')
    NOWRAP    = Alignment(vertical='top')
    THIN      = Side(style='thin', color='CCCCCC')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # ── Header row ────────────────────────────────────────────────────────
    for ci, (hdr, _, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    # ── Data rows ─────────────────────────────────────────────────────────
    sorted_rows = sorted(rows, key=sort_key)

    for ri, row in enumerate(sorted_rows, 2):
        tier      = int(row.get('tier') or 5)
        bg        = TIER_COLORS.get(tier, 'FFFFFF')
        fg        = TIER_TEXT.get(tier,  '000000')
        tier_fill = PatternFill('solid', start_color=bg)
        tier_font = Font(name='Arial', size=10, color=fg, bold=tier <= 2)

        for ci, (_, key, _) in enumerate(cols, 1):
            val = row.get(key, '')
            # Normalise all non-scalar types before writing to Excel
            if val is None:
                val = ''
            elif isinstance(val, bool):
                val = 'YES' if val else ''
            elif isinstance(val, list):
                val = ', '.join(str(v) for v in val if v not in (None, ''))
            elif isinstance(val, dict):
                val = '; '.join(f'{k}={v}' for k, v in val.items() if v not in (None, ''))
            elif not isinstance(val, (str, int, float)):
                val = str(val)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci == 1:
                cell.fill      = tier_fill
                cell.font      = tier_font
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif key in wrap_keys:
                cell.font      = DATA_FONT
                cell.alignment = WRAP
            else:
                cell.font      = DATA_FONT
                cell.alignment = NOWRAP

        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'


def make_header_cell(ws, row: int, col: int, value: str):
    """Write a bold dark-blue header cell on a summary sheet."""
    from openpyxl.styles import Font, PatternFill
    cell = ws.cell(row, col, value)
    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    cell.fill = PatternFill('solid', start_color='1F497D')
    return cell


def save_workbook(wb, out_path: Path, label: str = '') -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    tag = f'[{label}] ' if label else ''
    print(f'{tag}Saved → {out_path}', flush=True)
